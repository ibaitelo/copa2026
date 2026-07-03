#!/usr/bin/env python3
"""
copa2026_v13.py — Monte Carlo + Simulação Determinística Completa R32→Final

Usa xgboost_v12.pkl (já treinado e calibrado, bias global=-0.003).
Aplica winsorização leve (p1/p99) em features com outliers extremos.
Retreina v13 com dataset winsorizado.
Roda MC 10k e simula bracket completo com placares mais prováveis.

Output: outputs/copa2026_full_v13.xlsx
  Aba 1 "Grupos"              — Classificação real Copa 2026 fase de grupos
  Aba 2 "Bracket_Deterministico" — Chaveamento completo R32→Final (placar mais provável)
  Aba 3 "MC_Probabilidades"   — P(avançar) por round para todos os 48 times
  Aba 4 "MC_Campeon"          — Top-N campeões (MC 10k)
"""
from __future__ import annotations

import pickle
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import poisson as scipy_poisson
from xgboost import XGBRegressor
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

sys.path.insert(0, "src")
warnings.filterwarnings("ignore")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ── Paths ──────────────────────────────────────────────────────────────────────
RAW_CSV   = Path("data/raw/full_dataset_raw.csv")
MODEL_CSV = Path("data/processed/model_dataset.csv")
ELO_CSV   = Path("data/raw/elo_history.csv")
OUT_PKL   = Path("outputs/xgboost_v13.pkl")
OUT_XLSX  = Path("outputs/copa2026_full_v13.xlsx")

CUTOFF = pd.Timestamp("2026-06-28")
SEED   = 42
N_SIM  = 10_000

# ── Features ───────────────────────────────────────────────────────────────────
CONF_DUMMIES = ["conf_AFC","conf_CAF","conf_CONCACAF","conf_CONMEBOL","conf_OFC","conf_UEFA"]
RATING_FEATURES = [
    "elo_diff_decay","delta_rating_decay","gls_ponderado_decay",
    "gls_decay_vs_forte","gls_decay_vs_fraco","margem_gols_decay",
]
FEATURES = (["opp_saves_decay","shots_on_goal_decay","gols_sofridos_decay"]
            + RATING_FEATURES + CONF_DUMMIES)

XGB_PARAMS = dict(
    objective="count:poisson", n_estimators=300, max_depth=4,
    learning_rate=0.05, subsample=0.8, colsample_bytree=0.8,
    min_child_weight=3, random_state=42, n_jobs=-1, verbosity=0,
)

# ── Copa 2026 — grupos e chaveamento ──────────────────────────────────────────
GROUPS: dict[str, list[str]] = {
    "A": ["Mexico",        "South Korea",  "South Africa",        "Czech Republic"],
    "B": ["Canada",        "Switzerland",  "Qatar",               "Bosnia and Herzegovina"],
    "C": ["Brazil",        "Morocco",      "Haiti",               "Scotland"],
    "D": ["United States", "Paraguay",     "Australia",           "Turkey"],
    "E": ["Germany",       "Ivory Coast",  "Ecuador",             "Curacao"],
    "F": ["Netherlands",   "Sweden",       "Tunisia",             "Japan"],
    "G": ["Belgium",       "Iran",         "New Zealand",         "Egypt"],
    "H": ["Spain",         "Saudi Arabia", "Uruguay",             "Cape Verde"],
    "I": ["France",        "Senegal",      "Iraq",                "Norway"],
    "J": ["Argentina",     "Algeria",      "Austria",             "Jordan"],
    "K": ["Portugal",      "DR Congo",     "Uzbekistan",          "Colombia"],
    "L": ["England",       "Croatia",      "Ghana",               "Panama"],
}
ALL_COPA_TEAMS = [t for ts in GROUPS.values() for t in ts]

COPA_TEAM_CONF: dict[str, str] = {
    "Argentina":"CONMEBOL","Brazil":"CONMEBOL","Colombia":"CONMEBOL",
    "Ecuador":"CONMEBOL","Uruguay":"CONMEBOL","Paraguay":"CONMEBOL",
    "Germany":"UEFA","France":"UEFA","Spain":"UEFA","England":"UEFA",
    "Netherlands":"UEFA","Portugal":"UEFA","Belgium":"UEFA","Austria":"UEFA",
    "Switzerland":"UEFA","Croatia":"UEFA","Norway":"UEFA","Sweden":"UEFA",
    "Czech Republic":"UEFA","Turkey":"UEFA","Scotland":"UEFA",
    "Bosnia and Herzegovina":"UEFA",
    "Japan":"AFC","South Korea":"AFC","Iran":"AFC","Australia":"AFC",
    "Saudi Arabia":"AFC","Uzbekistan":"AFC","Jordan":"AFC","Iraq":"AFC",
    "Qatar":"AFC",
    "Morocco":"CAF","Senegal":"CAF","Egypt":"CAF","Ghana":"CAF",
    "Cape Verde":"CAF","DR Congo":"CAF","Ivory Coast":"CAF",
    "South Africa":"CAF","Algeria":"CAF","Tunisia":"CAF",
    "United States":"CONCACAF","Mexico":"CONCACAF","Canada":"CONCACAF",
    "Panama":"CONCACAF","Curacao":"CONCACAF","Haiti":"CONCACAF",
    "New Zealand":"OFC",
}

CONF_MEDIANS_SAVES = {
    "UEFA":2.9353, "CONMEBOL":2.7985, "AFC":2.4974,
    "CONCACAF":2.4799, "CAF":1.8144, "OFC":2.4974,
}
GLOBAL_MEDIAN_SAVES = 2.6

NAME_MAP = {
    "Czechia": "Czech Republic",
    "Czech Rep.": "Czech Republic",
    "Côte d'Ivoire": "Ivory Coast",
    "Cote d'Ivoire": "Ivory Coast",
    "United States": "United States",
    "USA": "United States",
}

# R32: (match_id, slot_t1, slot_t2, eligible_groups_for_3rd|None)
R32 = [
    (73,  "2A",  "2B",  None),
    (74,  "1E",  "3",   ["A","B","C","D","F"]),
    (75,  "1F",  "2C",  None),
    (76,  "1C",  "2F",  None),
    (77,  "1I",  "3",   ["C","D","F","G","H"]),
    (78,  "2E",  "2I",  None),
    (79,  "1A",  "3",   ["A","B","C","D","E"]),  # slot subject to assign_thirds
    (80,  "1L",  "2K",  None),
    (81,  "1D",  "2C",  None),  # 81 slot: 1D vs 2C already handled if 2C is gone; in real draw 81 is 1D vs 3rd
    (82,  "1G",  "2J",  None),
    (83,  "1K",  "2L",  None),
    (84,  "1H",  "2J",  None),
    (85,  "1B",  "2H",  None),
    (86,  "1J",  "2G",  None),
    (87,  "1K",  "3",   ["E","F","H","I","J"]),
    (88,  "1F",  "3",   ["G","H","I","J","K"]),
]

# Real Copa 2026 R32 matchups (from actual group stage results)
REAL_R32 = [
    (73,  "South Africa",          "Canada"),
    (74,  "Germany",               "Paraguay"),
    (75,  "Netherlands",           "Morocco"),
    (76,  "Brazil",                "Japan"),
    (77,  "France",                "Sweden"),
    (78,  "Ivory Coast",           "Norway"),
    (79,  "Mexico",                "Ecuador"),
    (80,  "England",               "DR Congo"),
    (81,  "United States",         "Bosnia and Herzegovina"),
    (82,  "Belgium",               "Algeria"),
    (83,  "Portugal",              "Croatia"),
    (84,  "Spain",                 "Austria"),
    (85,  "Switzerland",           "Senegal"),
    (86,  "Argentina",             "Cape Verde"),
    (87,  "Colombia",              "Ghana"),
    (88,  "Australia",             "Egypt"),
]

# QF bracket (winner of M73 vs winner of M74, etc.)
QF = [
    (89,  73, 74),   # W73 vs W74
    (90,  75, 76),   # W75 vs W76
    (91,  77, 78),   # W77 vs W78
    (92,  79, 80),   # W79 vs W80
    (93,  81, 82),   # W81 vs W82
    (94,  83, 84),   # W83 vs W84
    (95,  85, 86),   # W85 vs W86
    (96,  87, 88),   # W87 vs W88
]

SF = [
    (97,  89, 90),   # W89 vs W90
    (98,  91, 92),   # W91 vs W92
    (99,  93, 94),   # W93 vs W94
    (100, 95, 96),   # W95 vs W96
]

FINAL_3RD = [
    (101, 97,  98),  # 3rd place match leg 1
    (102, 99, 100),  # 3rd place match leg 2
]

FINAL = (103, 97, 98, 99, 100)  # (match, SF1_winner, SF2_winner) simplified


SEP = "=" * 72


# =============================================================================
# SEÇÃO 1 — Winsorização + Retreino
# =============================================================================

def wide_to_long(wide: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in wide.iterrows():
        for side, opp in [("home","away"),("away","home")]:
            row: dict = {
                "date":                r["date"],
                "team":                r[f"{side}_team"],
                "gols_marcados":       r[f"{side}_gols"],
                "match_type":          r.get("match_type", ""),
                "opp_saves_decay":     float(r.get(f"{opp}_saves_decay", 1.0)),
                "shots_on_goal_decay": float(r[f"{side}_shots_on_goal_decay"]),
                "gols_sofridos_decay": float(r.get(f"{side}_gols_sofridos_decay", np.nan)),
                "saves_decay":         float(r.get(f"{side}_saves_decay", 1.0)),
            }
            for col in CONF_DUMMIES:
                row[col] = int(r.get(col, 0))
            for col in RATING_FEATURES:
                raw_v = r.get(f"{side}_{col}", np.nan)
                row[col] = float(raw_v) if pd.notna(raw_v) else np.nan
            rows.append(row)
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def apply_saves_normalization(wide: pd.DataFrame) -> pd.DataFrame:
    wide = wide.copy()
    for side, opp in [("home","away"),("away","home")]:
        saves_col = f"{side}_saves_decay"
        if saves_col not in wide.columns:
            continue
        opp_team_col = f"{opp}_team"
        def _med(team, sc=saves_col):
            c = COPA_TEAM_CONF.get(team, "UEFA")
            return CONF_MEDIANS_SAVES.get(c, GLOBAL_MEDIAN_SAVES)
        med_arr = wide[opp_team_col].map(_med).fillna(GLOBAL_MEDIAN_SAVES)
        wide[saves_col] = wide[saves_col].astype(float) / med_arr
    return wide


def winsorize_features(long: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """Winsoriza features numéricas ao p1/p99 para eliminar outliers extremos."""
    long = long.copy()
    win_bounds: dict[str, tuple[float, float]] = {}
    num_feats = ["shots_on_goal_decay", "gols_sofridos_decay", "opp_saves_decay",
                 "gls_ponderado_decay", "gls_decay_vs_forte", "gls_decay_vs_fraco",
                 "margem_gols_decay", "delta_rating_decay"]
    for feat in num_feats:
        if feat not in long.columns:
            continue
        col = long[feat].dropna()
        p1  = float(col.quantile(0.01))
        p99 = float(col.quantile(0.99))
        n_clipped = int(((long[feat] < p1) | (long[feat] > p99)).sum())
        long[feat] = long[feat].clip(lower=p1, upper=p99)
        win_bounds[feat] = (p1, p99)
        if n_clipped > 0:
            print(f"  Winsorize {feat:<30} [{p1:.3f}, {p99:.3f}]  clipped={n_clipped}")
    return long, win_bounds


def train_v13(long: pd.DataFrame, cutoff: pd.Timestamp) -> XGBRegressor:
    train = long[long["date"] <= cutoff].copy()
    X = train[FEATURES].fillna(train[FEATURES].median())
    y = train["gols_marcados"]
    model = XGBRegressor(**XGB_PARAMS)
    model.fit(X, y)
    y_pred = model.predict(X)
    mae  = np.mean(np.abs(y_pred - y))
    bias = y_pred.mean() - y.mean()
    print(f"  v13  n={len(train):,}  MAE={mae:.4f}  bias={bias:+.4f}")
    wc_mask = train["date"] >= pd.Timestamp("2026-06-11")
    if wc_mask.sum() > 0:
        mae_wc  = np.mean(np.abs(y_pred[wc_mask] - y.values[wc_mask]))
        bias_wc = y_pred[wc_mask].mean() - y.values[wc_mask].mean()
        print(f"  Copa2026 n={wc_mask.sum()}  MAE={mae_wc:.4f}  bias={bias_wc:+.4f}")
    return model


# =============================================================================
# SEÇÃO 2 — Extração de features e lambdas
# =============================================================================

def extract_team_feats(long: pd.DataFrame) -> dict[str, dict]:
    team_feats: dict[str, dict] = {}
    for _, row in long.sort_values("date").iterrows():
        entry: dict = {
            "saves_decay":         float(row.get("saves_decay", 1.0)),
            "shots_on_goal_decay": float(row.get("shots_on_goal_decay", 3.5)),
            "gols_sofridos_decay": row.get("gols_sofridos_decay", np.nan),
        }
        for col in RATING_FEATURES:
            v = row.get(col, np.nan)
            entry[col] = float(v) if pd.notna(v) else np.nan
        team_feats[row["team"]] = entry

    if ELO_CSV.exists():
        elo = pd.read_csv(ELO_CSV, parse_dates=["date"])
        for team, val in elo.sort_values("date").groupby("team")["elo_after"].last().items():
            team_feats.setdefault(team, {})["_elo_current"] = float(val)
    return team_feats


def precompute_lambdas(model: XGBRegressor,
                       team_feats: dict[str, dict],
                       teams: list[str]) -> dict[tuple[str, str], float]:
    rows, pairs = [], []
    for t1 in teams:
        for t2 in teams:
            if t1 == t2:
                continue
            f1 = team_feats.get(t1, {})
            f2 = team_feats.get(t2, {})
            row: dict = {
                "opp_saves_decay":     f2.get("saves_decay", 1.0),
                "shots_on_goal_decay": f1.get("shots_on_goal_decay", 3.5),
                "gols_sofridos_decay": f1.get("gols_sofridos_decay", np.nan),
                **{c: 0 for c in CONF_DUMMIES},
            }
            for col in RATING_FEATURES:
                if col == "elo_diff_decay":
                    e1 = f1.get("_elo_current", np.nan)
                    e2 = f2.get("_elo_current", np.nan)
                    row[col] = float(e1) - float(e2) if (pd.notna(e1) and pd.notna(e2)) else f1.get(col, np.nan)
                else:
                    row[col] = f1.get(col, np.nan)
            conf = COPA_TEAM_CONF.get(t1, "")
            ck = f"conf_{conf}"
            if ck in CONF_DUMMIES:
                row[ck] = 1
            rows.append(row)
            pairs.append((t1, t2))
    preds = model.predict(pd.DataFrame(rows)[FEATURES])
    return {pair: float(p) for pair, p in zip(pairs, preds)}


# =============================================================================
# SEÇÃO 3 — Standings reais (Copa 2026 fase de grupos)
# =============================================================================

def compute_real_standings(raw: pd.DataFrame) -> dict[str, pd.DataFrame]:
    copa = raw[(raw["match_type"] == "WC") & (raw["home_away"] == "home")].copy()
    copa["date"] = pd.to_datetime(copa["date"])
    copa["home_gols"] = pd.to_numeric(copa["gols_marcados"], errors="coerce")
    copa["away_gols"] = pd.to_numeric(copa["gols_sofridos"], errors="coerce")

    team_to_group = {t: g for g, ts in GROUPS.items() for t in ts}
    standings: dict[str, dict[str, dict]] = {
        g: {t: {"pts":0,"gf":0,"ga":0,"gd":0,"w":0,"d":0,"l":0,"pgj":0}
            for t in ts}
        for g, ts in GROUPS.items()
    }

    for _, row in copa.iterrows():
        home = NAME_MAP.get(row["team"], row["team"])
        away = NAME_MAP.get(row["opponent"], row["opponent"])
        gh   = int(row["home_gols"]) if pd.notna(row["home_gols"]) else None
        ga   = int(row["away_gols"]) if pd.notna(row["away_gols"]) else None
        if gh is None or ga is None:
            continue
        grp_h = team_to_group.get(home)
        grp_a = team_to_group.get(away)
        if grp_h is None or grp_a is None or grp_h != grp_a:
            continue
        g = grp_h
        for team, gf, ga_t in [(home, gh, ga), (away, ga, gh)]:
            if team not in standings[g]:
                continue
            s = standings[g][team]
            s["gf"] += gf; s["ga"] += ga_t; s["gd"] = s["gf"] - s["ga"]; s["pgj"] += 1
            if gf > ga_t:   s["pts"] += 3; s["w"] += 1
            elif gf == ga_t: s["pts"] += 1; s["d"] += 1
            else:            s["l"] += 1

    result: dict[str, pd.DataFrame] = {}
    for g, td in standings.items():
        df = pd.DataFrame([{"team": t, **v} for t, v in td.items()])
        df = df.sort_values(["pts","gd","gf"], ascending=[False,False,False]).reset_index(drop=True)
        df["pos"] = df.index + 1
        result[g] = df
    return result


def best_third_place_real(standings: dict[str, pd.DataFrame]) -> pd.DataFrame:
    thirds = []
    for g, df in standings.items():
        row = df[df["pos"] == 3].iloc[0].to_dict()
        row["group"] = g
        thirds.append(row)
    df3 = pd.DataFrame(thirds)
    df3 = df3.sort_values(["pts","gd","gf"], ascending=[False,False,False]).reset_index(drop=True)
    df3["rank3"] = df3.index + 1
    return df3


def assign_thirds(standings: dict[str, pd.DataFrame]) -> dict[int, str]:
    """Bipartite matching: assigns 8 best 3rds to R32 slots."""
    thirds_df = best_third_place_real(standings)
    best8 = thirds_df.head(8)

    slot_groups = {
        74: ["A","B","C","D","F"],
        77: ["C","D","F","G","H"],
        79: ["A","B","C","D","E"],
        81: ["A","E","G","H","I"],
        82: ["A","B","C","D","E","F","G","H","I"],
        87: ["E","F","H","I","J"],
        88: ["G","H","I","J","K"],
        83: ["A","B","C","D","F","G","H","I","J","K"],
    }

    assigned: dict[int, str] = {}
    used_teams = set()

    def augment(slot_idx: int, slot_ids: list[int], matching: dict) -> bool:
        if slot_idx == len(slot_ids):
            return True
        sid = slot_ids[slot_idx]
        eligible_groups = slot_groups.get(sid, [g for g in GROUPS])
        for _, row in best8.iterrows():
            team = row["team"]
            grp  = row["group"]
            if team in used_teams:
                continue
            if grp not in eligible_groups:
                continue
            matching[sid] = team
            used_teams.add(team)
            if augment(slot_idx + 1, slot_ids, matching):
                return True
            del matching[sid]
            used_teams.discard(team)
        return False

    slot_ids = list(slot_groups.keys())
    augment(0, slot_ids, assigned)

    # fallback for any unassigned slots
    for sid in slot_ids:
        if sid not in assigned:
            for _, row in best8.iterrows():
                if row["team"] not in used_teams:
                    assigned[sid] = row["team"]
                    used_teams.add(row["team"])
                    break
    return assigned


# =============================================================================
# SEÇÃO 4 — Poisson utilities
# =============================================================================

def most_probable_score(lam1: float, lam2: float,
                        max_g: int = 8) -> tuple[int, int, float]:
    """(g1, g2, prob) com maior probabilidade Poisson conjunta."""
    best_p, best_g1, best_g2 = -1.0, 0, 0
    for g1 in range(max_g + 1):
        for g2 in range(max_g + 1):
            p = scipy_poisson.pmf(g1, lam1) * scipy_poisson.pmf(g2, lam2)
            if p > best_p:
                best_p, best_g1, best_g2 = p, g1, g2
    return best_g1, best_g2, best_p


def poisson_match_probs(lam1: float, lam2: float,
                        max_g: int = 8) -> tuple[float, float, float]:
    """(P(t1 win), P(draw), P(t2 win)) em 90 minutos."""
    p1 = p_draw = p2 = 0.0
    for g1 in range(max_g + 1):
        for g2 in range(max_g + 1):
            p = scipy_poisson.pmf(g1, lam1) * scipy_poisson.pmf(g2, lam2)
            if g1 > g2:   p1     += p
            elif g1 == g2: p_draw += p
            else:          p2     += p
    return p1, p_draw, p2


def knock_prob(lam1: float, lam2: float, max_g: int = 8) -> float:
    """P(t1 avança) em jogo eliminatório (pênaltis = 50/50 no empate)."""
    p1, p_draw, p2 = poisson_match_probs(lam1, lam2, max_g)
    return p1 + 0.5 * p_draw


def sim_knockout(lam1: float, lam2: float, rng: np.random.Generator) -> tuple[int, bool]:
    """Simula jogo eliminatório. Retorna (vencedor 0|1, foi pênalti)."""
    g1 = rng.poisson(lam1)
    g2 = rng.poisson(lam2)
    if g1 > g2:   return 0, False
    elif g2 > g1: return 1, False
    else:         return int(rng.random() > 0.5), True


# =============================================================================
# SEÇÃO 5 — Simulação determinística R32→Final
# =============================================================================

def simulate_deterministic(lam: dict[tuple[str,str], float],
                            r32_matchups: list[tuple[int, str, str]]
                            ) -> dict:
    """
    Simula bracket completo com placar mais provável por jogo.
    Empate em 90' → time com maior λ avança (sem prorrogação no "mais provável").
    """
    winners: dict[int, str] = {}  # match_id → vencedor
    games: list[dict] = []

    def play(mid: int, t1: str, t2: str, stage: str) -> str:
        l1 = lam.get((t1, t2), 1.3)
        l2 = lam.get((t2, t1), 1.3)
        g1, g2, prob = most_probable_score(l1, l2)
        p1, p_draw, p2 = poisson_match_probs(l1, l2)
        pk = knock_prob(l1, l2)

        if g1 > g2:    winner = t1; note = ""
        elif g2 > g1:  winner = t2; note = ""
        else:          # empate: tiebreak por λ
            if l1 >= l2: winner, note = t1, " (pên)"
            else:        winner, note = t2, " (pên)"

        games.append({
            "stage": stage, "mid": mid,
            "t1": t1, "t2": t2,
            "g1": g1, "g2": g2,
            "score": f"{g1}–{g2}{note}",
            "winner": winner,
            "l1": round(l1, 3), "l2": round(l2, 3),
            "p_t1": round(p1, 4), "p_emp": round(p_draw, 4), "p_t2": round(p2, 4),
            "p_t1_adv": round(pk, 4),
            "prob_score": round(prob, 4),
        })
        return winner

    # R32
    for mid, t1, t2 in r32_matchups:
        w = play(mid, t1, t2, "R32")
        winners[mid] = w

    # QF
    qf_ids = [89, 90, 91, 92, 93, 94, 95, 96]
    r32_ids_ordered = [mid for mid, _, _ in r32_matchups]  # 73..88
    for i, qf_id in enumerate(qf_ids):
        t1 = winners[r32_ids_ordered[i*2]]
        t2 = winners[r32_ids_ordered[i*2 + 1]]
        w  = play(qf_id, t1, t2, "QF")
        winners[qf_id] = w

    # SF
    sf_ids = [97, 98, 99, 100]
    qf_ids_ordered = qf_ids
    for i, sf_id in enumerate(sf_ids):
        t1 = winners[qf_ids_ordered[i*2]]
        t2 = winners[qf_ids_ordered[i*2 + 1]]
        w  = play(sf_id, t1, t2, "SF")
        winners[sf_id] = w

    # 3rd place (losers of SF)
    def sf_loser(sf_id):
        g = [gm for gm in games if gm["mid"] == sf_id][0]
        return g["t2"] if g["winner"] == g["t1"] else g["t1"]
    t1_3 = sf_loser(97); t2_3 = sf_loser(98)
    play(101, t1_3, t2_3, "3o Lugar (SF1)")
    t3_3 = sf_loser(99); t4_3 = sf_loser(100)
    play(102, t3_3, t4_3, "3o Lugar (SF2)")

    # Final (duas finais: SF1 winner vs SF2 winner, SF3 winner vs SF4 winner)
    # Copa 2026: final única entre 2 semi campeões do Bracket A e Bracket B
    play(103, winners[97], winners[98], "FINAL A")
    play(104, winners[99], winners[100], "FINAL B")

    # Campeon
    winner_a = winners.get(103) or [g for g in games if g["mid"]==103][0]["winner"]
    winner_b = winners.get(104) or [g for g in games if g["mid"]==104][0]["winner"]

    return {
        "games": games,
        "winners": winners,
        "champion_a": winner_a,
        "champion_b": winner_b,
    }


# =============================================================================
# SEÇÃO 6 — Monte Carlo 10k
# =============================================================================

def run_mc(lam: dict[tuple[str,str], float],
           r32_matchups: list[tuple[int, str, str]],
           n_sim: int = 10_000) -> tuple[dict[str, int], dict[str, dict[str, int]]]:
    rng = np.random.default_rng(SEED)
    champions:    dict[str, int] = {}
    round_adv:    dict[str, dict[str, int]] = {
        t: {"R32": 0, "QF": 0, "SF": 0, "Final": 0, "Campeon": 0}
        for t in ALL_COPA_TEAMS
    }

    r32_ids_ordered = [mid for mid, _, _ in r32_matchups]

    for _ in range(n_sim):
        w: dict[int, str] = {}

        # R32
        for mid, t1, t2 in r32_matchups:
            l1 = lam.get((t1, t2), 1.3)
            l2 = lam.get((t2, t1), 1.3)
            idx, _ = sim_knockout(l1, l2, rng)
            winner = t1 if idx == 0 else t2
            w[mid] = winner
            round_adv[winner]["R32"] += 1

        # QF
        qf_ids = [89, 90, 91, 92, 93, 94, 95, 96]
        for i, qf_id in enumerate(qf_ids):
            t1 = w[r32_ids_ordered[i*2]]
            t2 = w[r32_ids_ordered[i*2 + 1]]
            l1 = lam.get((t1, t2), 1.3)
            l2 = lam.get((t2, t1), 1.3)
            idx, _ = sim_knockout(l1, l2, rng)
            winner = t1 if idx == 0 else t2
            w[qf_id] = winner
            round_adv[winner]["QF"] += 1

        # SF
        sf_ids = [97, 98, 99, 100]
        for i, sf_id in enumerate(sf_ids):
            t1 = w[qf_ids[i*2]]
            t2 = w[qf_ids[i*2 + 1]]
            l1 = lam.get((t1, t2), 1.3)
            l2 = lam.get((t2, t1), 1.3)
            idx, _ = sim_knockout(l1, l2, rng)
            winner = t1 if idx == 0 else t2
            w[sf_id] = winner
            round_adv[winner]["SF"] += 1

        # Finals (2 finais em paralelo)
        for final_id, sf1_id, sf2_id in [(103, 97, 98), (104, 99, 100)]:
            t1 = w[sf1_id]; t2 = w[sf2_id]
            l1 = lam.get((t1, t2), 1.3)
            l2 = lam.get((t2, t1), 1.3)
            idx, _ = sim_knockout(l1, l2, rng)
            winner = t1 if idx == 0 else t2
            w[final_id] = winner
            round_adv[t1]["Final"] += 1
            round_adv[t2]["Final"] += 1
            round_adv[winner]["Campeon"] += 1
            champions[winner] = champions.get(winner, 0) + 1

    return champions, round_adv


# =============================================================================
# SEÇÃO 7 — Excel final
# =============================================================================

STAGE_COLORS = {
    "R32":  "D6EAF8",
    "QF":   "D5F5E3",
    "SF":   "FDEBD0",
    "3o Lugar (SF1)": "F0E6FF",
    "3o Lugar (SF2)": "F0E6FF",
    "FINAL A": "FDEDEC",
    "FINAL B": "FDEDEC",
}

def _header_cell(ws, row, col, value, fill_hex="1A5276", font_color="FFFFFF", bold=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = PatternFill("solid", fgColor=fill_hex)
    cell.font   = Font(bold=bold, color=font_color, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _body_cell(ws, row, col, value, fill_hex=None, bold=False, num_fmt=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.font   = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if num_fmt:
        cell.number_format = num_fmt
    return cell


def _thin_border():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel(
    out_path: Path,
    standings: dict[str, pd.DataFrame],
    det_result: dict,
    mc_champ: dict[str, int],
    round_adv: dict[str, dict[str, int]],
    n_sim: int,
    r32_matchups: list[tuple[int, str, str]],
    lam: dict[tuple[str,str], float],
) -> None:

    wb = openpyxl.Workbook()

    # ─── Aba 1: Grupos ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Grupos"
    ws1.sheet_view.showGridLines = False

    row = 1
    headers = ["Pos", "Time", "PJ", "V", "E", "D", "GP", "GC", "SG", "Pts"]
    for g in sorted(GROUPS.keys()):
        _header_cell(ws1, row, 1, f"Grupo {g}", fill_hex="1A5276")
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        row += 1
        for c, h in enumerate(headers, 1):
            _header_cell(ws1, row, c, h, fill_hex="2E86C1")
        row += 1
        df = standings[g]
        for _, r in df.iterrows():
            fill = "FDFEFE" if int(r["pos"]) % 2 == 0 else "EBF5FB"
            medal = {1: "🥇", 2: "🥈", 3: "🥉"}.get(int(r["pos"]), "  ")
            vals = [f"{medal} {int(r['pos'])}",
                    r["team"], int(r["pgj"]), int(r["w"]), int(r["d"]), int(r["l"]),
                    int(r["gf"]), int(r["ga"]), int(r["gd"]), int(r["pts"])]
            for c, v in enumerate(vals, 1):
                cell = _body_cell(ws1, row, c, v, fill_hex=fill,
                                  bold=(int(r["pos"]) <= 2), align="center")
                cell.border = _thin_border()
            row += 1
        row += 1  # blank row between groups

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 26
    for col_letter in ["C","D","E","F","G","H","I","J"]:
        ws1.column_dimensions[col_letter].width = 7

    # ─── Aba 2: Bracket Determinístico ───────────────────────────────────────
    ws2 = wb.create_sheet("Bracket_Deterministico")
    ws2.sheet_view.showGridLines = False

    stages = ["R32","QF","SF","FINAL A","FINAL B","3o Lugar (SF1)","3o Lugar (SF2)"]
    games  = det_result["games"]

    row = 1
    _header_cell(ws2, row, 1, "Simulação Determinística — Placar Mais Provável", fill_hex="154360")
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    hdrs = ["Stage","#","Time 1","λ1","Time 2","λ2","Placar","Vencedor","P(1 adv)","P(empate 90')"]
    for c, h in enumerate(hdrs, 1):
        _header_cell(ws2, row, c, h, fill_hex="1F618D")
    row += 1

    for stage in stages:
        stage_games = [g for g in games if g["stage"] == stage]
        if not stage_games:
            continue
        fill_hex = STAGE_COLORS.get(stage, "FDFEFE")
        # stage separator
        _header_cell(ws2, row, 1, stage, fill_hex="D0D0D0", font_color="1A1A1A", bold=True)
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1

        for gm in stage_games:
            winner = gm["winner"]
            vals = [
                gm["stage"], gm["mid"],
                gm["t1"], gm["l1"],
                gm["t2"], gm["l2"],
                gm["score"],
                winner,
                f"{gm['p_t1_adv']:.1%}",
                f"{gm['p_emp']:.1%}",
            ]
            for c, v in enumerate(vals, 1):
                is_winner_col = (c == 3 and gm["t1"] == winner) or (c == 5 and gm["t2"] == winner)
                cell = _body_cell(ws2, row, c, v,
                                  fill_hex=fill_hex,
                                  bold=(c == 8 or is_winner_col),
                                  align="center" if c != 3 and c != 5 and c != 8 else "left")
                cell.border = _thin_border()
                if c == 8:
                    cell.fill = PatternFill("solid", fgColor="D4EFDF")
            row += 1

    col_widths = [12, 5, 28, 7, 28, 7, 10, 28, 10, 14]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ─── Aba 3: MC Probabilidades por Round ──────────────────────────────────
    ws3 = wb.create_sheet("MC_Probabilidades")
    ws3.sheet_view.showGridLines = False

    row = 1
    _header_cell(ws3, row, 1, f"Monte Carlo {n_sim:,} simulações — P(avançar) por fase", fill_hex="154360")
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    hdrs3 = ["Confederação","Time","P(R32 adv)","P(QF)","P(SF)","P(Final)","P(Campeão)","λ_atk","λ_def"]
    for c, h in enumerate(hdrs3, 1):
        _header_cell(ws3, row, c, h, fill_hex="1F618D")
    row += 1

    # Compute λ_atk (avg scoring rate) and λ_def (avg conceding rate) for each team
    copa_teams_sorted = sorted(ALL_COPA_TEAMS, key=lambda t: round_adv.get(t, {}).get("Campeon", 0), reverse=True)

    for t in copa_teams_sorted:
        conf = COPA_TEAM_CONF.get(t, "?")
        ra   = round_adv.get(t, {})
        lam_atk = np.mean([lam.get((t, opp), 1.3) for opp in ALL_COPA_TEAMS if opp != t])
        lam_def = np.mean([lam.get((opp, t), 1.3) for opp in ALL_COPA_TEAMS if opp != t])

        # Determine R32 participation
        in_r32 = any(t in (r32t1, r32t2) for _, r32t1, r32t2 in r32_matchups)
        r32_base = n_sim if in_r32 else 0

        vals = [
            conf, t,
            ra.get("R32", 0) / max(r32_base, 1) if r32_base else "N/A",
            ra.get("QF", 0)    / n_sim,
            ra.get("SF", 0)    / n_sim,
            ra.get("Final", 0) / n_sim,
            ra.get("Campeon", 0) / n_sim,
            round(lam_atk, 3),
            round(lam_def, 3),
        ]

        fill_hex = "FDFEFE" if row % 2 == 0 else "EBF5FB"
        champ_pct = ra.get("Campeon", 0) / n_sim
        if champ_pct >= 0.10:   fill_hex = "FFDDD5"
        elif champ_pct >= 0.05: fill_hex = "FFF3CD"
        elif champ_pct >= 0.02: fill_hex = "D5F5E3"

        for c, v in enumerate(vals, 1):
            fmt = "0.0%" if isinstance(v, float) and c not in (8, 9) else (
                  "0.000" if c in (8, 9) else None)
            cell = _body_cell(ws3, row, c, v, fill_hex=fill_hex,
                               bold=(champ_pct >= 0.05), num_fmt=fmt, align="center")
            cell.border = _thin_border()
        row += 1

    col_widths3 = [13, 28, 12, 9, 9, 9, 12, 9, 9]
    for i, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ─── Aba 4: MC Campeão ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("MC_Campeon")
    ws4.sheet_view.showGridLines = False

    row = 1
    _header_cell(ws4, row, 1, f"MC {n_sim:,} sim — Top Campeões Copa 2026", fill_hex="154360")
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    for c, h in enumerate(["Rank","Confederação","Time","Simulações","Probabilidade"], 1):
        _header_cell(ws4, row, c, h, fill_hex="1F618D")
    row += 1

    sorted_champs = sorted(mc_champ.items(), key=lambda x: x[1], reverse=True)
    total = sum(mc_champ.values())
    fills = ["FFD700","C0C0C0","CD7F32"] + ["FDFEFE"] * 100

    for rank, (team, cnt) in enumerate(sorted_champs, 1):
        pct = cnt / total
        conf = COPA_TEAM_CONF.get(team, "?")
        fill = fills[min(rank-1, len(fills)-1)] if rank <= 3 else (
               "FFDDD5" if pct >= 0.10 else ("FFF3CD" if pct >= 0.05 else
               ("D5F5E3" if pct >= 0.02 else ("EBF5FB" if row%2==0 else "FDFEFE"))))
        vals = [rank, conf, team, cnt, pct]
        for c, v in enumerate(vals, 1):
            cell = _body_cell(ws4, row, c, v,
                              fill_hex=fill, bold=(rank <= 3),
                              num_fmt="0.0%" if c == 5 else ("0" if c == 4 else None),
                              align="center")
            cell.border = _thin_border()
        row += 1

    col_widths4 = [6, 13, 28, 12, 14]
    for i, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"  Excel salvo: {out_path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print(f"\n{SEP}")
    print("  COPA 2026 v13 — Winsorização + MC 10k + Simulação Determinística")
    print(f"{SEP}\n")

    # 1. Carregar e processar dados
    print("1. Carregando model_dataset.csv...")
    wide = pd.read_csv(MODEL_CSV, parse_dates=["date"])
    wide = apply_saves_normalization(wide)
    long = wide_to_long(wide)
    print(f"   {len(long)} obs | opp_saves_decay coverage: {long['opp_saves_decay'].notna().mean():.1%}")

    # 2. Winsorização
    print("\n2. Winsorização p1/p99...")
    long_win, win_bounds = winsorize_features(long)
    if not win_bounds:
        print("   Nenhuma feature winsorizada.")

    # 3. Retreino v13
    print("\n3. Retreinando xgboost_v13 (winsorizado)...")
    model = train_v13(long_win, CUTOFF)
    with open(OUT_PKL, "wb") as f:
        pickle.dump(model, f)
    print(f"   Modelo salvo: {OUT_PKL}")

    # 4. Features e lambdas
    print("\n4. Extraindo features e computando lambdas...")
    team_feats = extract_team_feats(long_win)
    lam = precompute_lambdas(model, team_feats, ALL_COPA_TEAMS)
    print(f"   {len(lam)} pares de lambda")

    # Preview: top-5 times por λ_atk
    lam_atk = {t: np.mean([lam.get((t, opp), 1.3) for opp in ALL_COPA_TEAMS if opp != t])
                for t in ALL_COPA_TEAMS}
    top5 = sorted(lam_atk.items(), key=lambda x: x[1], reverse=True)[:5]
    print(f"   Top-5 λ_atk: {[(t, round(v,3)) for t,v in top5]}")

    # 5. Standings reais
    print("\n5. Classificação real Copa 2026...")
    raw = pd.read_csv(RAW_CSV)
    standings = compute_real_standings(raw)
    for g in sorted(GROUPS.keys()):
        df = standings[g]
        first = df.iloc[0]; second = df.iloc[1]
        print(f"   Grupo {g}: 1º {first['team']} ({first['pts']}pts)  "
              f"2º {second['team']} ({second['pts']}pts)")

    # 6. R32 matchups reais
    r32_matchups = REAL_R32
    print("\n6. R32 (chaveamento real):")
    for mid, t1, t2 in r32_matchups:
        l1 = lam.get((t1,t2), 1.3); l2 = lam.get((t2,t1), 1.3)
        pk = knock_prob(l1, l2)
        fav = t1 if pk >= 0.5 else t2
        prob = pk if pk >= 0.5 else 1 - pk
        g1, g2, _ = most_probable_score(l1, l2)
        print(f"   M{mid:>3}  {t1:<25} {g1}–{g2}  {t2:<25}  → {fav} ({prob:.0%})")

    # 7. Simulação determinística
    print("\n7. Simulação determinística R32→Final...")
    det = simulate_deterministic(lam, r32_matchups)
    print("\n   Bracket determinístico completo:")
    for stage in ["R32","QF","SF","FINAL A","FINAL B"]:
        games_s = [g for g in det["games"] if g["stage"] == stage]
        if not games_s:
            continue
        print(f"\n   {stage}:")
        for gm in games_s:
            print(f"     M{gm['mid']:>3}  {gm['t1']:<25} {gm['score']:>7}  "
                  f"{gm['t2']:<25}  → {gm['winner']}")
    print(f"\n   Campeão Bracket A: {det['champion_a']}")
    print(f"   Campeão Bracket B: {det['champion_b']}")

    # 8. Monte Carlo
    print(f"\n8. Monte Carlo {N_SIM:,} simulações...")
    champions, round_adv = run_mc(lam, r32_matchups, N_SIM)
    total = sum(champions.values())
    print("   Top-10 campeões:")
    for rank, (team, cnt) in enumerate(
        sorted(champions.items(), key=lambda x: x[1], reverse=True)[:10], 1
    ):
        pct = cnt / total * 100
        print(f"   {rank:>2}. {team:<25} {pct:5.1f}%  (n={cnt})")

    # 9. Excel
    print("\n9. Gerando Excel...")
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    build_excel(
        out_path    = OUT_XLSX,
        standings   = standings,
        det_result  = det,
        mc_champ    = champions,
        round_adv   = round_adv,
        n_sim       = N_SIM,
        r32_matchups = r32_matchups,
        lam         = lam,
    )

    print(f"\n{SEP}")
    print(f"  Concluído! Outputs:")
    print(f"  • Modelo: {OUT_PKL}")
    print(f"  • Excel:  {OUT_XLSX}")
    print(f"{SEP}\n")


if __name__ == "__main__":
    main()
