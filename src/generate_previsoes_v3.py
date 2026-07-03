"""
generate_previsoes_v3.py — Gera copa2026_previsoes_v3.xlsx com dados do modelo v6.

Recria a estrutura do arquivo original (05/jun/2026) com:
  - "Fase de Grupos"      : 48 jogos com λ, V%/E%/D%, top-3 placares (v5)
  - "Mata-Mata"           : bracket determinístico v5
  - "Classificação Esperada": esperança analítica de pts/GF/GA por grupo
"""

import sys
import warnings
from pathlib import Path
from itertools import combinations

import numpy as np
import pandas as pd
from scipy.stats import poisson as sp_poisson
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent))
from simulate_copa2026 import (
    train_model, precompute_lambdas,
    simulate_all_groups, simulate_knockouts,
    GROUPS, R32,
)

OUT_XLSX = Path("outputs/copa2026_previsoes_v3.xlsx")

# ---------------------------------------------------------------------------
# Traduções
# ---------------------------------------------------------------------------

TEAM_PT: dict[str, str] = {
    "Mexico":                  "México",
    "South Korea":             "Coreia do Sul",
    "South Africa":            "África do Sul",
    "Czech Republic":          "Rep. Tcheca",
    "Canada":                  "Canadá",
    "Switzerland":             "Suíça",
    "Qatar":                   "Catar",
    "Bosnia and Herzegovina":  "Bósnia-Herz.",
    "Brazil":                  "Brasil",
    "Morocco":                 "Marrocos",
    "Haiti":                   "Haiti",
    "Scotland":                "Escócia",
    "United States":           "EUA",
    "Paraguay":                "Paraguai",
    "Australia":               "Austrália",
    "Turkey":                  "Turquia",
    "Germany":                 "Alemanha",
    "Ivory Coast":             "Costa do Marfim",
    "Ecuador":                 "Equador",
    "Curacao":                 "Curaçao",
    "Netherlands":             "Holanda",
    "Sweden":                  "Suécia",
    "Tunisia":                 "Tunísia",
    "Japan":                   "Japão",
    "Belgium":                 "Bélgica",
    "Iran":                    "Irã",
    "New Zealand":             "Nova Zelândia",
    "Egypt":                   "Egito",
    "Spain":                   "Espanha",
    "Saudi Arabia":            "Arábia Saudita",
    "Uruguay":                 "Uruguai",
    "Cape Verde":              "Cabo Verde",
    "France":                  "França",
    "Senegal":                 "Senegal",
    "Iraq":                    "Iraque",
    "Norway":                  "Noruega",
    "Argentina":               "Argentina",
    "Algeria":                 "Argélia",
    "Austria":                 "Áustria",
    "Jordan":                  "Jordânia",
    "Portugal":                "Portugal",
    "DR Congo":                "RD Congo",
    "Uzbekistan":              "Uzbequistão",
    "Colombia":                "Colômbia",
    "England":                 "Inglaterra",
    "Croatia":                 "Croácia",
    "Ghana":                   "Gana",
    "Panama":                  "Panamá",
}

GROUP_HEADER_PT: dict[str, str] = {
    "A": "GRUPO A  —  México  ·  Coreia do Sul  ·  África do Sul  ·  Rep. Tcheca",
    "B": "GRUPO B  —  Canadá  ·  Suíça  ·  Catar  ·  Bósnia-Herz.",
    "C": "GRUPO C  —  Brasil  ·  Marrocos  ·  Haiti  ·  Escócia",
    "D": "GRUPO D  —  EUA  ·  Paraguai  ·  Austrália  ·  Turquia",
    "E": "GRUPO E  —  Alemanha  ·  Costa do Marfim  ·  Equador  ·  Curaçao",
    "F": "GRUPO F  —  Holanda  ·  Suécia  ·  Tunísia  ·  Japão",
    "G": "GRUPO G  —  Bélgica  ·  Irã  ·  Nova Zelândia  ·  Egito",
    "H": "GRUPO H  —  Espanha  ·  Arábia Saudita  ·  Uruguai  ·  Cabo Verde",
    "I": "GRUPO I  —  França  ·  Senegal  ·  Iraque  ·  Noruega",
    "J": "GRUPO J  —  Argentina  ·  Argélia  ·  Áustria  ·  Jordânia",
    "K": "GRUPO K  —  Portugal  ·  RD Congo  ·  Uzbequistão  ·  Colômbia",
    "L": "GRUPO L  —  Inglaterra  ·  Croácia  ·  Gana  ·  Panamá",
}


def pt(team: str) -> str:
    return TEAM_PT.get(team, team)


# ---------------------------------------------------------------------------
# Poisson helpers
# ---------------------------------------------------------------------------

def poisson_probs(lh: float, la: float, max_g: int = 10) -> tuple[float, float, float]:
    p_w = p_d = p_l = 0.0
    for h in range(max_g + 1):
        ph = sp_poisson.pmf(h, lh)
        for a in range(max_g + 1):
            p = ph * sp_poisson.pmf(a, la)
            if h > a:    p_w += p
            elif h == a: p_d += p
            else:        p_l += p
    return p_w, p_d, p_l


def top3_scores(lh: float, la: float, max_g: int = 8) -> list[str]:
    scores = []
    for h in range(max_g + 1):
        ph = sp_poisson.pmf(h, lh)
        for a in range(max_g + 1):
            p = ph * sp_poisson.pmf(a, la)
            scores.append((h, a, p))
    scores.sort(key=lambda x: x[2], reverse=True)
    return [f"{h}–{a} ({p * 100:.1f}%)" for h, a, p in scores[:3]]


# ---------------------------------------------------------------------------
# ELO
# ---------------------------------------------------------------------------

def load_elo(team_feats: dict) -> dict[str, int]:
    elo: dict[str, int] = {}
    elo_path = Path("data/raw/elo_history.csv")
    if elo_path.exists():
        eh = pd.read_csv(elo_path, parse_dates=["date"])
        for team, val in (eh.sort_values("date")
                            .groupby("team")["elo_after"].last().items()):
            elo[team] = int(round(float(val)))
    for team, feats in team_feats.items():
        if team not in elo and "_elo_current" in feats:
            elo[team] = int(round(feats["_elo_current"]))
    return elo


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_grupos_rows(lam_cache: dict, elo: dict) -> list[list]:
    rows: list[list] = []
    rows.append(["COPA DO MUNDO 2026 — PREVISÃO DE PLACARES  (modelo XGBoost v6)"] + [None] * 14)
    rows.append(["Grp", "Time 1", "ELO", "λ1", "×",
                 "λ2", "ELO", "Time 2",
                 "V%", "E%", "D%", "Favorito",
                 "1º Placar", "2º Placar", "3º Placar"])

    for g in sorted(GROUPS.keys()):
        teams = GROUPS[g]
        rows.append([GROUP_HEADER_PT[g]] + [None] * 14)
        for t1, t2 in combinations(teams, 2):
            lh = lam_cache.get((t1, t2), 1.2)
            la = lam_cache.get((t2, t1), 1.0)
            pw, pd_, pl = poisson_probs(lh, la)
            fav = pt(t1) if pw >= pl else pt(t2)
            sc = top3_scores(lh, la)
            rows.append([
                g, pt(t1), elo.get(t1, ""), round(lh, 3),
                "×", round(la, 3), elo.get(t2, ""), pt(t2),
                f"{pw * 100:.1f}%", f"{pd_ * 100:.1f}%", f"{pl * 100:.1f}%",
                fav, sc[0], sc[1], sc[2],
            ])
    return rows


def build_matamata_rows(lam_cache: dict, elo: dict,
                        det_standings, det_ko) -> list[list]:
    rows: list[list] = []
    rows.append(["COPA DO MUNDO 2026 — MATA-MATA DETERMINÍSTICO  (modelo XGBoost v6)"] + [None] * 14)
    rows.append(["Fase", "Time 1", "ELO", "λ1", "×",
                 "λ2", "ELO", "Time 2",
                 "V%", "E%", "D%", "Favorito",
                 "1º Placar", "2º Placar", "3º Placar"])

    mt = det_ko["teams"]

    def _add_phase(phase_label: str, match_ids: list[int]) -> None:
        rows.append([phase_label] + [None] * 14)
        for mid in match_ids:
            t1, t2 = mt[mid]
            lh = lam_cache.get((t1, t2), 1.2)
            la = lam_cache.get((t2, t1), 1.0)
            pw, pd_, pl = poisson_probs(lh, la)
            # Favorito = quem o bracket determinístico classifica (não apenas λ regular)
            fav = pt(det_ko["winner"][mid])
            sc = top3_scores(lh, la)
            rows.append([
                phase_label, pt(t1), elo.get(t1, ""), round(lh, 3),
                "×", round(la, 3), elo.get(t2, ""), pt(t2),
                f"{pw * 100:.1f}%", f"{pd_ * 100:.1f}%", f"{pl * 100:.1f}%",
                fav, sc[0], sc[1], sc[2],
            ])

    _add_phase("DEZESSEIS-AVOS",    [r[0] for r in R32])
    _add_phase("OITAVAS",    [89, 90, 91, 92, 93, 94, 95, 96])
    _add_phase("QUARTAS",       [97, 98, 99, 100])
    _add_phase("SEMIFINAIS", [101, 102])
    _add_phase("3º LUGAR", [103])
    _add_phase("FINAL",      [104])

    return rows


def build_classificacao_rows(lam_cache: dict) -> list[list]:
    # Analytical expected pts / GF / GA
    exp: dict[str, dict] = {
        t: {"pts": 0.0, "gf": 0.0, "ga": 0.0}
        for ts in GROUPS.values() for t in ts
    }
    for teams in GROUPS.values():
        for t1, t2 in combinations(teams, 2):
            lh = lam_cache.get((t1, t2), 1.2)
            la = lam_cache.get((t2, t1), 1.0)
            pw, pd_, pl = poisson_probs(lh, la)
            exp[t1]["pts"] += 3 * pw + pd_
            exp[t2]["pts"] += 3 * pl + pd_
            exp[t1]["gf"] += lh;  exp[t1]["ga"] += la
            exp[t2]["gf"] += la;  exp[t2]["ga"] += lh

    def sort_key(t):
        e = exp[t]
        return (e["pts"], e["gf"] - e["ga"], e["gf"])

    # Identify top-8 third-place teams
    thirds: list[dict] = []
    for g in sorted(GROUPS.keys()):
        ranked = sorted(GROUPS[g], key=sort_key, reverse=True)
        t3 = ranked[2]
        thirds.append({"team": t3, "pts": exp[t3]["pts"]})
    thirds.sort(key=lambda x: x["pts"], reverse=True)
    qual_thirds = {d["team"] for d in thirds[:8]}

    rows: list[list] = []
    rows.append(["COPA DO MUNDO 2026 — CLASSIFICAÇÃO ESPERADA (pontos × probabilidades)"] + [None] * 7)
    rows.append(["Pos", "Time", "Pts Esp.", "GF Esp.", "GA Esp.", "Saldo", "Grupo", "Obs."])

    for g in sorted(GROUPS.keys()):
        ranked = sorted(GROUPS[g], key=sort_key, reverse=True)
        for pos, team in enumerate(ranked, 1):
            e = exp[team]
            saldo = e["gf"] - e["ga"]
            if pos <= 2:
                obs = "Classifica"
            elif pos == 3 and team in qual_thirds:
                obs = "3º qualif."
            else:
                obs = None
            rows.append([
                pos, pt(team),
                round(e["pts"], 2), round(e["gf"], 2), round(e["ga"], 2),
                round(saldo, 2), g, obs,
            ])
        rows.append([None] * 8)

    # Remove trailing spacer
    if rows and all(v is None for v in rows[-1]):
        rows.pop()

    return rows


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

FILL_TITLE  = PatternFill("solid", fgColor="1F4E79")   # dark blue
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")   # medium blue
FILL_GROUP  = PatternFill("solid", fgColor="D6E4F0")   # light blue
FILL_PHASE  = PatternFill("solid", fgColor="E2EFDA")   # light green

FONT_TITLE  = Font(bold=True, color="FFFFFF", size=12)
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONT_GROUP  = Font(bold=True, color="1F4E79", size=10)
FONT_PHASE  = Font(bold=True, color="375623", size=10)
FONT_NORMAL = Font(size=10)


def _style_row(ws, r_idx: int, fill, font, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=r_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_grupos(ws, rows: list[list]) -> None:
    n_cols = 15
    col_widths = [5, 22, 6, 7, 3, 7, 6, 22, 7, 7, 7, 20, 16, 16, 16]
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _style_row(ws, 1, FILL_TITLE, FONT_TITLE, n_cols)

    # Header row
    _style_row(ws, 2, FILL_HEADER, FONT_HEADER, n_cols)

    # Group/match rows
    for r_idx, row in enumerate(rows, 1):
        if r_idx <= 2:
            continue
        first = row[0]
        if isinstance(first, str) and first.startswith("GRUPO"):
            ws.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=n_cols)
            _style_row(ws, r_idx, FILL_GROUP, FONT_GROUP, n_cols)

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_matamata(ws, rows: list[list]) -> None:
    n_cols = 15
    col_widths = [12, 22, 6, 7, 3, 7, 6, 22, 7, 7, 7, 20, 16, 16, 16]
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    PHASE_FILLS = {
        "DEZESSEIS-AVOS": PatternFill("solid", fgColor="E2EFDA"),
        "OITAVAS":    PatternFill("solid", fgColor="FFF2CC"),
        "QUARTAS":    PatternFill("solid", fgColor="FCE4D6"),
        "SEMIFINAIS": PatternFill("solid", fgColor="F4B8C1"),
        "3º LUGAR": PatternFill("solid", fgColor="E2EFDA"),
        "FINAL":      PatternFill("solid", fgColor="FFD700"),
    }
    PHASE_FONTS = {
        "DEZESSEIS-AVOS": Font(bold=True, color="375623", size=10),
        "OITAVAS":    Font(bold=True, color="7F6000", size=10),
        "QUARTAS":    Font(bold=True, color="843C0C", size=10),
        "SEMIFINAIS": Font(bold=True, color="99002B", size=10),
        "3º LUGAR": Font(bold=True, color="375623", size=10),
        "FINAL":      Font(bold=True, color="7F4800", size=10),
    }

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center")

    _style_row(ws, 1, FILL_TITLE, FONT_TITLE, n_cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _style_row(ws, 2, FILL_HEADER, FONT_HEADER, n_cols)

    for r_idx, row in enumerate(rows, 1):
        if r_idx <= 2:
            continue
        phase = row[0]
        if isinstance(phase, str) and row[1] is None:
            # Section header row
            ws.merge_cells(start_row=r_idx, start_column=1,
                           end_row=r_idx, end_column=n_cols)
            fill = PHASE_FILLS.get(phase, FILL_PHASE)
            font = PHASE_FONTS.get(phase, FONT_PHASE)
            _style_row(ws, r_idx, fill, font, n_cols)
        elif isinstance(phase, str) and phase in PHASE_FILLS:
            fill = PHASE_FILLS.get(phase, FILL_PHASE)
            for c in range(1, n_cols + 1):
                ws.cell(row=r_idx, column=c).fill = fill

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_classificacao(ws, rows: list[list]) -> None:
    n_cols = 8
    col_widths = [5, 22, 10, 10, 10, 8, 7, 12]
    ws.row_dimensions[1].height = 20

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center")

    _style_row(ws, 1, FILL_TITLE, FONT_TITLE, n_cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _style_row(ws, 2, FILL_HEADER, FONT_HEADER, n_cols)

    # Color rows by position
    FILL_1ST = PatternFill("solid", fgColor="C6EFCE")
    FILL_2ND = PatternFill("solid", fgColor="DEEAF1")
    FILL_3RD = PatternFill("solid", fgColor="FCE4D6")

    for r_idx, row in enumerate(rows, 1):
        if r_idx <= 2:
            continue
        pos = row[0]
        if pos == 1:
            fill = FILL_1ST
        elif pos == 2:
            fill = FILL_2ND
        elif pos == 3:
            fill = FILL_3RD
        else:
            continue
        for c in range(1, n_cols + 1):
            ws.cell(row=r_idx, column=c).fill = fill

    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 65)
    print("  generate_previsoes_v2.py — Copa 2026 Excel v2")
    print("  Modelo: XGBoost v6 (decay temporal ξ=0.001 + dados abr-jun 2026)")
    print("=" * 65)

    print("\n  Carregando modelo e extraindo features...")
    model, team_feats = train_model()

    all_teams = [t for ts in GROUPS.values() for t in ts]
    missing = [t for t in all_teams if t not in team_feats]
    if missing:
        print(f"  AVISO: {len(missing)} times sem features — usando defaults")
        for t in missing:
            team_feats[t] = {"gls_decay": 1.0, "shots_on_goal_decay": 3.0,
                             "win_rate_decay": 0.35, "saves_decay": 1.8}

    print("  Pré-computando lambdas...")
    lam_cache = precompute_lambdas(model, team_feats, all_teams)

    print("  Carregando ELO...")
    elo = load_elo(team_feats)

    print("  Rodando simulação determinística dos grupos...")
    np.random.seed(42)
    det_standings = simulate_all_groups(lam_cache, stochastic=False)

    print("  Rodando bracket determinístico...")
    det_ko = simulate_knockouts(det_standings, lam_cache, stochastic=False)

    champion = pt(det_ko["winner"][104])
    finalist = pt(det_ko["teams"][104][1])
    print(f"  Bracket det. → Final: {finalist} vs {champion}  ← Campeão: {champion}")

    print("  Construindo abas...")
    grupos_rows     = build_grupos_rows(lam_cache, elo)
    matamata_rows   = build_matamata_rows(lam_cache, elo, det_standings, det_ko)
    classif_rows    = build_classificacao_rows(lam_cache)

    print(f"  Fase de Grupos: {len(grupos_rows)} linhas")
    print(f"  Mata-Mata:      {len(matamata_rows)} linhas")
    print(f"  Classificação:  {len(classif_rows)} linhas")

    print("  Gerando Excel...")
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Fase de Grupos"
    write_grupos(ws1, grupos_rows)

    ws2 = wb.create_sheet("Mata-Mata")
    write_matamata(ws2, matamata_rows)

    ws3 = wb.create_sheet("Classificação Esperada")
    write_classificacao(ws3, classif_rows)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"\n  Salvo: {OUT_XLSX}")
    print("=" * 65)


if __name__ == "__main__":
    main()
