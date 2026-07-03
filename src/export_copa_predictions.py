"""
export_copa_predictions.py — Excel com previsões de todos os jogos da Copa 2026.

Abas:
  1. Fase de Grupos  — 72 jogos
  2. Mata-Mata       — R32 → R16 → QF → SF → 3º lugar → Final (47 jogos)
  3. Classificação Esperada
"""

import pickle
import sys
from itertools import combinations
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import poisson as sci_poisson

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

OUT_PKL   = Path("outputs/xgboost_v4.pkl")
INPUT_CSV = Path("data/processed/model_dataset.csv")
ELO_CSV   = Path("data/raw/elo_history.csv")
OUTPUT    = Path("outputs/copa2026_previsoes.xlsx")
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

BASE_FEATURES = ["gls_avg5","shots_on_goal_avg5","win_rate_avg5","opp_saves_avg5","home_advantage"]
RATING_FEATURES = [
    "delta_rating_avg5","gls_avg5_vs_forte","gls_avg5_vs_fraco","shots_avg5_vs_forte",
    "gls_ponderado_avg5","elo_diff_avg5","rating_titulares_avg5",
]
CONF_DUMMIES = ["conf_AFC","conf_CAF","conf_CONCACAF","conf_CONMEBOL","conf_FRIENDLY","conf_OFC","conf_UEFA"]
FEATURES = BASE_FEATURES + RATING_FEATURES + CONF_DUMMIES + ["match_type_wcq"]

GROUPS = {
    "A": ["Mexico",        "South Korea",  "South Africa",     "Czech Republic"],
    "B": ["Canada",        "Switzerland",  "Qatar",            "Bosnia and Herzegovina"],
    "C": ["Brazil",        "Morocco",      "Haiti",            "Scotland"],
    "D": ["United States", "Paraguay",     "Australia",        "Turkey"],
    "E": ["Germany",       "Ivory Coast",  "Ecuador",          "Curacao"],
    "F": ["Netherlands",   "Sweden",       "Tunisia",          "Japan"],
    "G": ["Belgium",       "Iran",         "New Zealand",      "Egypt"],
    "H": ["Spain",         "Saudi Arabia", "Uruguay",          "Cape Verde"],
    "I": ["France",        "Senegal",      "Iraq",             "Norway"],
    "J": ["Argentina",     "Algeria",      "Austria",          "Jordan"],
    "K": ["Portugal",      "DR Congo",     "Uzbekistan",       "Colombia"],
    "L": ["England",       "Croatia",      "Ghana",            "Panama"],
}

PT_BR = {
    "Mexico": "México", "South Korea": "Coreia do Sul", "South Africa": "África do Sul",
    "Czech Republic": "Rep. Tcheca", "Canada": "Canadá", "Switzerland": "Suíça",
    "Qatar": "Catar", "Bosnia and Herzegovina": "Bósnia-Herz.", "Brazil": "Brasil",
    "Morocco": "Marrocos", "Haiti": "Haiti", "Scotland": "Escócia",
    "United States": "EUA", "Paraguay": "Paraguai", "Australia": "Austrália",
    "Turkey": "Turquia", "Germany": "Alemanha", "Ivory Coast": "Costa do Marfim",
    "Ecuador": "Equador", "Curacao": "Curaçao", "Netherlands": "Holanda",
    "Sweden": "Suécia", "Tunisia": "Tunísia", "Japan": "Japão",
    "Belgium": "Bélgica", "Iran": "Irã", "New Zealand": "Nova Zelândia",
    "Egypt": "Egito", "Spain": "Espanha", "Saudi Arabia": "Arábia Saudita",
    "Uruguay": "Uruguai", "Cape Verde": "Cabo Verde", "France": "França",
    "Senegal": "Senegal", "Iraq": "Iraque", "Norway": "Noruega",
    "Argentina": "Argentina", "Algeria": "Argélia", "Austria": "Áustria",
    "Jordan": "Jordânia", "Portugal": "Portugal", "DR Congo": "RD Congo",
    "Uzbekistan": "Uzbequistão", "Colombia": "Colômbia", "England": "Inglaterra",
    "Croatia": "Croácia", "Ghana": "Gana", "Panama": "Panamá",
}

# ─── Bracket R32 (chaveamento oficial Copa 2026) ──────────────────────────────
# (match_id, slot1, slot2)  — slot: "1X"=1º grp X, "2X"=2º grp X, "3"=3º qualif.
R32 = [
    (73,  "2A", "2B"),
    (74,  "1E", "3"),   # 3º de A/B/C/D/F
    (75,  "1F", "2C"),
    (76,  "1C", "2F"),
    (77,  "1I", "3"),   # 3º de C/D/F/G/H
    (78,  "2E", "2I"),
    (79,  "1A", "3"),   # 3º de C/E/F/H/I
    (80,  "1L", "3"),   # 3º de E/H/I/J/K
    (81,  "1D", "3"),   # 3º de B/E/F/I/J
    (82,  "1G", "3"),   # 3º de A/E/H/I/J
    (83,  "2K", "2L"),
    (84,  "1H", "2J"),
    (85,  "1B", "3"),   # 3º de E/F/G/I/J
    (86,  "1J", "2H"),
    (87,  "1K", "3"),   # 3º de D/E/I/J/L
    (88,  "2D", "2G"),
]

THIRD_ELIGIBLE = {
    74: ["A","B","C","D","F"],
    77: ["C","D","F","G","H"],
    79: ["C","E","F","H","I"],
    80: ["E","H","I","J","K"],
    81: ["B","E","F","I","J"],
    82: ["A","E","H","I","J"],
    85: ["E","F","G","I","J"],
    87: ["D","E","I","J","L"],
}

# R32 → R16 mapping
R32_TO_R16 = {73:90,75:90, 74:89,77:89, 76:91,78:91, 79:92,80:92,
               83:93,84:93, 81:94,82:94, 86:95,88:95, 85:96,87:96}

# Fases eliminatórias na ordem
KO_ROUNDS = [
    ("Oitavas (R32)",    "R32",  R32),
    ("Quartas (R16)",    "R16",  [(89,(74,77)),(90,(73,75)),(91,(76,78)),(92,(79,80)),
                                  (93,(83,84)),(94,(81,82)),(95,(86,88)),(96,(85,87))]),
    ("Semifinais (QF)",  "QF",   [(97,(89,90)),(98,(93,94)),(99,(91,92)),(100,(95,96))]),
    ("Semifinais (SF)",  "SF",   [(101,(97,98)),(102,(99,100))]),
    ("3º Lugar / Final","3F",   [(103,(101,102)),(104,(101,102))]),  # especial
]

ROUND_BG = {
    "R32": "D6E4F0", "R16": "D5E8D4", "QF": "FFF2CC",
    "SF": "FCE4D6", "3F": "F8CECC",
}
ROUND_HEADER_BG = {
    "R32": "2E75B6", "R16": "548235", "QF": "9C6500",
    "SF": "C55A11", "3F": "C00000",
}
ROUND_LABEL = {
    "R32": "OITAVAS DE FINAL", "R16": "QUARTAS DE FINAL",
    "QF":  "SEMIFINAIS", "SF": "FINAL FOUR",
    "3F":  "3º LUGAR  &  FINAL",
}

# ─── Estilos ──────────────────────────────────────────────────────────────────
THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT   = Alignment(horizontal="left",   vertical="center")

CONF_BG = {
    "A": "D6E4F0", "B": "D5E8D4", "C": "FFE6CC", "D": "E1D5E7",
    "E": "DAE8FC", "F": "D5E8D4", "G": "FFF2CC", "H": "F8CECC",
    "I": "D6E4F0", "J": "E1D5E7", "K": "FFE6CC", "L": "DAE8FC",
}
GROUP_HEADER_BG = {
    "A": "2E75B6", "B": "548235", "C": "C55A11", "D": "7030A0",
    "E": "1F6DC6", "F": "375623", "G": "9C6500", "H": "C00000",
    "I": "1F4E79", "J": "4B0082", "K": "833C00", "L": "203864",
}


# ─── Helpers de célula ────────────────────────────────────────────────────────
def cell(ws, r, c, value="", fill_hex=None, bold=False, align=CENTER,
         font_color="000000", size=10, border=True):
    cl = ws.cell(row=r, column=c, value=value)
    if fill_hex:
        cl.fill = PatternFill("solid", fgColor=fill_hex)
    cl.font  = Font(bold=bold, color=font_color, size=size)
    cl.alignment = align
    if border:
        cl.border = BORDER
    return cl


def section_header(ws, row_idx, ncols, text, fill_hex, font_color="FFFFFF"):
    ws.merge_cells(f"A{row_idx}:{get_column_letter(ncols)}{row_idx}")
    c = ws[f"A{row_idx}"]
    c.value = f"  {text}"
    c.fill  = PatternFill("solid", fgColor=fill_hex)
    c.font  = Font(bold=True, color=font_color, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = BORDER
    ws.row_dimensions[row_idx].height = 15


def write_col_headers(ws, row_idx, headers, widths, fill_hex="2E75B6"):
    for j, h in enumerate(headers, 1):
        cell(ws, row_idx, j, h, fill_hex=fill_hex, bold=True,
             font_color="FFFFFF", size=10)
    ws.row_dimensions[row_idx].height = 16
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_game_row(ws, row_idx, rec, bg, ncols=15):
    pv, pe, pd_ = rec["pv"], rec["pe"], rec["pd"]
    if pv > pd_:   pv_fc, pd_fc = "375623", "595959"
    elif pd_ > pv: pv_fc, pd_fc = "595959", "833C00"
    else:          pv_fc, pd_fc = "595959", "595959"

    vals = [
        (rec.get("fase",""),     bg, True,  CENTER, "1F3864"),
        (rec["t1_pt"],           bg, False, LEFT,   "000000"),
        (rec["elo1"],            bg, False, CENTER, "595959"),
        (rec["lam1"],            bg, True,  CENTER, "1F3864"),
        ("×",                    bg, False, CENTER, "BFBFBF"),
        (rec["lam2"],            bg, True,  CENTER, "1F3864"),
        (rec["elo2"],            bg, False, CENTER, "595959"),
        (rec["t2_pt"],           bg, False, LEFT,   "000000"),
        (f"{pv}%",               bg, True,  CENTER, pv_fc),
        (f"{pe}%",               bg, False, CENTER, "595959"),
        (f"{pd_}%",              bg, True,  CENTER, pd_fc),
        (rec["fav"],             bg, True,  CENTER, "1F3864"),
        (rec["s1"],              bg, False, CENTER, "000000"),
        (rec["s2"],              bg, False, CENTER, "595959"),
        (rec["s3"],              bg, False, CENTER, "BFBFBF"),
    ]
    for j, (val, bh, bold, aln, fc) in enumerate(vals, 1):
        cell(ws, row_idx, j, val, fill_hex=bh, bold=bold, align=aln, font_color=fc)
    ws.row_dimensions[row_idx].height = 14


# ─── Model helpers ────────────────────────────────────────────────────────────
def load_team_features():
    wide = pd.read_csv(INPUT_CSV, parse_dates=["date"])
    rows = []
    for _, r in wide.iterrows():
        for side, opp in [("home","away"),("away","home")]:
            row = {
                "date": r["date"], "team": r[f"{side}_team"],
                "gls_avg5":           r.get(f"{side}_gls_avg5", np.nan),
                "shots_on_goal_avg5": r.get(f"{side}_shots_on_goal_avg5", np.nan),
                "win_rate_avg5":      r.get(f"{side}_win_rate_avg5", np.nan),
                "saves_avg5":         r.get(f"{side}_saves_avg5", np.nan),
            }
            for col in RATING_FEATURES:
                row[col] = r.get(f"{side}_{col}", np.nan)
            rows.append(row)
    long = pd.DataFrame(rows).sort_values("date")
    tf = {}
    for _, row in long.iterrows():
        e = {
            "gls_avg5":           float(row["gls_avg5"]) if pd.notna(row["gls_avg5"]) else 1.2,
            "shots_on_goal_avg5": float(row["shots_on_goal_avg5"]) if pd.notna(row["shots_on_goal_avg5"]) else 3.5,
            "win_rate_avg5":      float(row["win_rate_avg5"]) if pd.notna(row["win_rate_avg5"]) else 0.4,
            "saves_avg5":         float(row["saves_avg5"]) if pd.notna(row["saves_avg5"]) else 1.8,
        }
        for col in RATING_FEATURES:
            v = row.get(col, np.nan)
            e[col] = float(v) if pd.notna(v) else np.nan
        tf[row["team"]] = e
    if ELO_CSV.exists():
        elo_hist = pd.read_csv(ELO_CSV, parse_dates=["date"])
        cur_elo = elo_hist.sort_values("date").groupby("team")["elo_after"].last().to_dict()
        for t, v in cur_elo.items():
            if t in tf: tf[t]["_elo_current"] = float(v)
            else:       tf[t] = {"_elo_current": float(v)}
    return tf


def get_lambda(model, t1, t2, tf):
    f1, f2 = tf.get(t1, {}), tf.get(t2, {})
    row = {
        "gls_avg5":           f1.get("gls_avg5", 1.2),
        "shots_on_goal_avg5": f1.get("shots_on_goal_avg5", 3.5),
        "win_rate_avg5":      f1.get("win_rate_avg5", 0.4),
        "opp_saves_avg5":     f2.get("saves_avg5", 2.0),
        "home_advantage": 0, "match_type_wcq": 0,
        **{c: 0 for c in CONF_DUMMIES},
    }
    for col in RATING_FEATURES:
        if col == "elo_diff_avg5":
            e1, e2 = f1.get("_elo_current", np.nan), f2.get("_elo_current", np.nan)
            row[col] = (float(e1)-float(e2)) if (pd.notna(e1) and pd.notna(e2)) else f1.get(col, np.nan)
        else:
            row[col] = f1.get(col, np.nan)
    return float(model.predict(pd.DataFrame([row])[FEATURES])[0])


def match_stats(lam1, lam2, max_g=8):
    probs = {(g1,g2): sci_poisson.pmf(g1,lam1)*sci_poisson.pmf(g2,lam2)
             for g1 in range(max_g+1) for g2 in range(max_g+1)}
    pv  = sum(p for (g1,g2),p in probs.items() if g1>g2)
    pe  = sum(p for (g1,g2),p in probs.items() if g1==g2)
    pd_ = sum(p for (g1,g2),p in probs.items() if g1<g2)
    top3 = sorted(probs.items(), key=lambda x:-x[1])[:3]
    return pv, pe, pd_, top3


def make_record(model, tf, t1, t2, fase=""):
    l1 = get_lambda(model, t1, t2, tf)
    l2 = get_lambda(model, t2, t1, tf)
    pv, pe, pd_, top3 = match_stats(l1, l2)
    elo1 = tf.get(t1, {}).get("_elo_current", np.nan)
    elo2 = tf.get(t2, {}).get("_elo_current", np.nan)
    if pv > pd_:   fav = PT_BR.get(t1, t1)
    elif pd_ > pv: fav = PT_BR.get(t2, t2)
    else:          fav = "Eq."
    scores = [f"{g1}–{g2} ({p*100:.1f}%)" for (g1,g2),p in top3]
    # win_prob: P(vencer inclusive no mata-mata — pen.)
    pko1 = pv + pe*0.5  # empate → 50% penaltis
    return {
        "fase": fase,
        "t1": t1, "t1_pt": PT_BR.get(t1, t1),
        "t2": t2, "t2_pt": PT_BR.get(t2, t2),
        "lam1": round(l1, 3), "lam2": round(l2, 3),
        "pv": round(pv*100,1), "pe": round(pe*100,1), "pd": round(pd_*100,1),
        "pko1": pko1,   # prob de avançar (t1), inclui pênaltis
        "elo1": round(elo1) if pd.notna(elo1) else "-",
        "elo2": round(elo2) if pd.notna(elo2) else "-",
        "fav": fav,
        "s1": scores[0] if len(scores)>0 else "",
        "s2": scores[1] if len(scores)>1 else "",
        "s3": scores[2] if len(scores)>2 else "",
        "winner": t1 if pko1 >= 0.5 else t2,
    }


# ─── Fase de Grupos ───────────────────────────────────────────────────────────
def build_group_records(model, tf):
    recs = []
    for g, teams in GROUPS.items():
        for t1, t2 in combinations(teams, 2):
            r = make_record(model, tf, t1, t2, fase=g)
            r["grupo"] = g
            recs.append(r)
    return recs


# ─── Classificação esperada (pts esperados) ───────────────────────────────────
def expected_standings(group_records) -> dict[str, list[dict]]:
    """Retorna classificação por grupo ordenada por pts esperados."""
    pts: dict[str, float] = {}
    gf:  dict[str, float] = {}
    ga:  dict[str, float] = {}
    team_group: dict[str, str] = {}

    for rec in group_records:
        t1, t2 = rec["t1"], rec["t2"]
        pv, pe, pd_ = rec["pv"]/100, rec["pe"]/100, rec["pd"]/100
        for t in [t1, t2]:
            pts.setdefault(t, 0.0)
            gf.setdefault(t, 0.0)
            ga.setdefault(t, 0.0)
            team_group[t] = rec["grupo"]
        pts[t1] += pv*3 + pe; pts[t2] += pd_*3 + pe
        gf[t1]  += rec["lam1"]; gf[t2] += rec["lam2"]
        ga[t1]  += rec["lam2"]; ga[t2] += rec["lam1"]

    by_group: dict[str, list] = {}
    for t in pts:
        g = team_group[t]
        by_group.setdefault(g, [])
        by_group[g].append({
            "team": t, "group": g,
            "pts": pts[t], "gf": gf[t], "ga": ga[t],
            "gd": gf[t]-ga[t],
        })

    standings: dict[str, list] = {}
    for g, rows in by_group.items():
        standings[g] = sorted(rows, key=lambda x: (-x["pts"], -x["gd"], -x["gf"]))
        for i, r in enumerate(standings[g]):
            r["pos"] = i + 1
    return standings


# ─── 3ºs colocados ───────────────────────────────────────────────────────────
def _augment(mid, available, assignment, visited):
    for group in available.get(mid, []):
        am = next((m for m, g in assignment.items() if g == group), None)
        if am is None:
            assignment[mid] = group; return True
        if am not in visited:
            visited.add(am)
            if _augment(am, available, assignment, visited):
                assignment[mid] = group; return True
    return False


def assign_thirds(qualifying_thirds):
    q_groups = {t["group"] for t in qualifying_thirds}
    available = {mid: [g for g in gs if g in q_groups]
                 for mid, gs in THIRD_ELIGIBLE.items()}
    match_order = sorted(available, key=lambda m: len(available[m]))
    assignment = {}
    for mid in match_order:
        _augment(mid, available, assignment, {mid})
    return assignment


# ─── Resolução do bracket eliminatório ───────────────────────────────────────
def build_knockout_records(model, tf, standings):
    """
    Resolve deterministicamente cada fase, usando pko1 >= 0.5 como critério.
    Retorna lista de records por fase + dict winner[match_id].
    """
    qualifying_thirds = []
    for g, rows in standings.items():
        r = rows[2]  # 3º colocado (índice 2)
        qualifying_thirds.append({"group": g, "team": r["team"],
                                   "pts": r["pts"], "gd": r["gd"], "gf": r["gf"]})
    qualifying_thirds.sort(key=lambda x: (-x["pts"], -x["gd"], -x["gf"]))
    qualifying_thirds = qualifying_thirds[:8]
    thirds_by_group = {t["group"]: t for t in qualifying_thirds}
    third_assignment = assign_thirds(qualifying_thirds)

    winner: dict[int, str] = {}
    loser:  dict[int, str] = {}
    ko_records: list[dict] = []

    def resolve(slot, mid):
        if slot.startswith("1"): return standings[slot[1]][0]["team"]
        if slot.startswith("2"): return standings[slot[1]][1]["team"]
        if slot == "3":
            g = third_assignment.get(mid)
            return thirds_by_group[g]["team"] if g else "TBD"
        return "TBD"

    # ── R32 ──
    for mid, s1, s2 in R32:
        t1, t2 = resolve(s1, mid), resolve(s2, mid)
        rec = make_record(model, tf, t1, t2, fase="R32")
        rec["match_id"] = mid
        ko_records.append(rec)
        winner[mid] = rec["winner"]
        loser[mid]  = t2 if rec["winner"] == t1 else t1

    # ── R16 ──
    r16 = [(89,(74,77)),(90,(73,75)),(91,(76,78)),(92,(79,80)),
           (93,(83,84)),(94,(81,82)),(95,(86,88)),(96,(85,87))]
    for mid, (m1, m2) in r16:
        t1, t2 = winner[m1], winner[m2]
        rec = make_record(model, tf, t1, t2, fase="R16")
        rec["match_id"] = mid
        ko_records.append(rec)
        winner[mid] = rec["winner"]
        loser[mid]  = t2 if rec["winner"] == t1 else t1

    # ── QF ──
    for mid, (m1, m2) in [(97,(89,90)),(98,(93,94)),(99,(91,92)),(100,(95,96))]:
        t1, t2 = winner[m1], winner[m2]
        rec = make_record(model, tf, t1, t2, fase="QF")
        rec["match_id"] = mid
        ko_records.append(rec)
        winner[mid] = rec["winner"]
        loser[mid]  = t2 if rec["winner"] == t1 else t1

    # ── SF ──
    for mid, (m1, m2) in [(101,(97,98)),(102,(99,100))]:
        t1, t2 = winner[m1], winner[m2]
        rec = make_record(model, tf, t1, t2, fase="SF")
        rec["match_id"] = mid
        ko_records.append(rec)
        winner[mid] = rec["winner"]
        loser[mid]  = t2 if rec["winner"] == t1 else t1

    # ── 3º lugar ──
    t1, t2 = loser[101], loser[102]
    rec = make_record(model, tf, t1, t2, fase="3P")
    rec["match_id"] = 103
    rec["fase_label"] = "3º Lugar"
    ko_records.append(rec)

    # ── Final ──
    t1, t2 = winner[101], winner[102]
    rec = make_record(model, tf, t1, t2, fase="FIN")
    rec["match_id"] = 104
    rec["fase_label"] = "FINAL"
    ko_records.append(rec)
    winner[104] = rec["winner"]

    print(f"  Campeão previsto: {winner[104]}")
    return ko_records, winner, qualifying_thirds


# ─── Escrita das abas ─────────────────────────────────────────────────────────
KO_HEADERS = [
    "Fase", "Time 1", "ELO", "λ1", "×", "λ2", "ELO", "Time 2",
    "V%", "E%", "D%", "Favorito", "1º Placar", "2º Placar", "3º Placar",
]
KO_WIDTHS = [10, 20, 6, 7, 3, 7, 6, 20, 7, 7, 7, 18, 16, 16, 16]

GRP_HEADERS = [
    "Grp", "Time 1", "ELO", "λ1", "×", "λ2", "ELO", "Time 2",
    "V%", "E%", "D%", "Favorito", "1º Placar", "2º Placar", "3º Placar",
]
GRP_WIDTHS = [5, 18, 6, 7, 3, 7, 6, 18, 7, 7, 7, 16, 15, 15, 15]

FASE_LABEL = {
    "R32": "OITAVAS", "R16": "QUARTAS",
    "QF":  "SEMI", "SF": "SEMI-FINAL",
    "3P":  "3º LUGAR", "FIN": "FINAL",
}
FASE_BG = {
    "R32": "D6E4F0", "R16": "D5E8D4",
    "QF": "FFF2CC", "SF": "FCE4D6",
    "3P": "EDEDED", "FIN": "FFE6CC",
}
FASE_HEADER_BG = {
    "R32": "2E75B6", "R16": "548235",
    "QF": "9C6500", "SF": "C55A11",
    "3P": "595959", "FIN": "C00000",
}


def write_sheet_grupos(wb, group_records):
    ws = wb.active
    ws.title = "Fase de Grupos"

    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = "COPA DO MUNDO 2026 — PREVISÃO DE PLACARES  (modelo XGBoost v4)"
    c.fill  = PatternFill("solid", fgColor="1F3864")
    c.font  = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    write_col_headers(ws, 2, GRP_HEADERS, GRP_WIDTHS)

    prev_group = None
    row_idx = 3
    for rec in group_records:
        g = rec["grupo"]
        if g != prev_group:
            section_header(ws, row_idx, 15,
                           f"GRUPO {g}  —  {'  ·  '.join(PT_BR.get(t,t) for t in GROUPS[g])}",
                           GROUP_HEADER_BG[g])
            row_idx += 1
            prev_group = g
        write_game_row(ws, row_idx, rec, CONF_BG[g])
        row_idx += 1

    ws.freeze_panes = "A3"
    print(f"  Aba 'Fase de Grupos': {len(group_records)} jogos")


def write_sheet_knockout(wb, ko_records):
    ws = wb.create_sheet("Mata-Mata")

    ws.merge_cells("A1:O1")
    c = ws["A1"]
    c.value = "COPA DO MUNDO 2026 — MATA-MATA DETERMINÍSTICO  (modelo XGBoost v4)"
    c.fill  = PatternFill("solid", fgColor="1F3864")
    c.font  = Font(bold=True, color="FFFFFF", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    write_col_headers(ws, 2, KO_HEADERS, KO_WIDTHS)

    prev_fase = None
    row_idx = 3
    game_num = {f: 1 for f in FASE_BG}

    for rec in ko_records:
        fase = rec["fase"]
        if fase != prev_fase:
            label = FASE_LABEL.get(fase, fase)
            section_header(ws, row_idx, 15, label, FASE_HEADER_BG.get(fase, "1F3864"))
            row_idx += 1
            prev_fase = fase

        # Label da fase para a coluna A
        display_fase = rec.get("fase_label", FASE_LABEL.get(fase, fase))
        rec_copy = dict(rec)
        rec_copy["fase"] = display_fase

        write_game_row(ws, row_idx, rec_copy, FASE_BG.get(fase, "FFFFFF"))

        # Destaque para o vencedor (negrito na célula do time vencedor)
        w = rec["winner"]
        col = 2 if rec["t1"] == w else 8
        c = ws.cell(row=row_idx, column=col)
        c.font = Font(bold=True, color="375623", size=10)

        row_idx += 1

    ws.freeze_panes = "A3"
    print(f"  Aba 'Mata-Mata': {len(ko_records)} jogos")


def write_sheet_standings(wb, group_records, standings, qualifying_thirds):
    ws = wb.create_sheet("Classificação Esperada")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "COPA DO MUNDO 2026 — CLASSIFICAÇÃO ESPERADA (pontos × probabilidades)"
    c.fill  = PatternFill("solid", fgColor="1F3864")
    c.font  = Font(bold=True, color="FFFFFF", size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["Pos", "Time", "Pts Esp.", "GF Esp.", "GA Esp.", "Saldo", "Grupo", "Obs."]
    widths   = [5, 22, 10, 10, 10, 10, 7, 14]
    for j, h in enumerate(headers, 1):
        c2 = ws.cell(row=2, column=j, value=h)
        c2.fill = PatternFill("solid", fgColor="2E75B6")
        c2.font = Font(bold=True, color="FFFFFF", size=10)
        c2.alignment = CENTER
        c2.border = BORDER
    ws.row_dimensions[2].height = 16
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thirds_set = {t["group"] for t in qualifying_thirds}
    row_idx = 3
    for g in sorted(standings.keys()):
        rows = standings[g]
        bg = CONF_BG.get(g, "FFFFFF")
        for r in rows:
            pos = r["pos"]
            if pos <= 2:   obs, bg_use = "Classifica", bg
            elif g in thirds_set and pos == 3: obs, bg_use = "3º qualif.", "FFFDE7"
            else: obs, bg_use = "", "F5F5F5"
            fc = "000000" if pos <= 2 else "888888"
            vals = [pos, PT_BR.get(r["team"], r["team"]),
                    round(r["pts"],2), round(r["gf"],2), round(r["ga"],2),
                    round(r["gd"],2), g, obs]
            for j, v in enumerate(vals, 1):
                c3 = ws.cell(row=row_idx, column=j, value=v)
                c3.fill = PatternFill("solid", fgColor=bg_use)
                c3.font = Font(bold=(pos<=2), color=fc, size=10)
                c3.alignment = CENTER if j != 2 else LEFT
                c3.border = BORDER
            ws.row_dimensions[row_idx].height = 14
            row_idx += 1
        row_idx += 1  # separador entre grupos

    ws.freeze_panes = "A3"
    print(f"  Aba 'Classificação Esperada': {sum(len(v) for v in standings.values())} times")


# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 55)
    print("  export_copa_predictions.py — Copa 2026 Completa")
    print("=" * 55)

    with open(OUT_PKL, "rb") as f:
        model = pickle.load(f)
    print("  Modelo v4 carregado.")

    tf = load_team_features()
    print(f"  Features de {len(tf)} seleções.")

    # Grupos
    group_records = build_group_records(model, tf)
    print(f"  {len(group_records)} jogos de grupo calculados.")

    # Classificação esperada
    standings = expected_standings(group_records)

    # Mata-mata
    ko_records, winner, qualifying_thirds = build_knockout_records(model, tf, standings)
    print(f"  {len(ko_records)} jogos de mata-mata calculados.")

    # Excel
    wb = openpyxl.Workbook()
    write_sheet_grupos(wb, group_records)
    write_sheet_knockout(wb, ko_records)
    write_sheet_standings(wb, group_records, standings, qualifying_thirds)

    wb.save(OUTPUT)
    total = len(group_records) + len(ko_records)
    print(f"\n  → Salvo: {OUTPUT}  ({total} jogos no total)")
    print(f"  Abas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
