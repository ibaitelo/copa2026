#!/usr/bin/env python3
"""
generate_simulacao_v11.py — Copa 2026 Simulação v11
Modelo: xgboost_v10_clean.pkl (sem host_factor no modelo;
        multiplicador pós-predição: λ_final = λ_base × (1 + bonus))

Saída: outputs/copa2026_simulacao_v11.xlsx
  Aba 1 — Fase de Grupos (determinístico)
  Aba 2 — Monte Carlo (10 000 iterações)
  Aba 3 — Diagnóstico opp_saves_decay por time
"""
from __future__ import annotations

import sys
sys.path.insert(0, "src")

import pickle
import random
import numpy as np
import pandas as pd
from itertools import combinations
from pathlib import Path
from xgboost import XGBRegressor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from simulate_copa2026 import (
    GROUPS, R32, MATCH_VENUES, GROUP_VENUE_COUNTRY, TEAM_TO_GROUP,
    HOST_TEAM_COUNTRY, assign_thirds, simulate_game_from_lambdas,
)

# ── Paths ─────────────────────────────────────────────────────────────────────
MODEL_PKL       = Path("outputs/xgboost_v10_clean.pkl")
MODEL_DATASET   = Path("data/processed/model_dataset.csv")
HOST_FACTOR_CSV = Path("data/external/host_factor_copa2026.csv")
ELO_CSV         = Path("data/raw/elo_history.csv")
OUT_XLSX        = Path("outputs/copa2026_simulacao_v11.xlsx")
N_SIM           = 10_000
SEED            = 42

# ── Features (espelha validate_corrections.py) ────────────────────────────────
CONF_DUMMIES = ["conf_AFC","conf_CAF","conf_CONCACAF","conf_CONMEBOL","conf_OFC","conf_UEFA"]
RATING_FEATURES = [
    "elo_diff_decay","delta_rating_decay","gls_ponderado_decay",
    "gls_decay_vs_forte","gls_decay_vs_fraco","margem_gols_decay",
]
FEATURES = (["opp_saves_decay","shots_on_goal_decay","gols_sofridos_decay"]
            + RATING_FEATURES + CONF_DUMMIES)  # 15 features

ALL_COPA_TEAMS = [t for ts in GROUPS.values() for t in ts]  # 48 times

# ── Multiplicador pós-predição (Correção 2) ───────────────────────────────────
def load_cohost_factors() -> dict[str, float]:
    df = pd.read_csv(HOST_FACTOR_CSV)
    return {r["team"]: float(r["host_factor_cohost"]) for _, r in df.iterrows()}

def get_home_bonus(team: str, venue: str, hf_cohost: dict) -> float:
    tc = HOST_TEAM_COUNTRY.get(team)
    return hf_cohost.get(team, 0.0) * 0.5 if (tc and tc == venue) else 0.0

# ── Carregar modelo ───────────────────────────────────────────────────────────
def load_model() -> XGBRegressor:
    with open(MODEL_PKL, "rb") as f:
        return pickle.load(f)

# ── Extrair team_feats do dataset ─────────────────────────────────────────────
def extract_team_feats(wide: pd.DataFrame) -> dict[str, dict]:
    rows = []
    for _, r in wide.iterrows():
        for side, opp in [("home","away"), ("away","home")]:
            row: dict = {
                "date":                r["date"],
                "team":                r[f"{side}_team"],
                "saves_decay":         float(r.get(f"{side}_saves_decay", 2.5)),
                "shots_on_goal_decay": float(r[f"{side}_shots_on_goal_decay"]),
                "gols_sofridos_decay": float(r.get(f"{side}_gols_sofridos_decay", np.nan)),
            }
            for col in RATING_FEATURES:
                raw = r.get(f"{side}_{col}", np.nan)
                row[col] = float(raw) if pd.notna(raw) else np.nan
            rows.append(row)

    long = pd.DataFrame(rows).sort_values("date").reset_index(drop=True)
    team_feats: dict[str, dict] = {}
    for _, row in long.iterrows():
        entry: dict = {k: row[k] for k in ["saves_decay","shots_on_goal_decay","gols_sofridos_decay"]}
        for col in RATING_FEATURES:
            v = row.get(col, np.nan)
            entry[col] = float(v) if pd.notna(v) else np.nan
        team_feats[row["team"]] = entry

    if ELO_CSV.exists():
        elo = pd.read_csv(ELO_CSV, parse_dates=["date"])
        for team, val in elo.sort_values("date").groupby("team")["elo_after"].last().items():
            team_feats.setdefault(team, {})["_elo_current"] = float(val)
    return team_feats

# ── Lambda cache (único, sem host_factor) ─────────────────────────────────────
def precompute_lambdas(model: XGBRegressor,
                       team_feats: dict,
                       teams: list[str]) -> dict[tuple[str,str], float]:
    rows, pairs = [], []
    for t1 in teams:
        for t2 in teams:
            if t1 == t2:
                continue
            f1 = team_feats.get(t1, {})
            f2 = team_feats.get(t2, {})
            row: dict = {
                "opp_saves_decay":     f2.get("saves_decay", 2.5),
                "shots_on_goal_decay": f1.get("shots_on_goal_decay", 3.5),
                "gols_sofridos_decay": f1.get("gols_sofridos_decay", np.nan),
                **{c: 0 for c in CONF_DUMMIES},
            }
            for col in RATING_FEATURES:
                if col == "elo_diff_decay":
                    e1, e2 = f1.get("_elo_current", np.nan), f2.get("_elo_current", np.nan)
                    row[col] = (float(e1)-float(e2)
                                if pd.notna(e1) and pd.notna(e2)
                                else f1.get(col, np.nan))
                else:
                    row[col] = f1.get(col, np.nan)
            rows.append(row)
            pairs.append((t1, t2))

    preds = model.predict(pd.DataFrame(rows)[FEATURES])
    return {pair: float(p) for pair, p in zip(pairs, preds)}

def get_match_lambdas(t1: str, t2: str, venue: str,
                      lam: dict, hf_cohost: dict) -> tuple[float, float]:
    lh = lam.get((t1,t2), 1.2) * (1 + get_home_bonus(t1, venue, hf_cohost))
    la = lam.get((t2,t1), 1.0) * (1 + get_home_bonus(t2, venue, hf_cohost))
    return lh, la

# ── Fase de grupos ────────────────────────────────────────────────────────────
def simulate_group(gid: str, teams: list[str],
                   lam: dict, hf_cohost: dict,
                   stochastic: bool) -> pd.DataFrame:
    venue = GROUP_VENUE_COUNTRY[gid]
    stats = {t: {"pts":0,"gf":0,"ga":0,"gd":0} for t in teams}
    for t1, t2 in combinations(teams, 2):
        lh, la = get_match_lambdas(t1, t2, venue, lam, hf_cohost)
        g1, g2 = simulate_game_from_lambdas(lh, la, stochastic)
        if   g1 > g2: stats[t1]["pts"] += 3
        elif g2 > g1: stats[t2]["pts"] += 3
        else:         stats[t1]["pts"] += 1; stats[t2]["pts"] += 1
        for t, gf, ga in [(t1,g1,g2),(t2,g2,g1)]:
            stats[t]["gf"] += gf; stats[t]["ga"] += ga; stats[t]["gd"] += gf-ga
    df = pd.DataFrame([{"team":t,**v} for t,v in stats.items()])
    df = df.sort_values(["pts","gd","gf"], ascending=False).reset_index(drop=True)
    df["pos"] = df.index + 1
    return df

def simulate_all_groups(lam, hf_cohost, stochastic):
    return {g: simulate_group(g, ts, lam, hf_cohost, stochastic) for g, ts in GROUPS.items()}

# ── 8 melhores 3ºs — Correção: random tiebreaker no modo MC ──────────────────
def best_third_place(standings: dict, mc_mode: bool = False) -> list[dict]:
    thirds = []
    for g, df in standings.items():
        row = df[df["pos"] == 3].iloc[0]
        thirds.append({
            "group": g, "team": row["team"],
            "pts": row["pts"], "gd": row["gd"], "gf": row["gf"],
            "_tb": random.random() if mc_mode else 0.0,
        })
    thirds.sort(key=lambda x: (x["pts"], x["gd"], x["gf"], x["_tb"]), reverse=True)
    return thirds[:8]

# ── Fase eliminatória ─────────────────────────────────────────────────────────
def sim_ko(t1: str, t2: str, mid: int,
           lam: dict, hf_cohost: dict, stochastic: bool) -> tuple[str, str, str]:
    venue = MATCH_VENUES.get(mid, "USA")
    lh, la = get_match_lambdas(t1, t2, venue, lam, hf_cohost)
    if stochastic:
        g1, g2 = int(np.random.poisson(lh)), int(np.random.poisson(la))
    else:
        g1, g2 = simulate_game_from_lambdas(lh, la, False)
    if g1 != g2:
        return (t1, t2, "normal") if g1 > g2 else (t2, t1, "normal")
    e1 = int(np.random.poisson(lh/3)) if stochastic else 0
    e2 = int(np.random.poisson(la/3)) if stochastic else 0
    g1 += e1; g2 += e2
    if g1 != g2:
        return (t1, t2, "prorrogacao") if g1 > g2 else (t2, t1, "prorrogacao")
    winner = (t1 if random.random() < 0.5 else t2) if stochastic else (t1 if lh >= la else t2)
    return winner, (t2 if winner == t1 else t1), "penaltis"

def simulate_knockouts(standings, lam, hf_cohost, stochastic, q_thirds=None, mc_mode=False):
    if q_thirds is None:
        q_thirds = best_third_place(standings, mc_mode=mc_mode)
    thirds_by_group = {t["group"]: t["team"] for t in q_thirds}
    third_asgn      = assign_thirds(q_thirds)           # {match_id → group}

    def pos(g: str, n: int) -> str:
        return standings[g][standings[g]["pos"] == n].iloc[0]["team"]

    def resolve(slot: str, mid: int) -> str:
        if slot[0] == "1": return pos(slot[1], 1)
        if slot[0] == "2": return pos(slot[1], 2)
        g = third_asgn.get(mid)
        return thirds_by_group.get(g, next(iter(thirds_by_group.values())))

    mw: dict[int, str] = {}
    mm: dict[int, str] = {}

    for mid, a_slot, b_slot, _ in R32:
        t1, t2 = resolve(a_slot, mid), resolve(b_slot, mid)
        w, l, mode = sim_ko(t1, t2, mid, lam, hf_cohost, stochastic)
        mw[mid] = w; mm[mid] = mode

    r16 = [(89,73,74),(90,75,76),(91,77,78),(92,79,80),
           (93,81,82),(94,83,84),(95,85,86),(96,87,88)]
    for mid, m1, m2 in r16:
        w, l, mode = sim_ko(mw[m1], mw[m2], mid, lam, hf_cohost, stochastic)
        mw[mid] = w; mm[mid] = mode

    qf = [(97,89,90),(98,91,92),(99,93,94),(100,95,96)]
    for mid, m1, m2 in qf:
        w, l, mode = sim_ko(mw[m1], mw[m2], mid, lam, hf_cohost, stochastic)
        mw[mid] = w; mm[mid] = mode

    sf_losers = {}
    for mid, m1, m2 in [(101,97,98),(102,99,100)]:
        w, l, mode = sim_ko(mw[m1], mw[m2], mid, lam, hf_cohost, stochastic)
        mw[mid] = w; mm[mid] = mode; sf_losers[mid] = l

    w3, _, mode3 = sim_ko(sf_losers[101], sf_losers[102], 103, lam, hf_cohost, stochastic)
    mw[103] = w3; mm[103] = mode3
    wf, _, modef = sim_ko(mw[101], mw[102], 104, lam, hf_cohost, stochastic)
    mw[104] = wf; mm[104] = modef

    return mw, mm, q_thirds

# ── Bracket determinístico com scores completos ───────────────────────────────
def sim_ko_det(t1: str, t2: str, mid: int,
               lam: dict, hf_cohost: dict) -> dict:
    """Deterministic KO match → full result dict."""
    venue = MATCH_VENUES.get(mid, "USA")
    lh, la = get_match_lambdas(t1, t2, venue, lam, hf_cohost)
    g1, g2 = simulate_game_from_lambdas(lh, la, False)
    if g1 != g2:
        w = t1 if g1 > g2 else t2
        mode = "Normal"
    else:
        w = t1 if lh >= la else t2   # ties decided by higher lambda
        mode = "Pênaltis"
    l = t2 if w == t1 else t1
    return {
        "mid": mid, "t1": t1, "t2": t2, "g1": g1, "g2": g2,
        "winner": w, "loser": l, "mode": mode,
        "lh": round(lh, 3), "la": round(la, 3), "venue": venue,
    }

def simulate_knockouts_det(standings: dict, lam: dict,
                           hf_cohost: dict, q_thirds: list | None = None
                           ) -> tuple[dict, dict, dict, list]:
    """Deterministic bracket → (mw, match_details, slot_desc, q_thirds)."""
    if q_thirds is None:
        q_thirds = best_third_place(standings, mc_mode=False)
    thirds_by_group = {t["group"]: t["team"] for t in q_thirds}
    third_asgn      = assign_thirds(q_thirds)

    def pos(g: str, n: int) -> str:
        return standings[g][standings[g]["pos"] == n].iloc[0]["team"]

    def resolve(slot: str, mid: int) -> str:
        if slot[0] == "1": return pos(slot[1], 1)
        if slot[0] == "2": return pos(slot[1], 2)
        g = third_asgn.get(mid)
        return thirds_by_group.get(g, next(iter(thirds_by_group.values())))

    def fmt_slot(slot: str, mid: int) -> str:
        if slot[0] == "1": return f"1º Grp {slot[1]}"
        if slot[0] == "2": return f"2º Grp {slot[1]}"
        g = third_asgn.get(mid, "?")
        return f"3º Grp {g}"

    mw: dict[int, str]        = {}
    details: dict[int, dict]  = {}
    slot_desc: dict[int, tuple] = {}

    for mid, a_slot, b_slot, _ in R32:
        t1, t2 = resolve(a_slot, mid), resolve(b_slot, mid)
        d = sim_ko_det(t1, t2, mid, lam, hf_cohost)
        mw[mid] = d["winner"]; details[mid] = d
        slot_desc[mid] = (fmt_slot(a_slot, mid), fmt_slot(b_slot, mid))

    r16 = [(89,73,74),(90,75,76),(91,77,78),(92,79,80),
           (93,81,82),(94,83,84),(95,85,86),(96,87,88)]
    for mid, m1, m2 in r16:
        d = sim_ko_det(mw[m1], mw[m2], mid, lam, hf_cohost)
        mw[mid] = d["winner"]; details[mid] = d
        slot_desc[mid] = (f"W M{m1}", f"W M{m2}")

    for mid, m1, m2 in [(97,89,90),(98,91,92),(99,93,94),(100,95,96)]:
        d = sim_ko_det(mw[m1], mw[m2], mid, lam, hf_cohost)
        mw[mid] = d["winner"]; details[mid] = d
        slot_desc[mid] = (f"W M{m1}", f"W M{m2}")

    sf_losers: dict[int, str] = {}
    for mid, m1, m2 in [(101,97,98),(102,99,100)]:
        d = sim_ko_det(mw[m1], mw[m2], mid, lam, hf_cohost)
        mw[mid] = d["winner"]; details[mid] = d
        sf_losers[mid] = d["loser"]
        slot_desc[mid] = (f"W M{m1}", f"W M{m2}")

    d = sim_ko_det(sf_losers[101], sf_losers[102], 103, lam, hf_cohost)
    mw[103] = d["winner"]; details[103] = d
    slot_desc[103] = ("Perd. SF1", "Perd. SF2")

    d = sim_ko_det(mw[101], mw[102], 104, lam, hf_cohost)
    mw[104] = d["winner"]; details[104] = d
    slot_desc[104] = ("W SF1", "W SF2")

    return mw, details, slot_desc, q_thirds

# ── Monte Carlo ───────────────────────────────────────────────────────────────
def run_mc(lam, hf_cohost, n_sim=N_SIM, seed=SEED) -> pd.DataFrame:
    np.random.seed(seed); random.seed(seed)
    stages = ["grupo","r32","r16","qf","sf","3lugar","campeon"]
    counts = {s: {t: 0 for t in ALL_COPA_TEAMS} for s in stages}

    for i in range(n_sim):
        if i % (n_sim // 10) == 0:
            pct = i * 100 // n_sim
            print(f"    {pct:3d}%", end="\r", flush=True)

        stngs = simulate_all_groups(lam, hf_cohost, stochastic=True)

        # Qualificados da fase de grupos (top 2 + 8 melhores 3ºs)
        q_thirds = best_third_place(stngs, mc_mode=True)
        q3_teams = {t["team"] for t in q_thirds}
        for g, df in stngs.items():
            for _, row in df.iterrows():
                if row["pos"] <= 2 or row["team"] in q3_teams:
                    counts["grupo"][row["team"]] += 1

        mw, mm, _ = simulate_knockouts(stngs, lam, hf_cohost, True, q_thirds=q_thirds)

        for mid in range(73, 89):   counts["r32"][mw[mid]] += 1
        for mid in range(89, 97):   counts["r16"][mw[mid]] += 1
        for mid in range(97, 101):  counts["qf"][mw[mid]]  += 1
        for mid in [101, 102]:      counts["sf"][mw[mid]]  += 1
        counts["3lugar"][mw[103]] += 1
        counts["campeon"][mw[104]] += 1

    print("    100% — concluído.")
    rows = []
    for t in ALL_COPA_TEAMS:
        rows.append({
            "time":     t,
            "grupo":    TEAM_TO_GROUP[t],
            "grupo%":   counts["grupo"][t]  / n_sim * 100,
            "r32%":     counts["r32"][t]    / n_sim * 100,
            "r16%":     counts["r16"][t]    / n_sim * 100,
            "qf%":      counts["qf"][t]     / n_sim * 100,
            "sf%":      counts["sf"][t]     / n_sim * 100,
            "3lugar%":  counts["3lugar"][t] / n_sim * 100,
            "campeon%": counts["campeon"][t]/ n_sim * 100,
        })
    return pd.DataFrame(rows).sort_values("campeon%", ascending=False).reset_index(drop=True)

# ── Diagnóstico opp_saves ─────────────────────────────────────────────────────
def build_opp_saves_diag(wide: pd.DataFrame,
                         team_feats: dict,
                         lam: dict) -> pd.DataFrame:
    # Confederação: de rows onde o time é home_team
    conf_map: dict[str, str] = {}
    for _, row in wide.iterrows():
        team = row["home_team"]
        if team not in conf_map:
            for c in CONF_DUMMIES:
                if row.get(c, 0):
                    conf_map[team] = c.replace("conf_", "")
                    break

    # saves_decay info por time (última ocorrência)
    saves_info: dict[str, dict] = {}
    for _, row in wide.sort_values("date").iterrows():
        for side in ["home","away"]:
            team     = row[f"{side}_team"]
            saves    = float(row[f"{side}_saves_decay"])
            imputed  = bool(row.get(f"{side}_saves_decay_imputed", False))
            has_data = bool(row.get(f"{side}_has_decay_data", True))
            if team not in saves_info:
                saves_info[team] = {"n_real": 0, "last_saves": saves, "last_imputed": imputed}
            info = saves_info[team]
            if not imputed and saves > 0 and has_data:
                info["n_real"] += 1
            info["last_saves"]   = saves
            info["last_imputed"] = imputed

    # lambda médio previsto (sem bonus, para diagnóstico puro)
    lambda_medio: dict[str, float] = {}
    for team in ALL_COPA_TEAMS:
        vals = [lam.get((team, opp), 1.2) for opp in ALL_COPA_TEAMS if opp != team]
        lambda_medio[team] = float(np.mean(vals)) if vals else 1.2

    rows = []
    for team in ALL_COPA_TEAMS:
        info = saves_info.get(team, {"last_saves": 2.5, "last_imputed": True, "n_real": 0})
        rows.append({
            "time":                team,
            "confederacao":        conf_map.get(team, "?"),
            "opp_saves_decay_valor": round(info["last_saves"], 4),
            "opp_saves_imputed":   info["last_imputed"],
            "n_jogos_saves_reais": info["n_real"],
            "lambda_medio_previsto": round(lambda_medio.get(team, 1.2), 4),
        })

    df = pd.DataFrame(rows)
    df = df.sort_values("opp_saves_decay_valor", ascending=True).reset_index(drop=True)
    return df

# ── Excel helpers ─────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # azul escuro
FILL_TOP2   = PatternFill("solid", fgColor="C6EFCE")   # verde claro
FILL_3RD    = PatternFill("solid", fgColor="FFEB9C")   # amarelo
FILL_4TH    = PatternFill("solid", fgColor="FFCCCC")   # vermelho claro
FILL_ALT    = PatternFill("solid", fgColor="EBF0FF")   # azul bem claro
FILL_NONE   = PatternFill("solid", fgColor="FFFFFF")

FONT_HEAD   = Font(bold=True, color="FFFFFF", size=10)
FONT_BOLD   = Font(bold=True, size=10)
FONT_NORM   = Font(size=10)

BORDER_THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

GROUP_COLORS = [
    "BDD7EE","FCE4D6","E2EFDA","FFF2CC","DAEEF3","F2CEEF",
    "E7E6E6","FFD7D7","D9F2E4","EDE7F6","FFF8DC","E8F5E9",
]


def _cell(ws, row, col, value="", bold=False, fill=None, align="left", color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=10, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = BORDER_THIN
    if fill:
        c.fill = fill
    return c


def _header(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_HEAD; c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER_THIN


# ── Constantes de rounds (usadas no bracket e no print) ──────────────────────
ROUND_NAMES = {
    "R32": "DEZESSEIS AVOS DE FINAL",
    "R16": "OITAVAS DE FINAL",
    "QF":  "QUARTAS DE FINAL",
    "SF":  "SEMIFINAIS",
    "3P":  "TERCEIRO LUGAR",
    "F":   "FINAL",
}
ROUND_COLORS = {
    "R32": "1F3864",   # azul marinho
    "R16": "2E5F8A",   # azul médio
    "QF":  "C55A11",   # laranja escuro
    "SF":  "7B2C00",   # marrom avermelhado
    "3P":  "7030A0",   # roxo
    "F":   "806000",   # dourado escuro
}
ROUND_MIDS = [
    ("R32", list(range(73, 89))),
    ("R16", list(range(89, 97))),
    ("QF",  list(range(97, 101))),
    ("SF",  [101, 102]),
    ("3P",  [103]),
    ("F",   [104]),
]


# ── Aba Mata-Mata ─────────────────────────────────────────────────────────────
def write_bracket(wb, bracket_details: dict, slot_desc: dict) -> None:
    ws = wb.create_sheet("Mata-Mata")
    ws.sheet_view.showGridLines = False

    # Colunas: A=Jogo B=OrigemA C=TimeA D=λA E=Placar F=λB G=TimeB H=OrigemB I=Decisão J=Sede
    for col, w in zip("ABCDEFGHIJ", [7, 12, 22, 7, 9, 7, 22, 12, 13, 10]):
        ws.column_dimensions[col].width = w

    N_COLS = 10
    COLS = ["Jogo", "Origem A", "Time A", "λA", "Placar", "λB", "Time B", "Origem B", "Decisão", "Sede"]

    FILL_WIN   = PatternFill("solid", fgColor="C6EFCE")  # verde — vencedor
    FILL_LOSE  = PatternFill("solid", fgColor="FFD7D7")  # salmão — eliminado
    FILL_PEN   = PatternFill("solid", fgColor="FFF2CC")  # amarelo — pênaltis
    FILL_GOLD  = PatternFill("solid", fgColor="FFD700")  # ouro — campeão
    FILL_ORIG  = PatternFill("solid", fgColor="F2F2F2")  # cinza claro — origem

    r = 1
    # Título geral
    c = ws.cell(row=r, column=1,
                value="Copa 2026 — Mata-Mata | Simulação Determinística | modelo v10_clean")
    c.font = Font(bold=True, size=12, color="FFFFFF"); c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER_THIN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COLS)
    ws.row_dimensions[r].height = 22
    r += 2

    for round_key, mids in ROUND_MIDS:
        rfill = PatternFill("solid", fgColor=ROUND_COLORS[round_key])
        is_final = round_key == "F"

        # Cabeçalho do round
        c = ws.cell(row=r, column=1, value=ROUND_NAMES[round_key])
        c.font = Font(bold=True, size=11, color="FFFFFF"); c.fill = rfill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER_THIN
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=N_COLS)
        ws.row_dimensions[r].height = 18
        r += 1

        # Headers colunas
        for ci, h in enumerate(COLS, 1):
            cc = ws.cell(row=r, column=ci, value=h)
            cc.font = Font(bold=True, size=9, color="FFFFFF"); cc.fill = rfill
            cc.alignment = Alignment(horizontal="center", vertical="center")
            cc.border = BORDER_THIN
        r += 1

        for mid in mids:
            d          = bracket_details[mid]
            sd         = slot_desc.get(mid, ("", ""))
            t1, t2     = d["t1"], d["t2"]
            winner     = d["winner"]
            g1, g2     = d["g1"], d["g2"]
            lh, la     = d["lh"], d["la"]
            mode       = d["mode"]
            venue      = d["venue"]
            is_pen     = mode == "Pênaltis"

            if is_final:
                fa = FILL_GOLD if t1 == winner else FILL_LOSE
                fb = FILL_GOLD if t2 == winner else FILL_LOSE
            elif is_pen:
                fa = FILL_PEN if t1 == winner else FILL_LOSE
                fb = FILL_PEN if t2 == winner else FILL_LOSE
            else:
                fa = FILL_WIN if t1 == winner else FILL_LOSE
                fb = FILL_WIN if t2 == winner else FILL_LOSE

            placar = f"{g1} - {g2}" + (" *" if is_pen else "")
            fscore = FILL_PEN if is_pen else FILL_NONE

            _cell(ws, r, 1,  f"M{mid}",     fill=rfill,    align="center", bold=True, color="FFFFFF")
            _cell(ws, r, 2,  sd[0],         fill=FILL_ORIG, align="center", color="595959")
            _cell(ws, r, 3,  t1,            fill=fa,        align="left",  bold=(t1 == winner))
            _cell(ws, r, 4,  f"{lh:.3f}",   fill=fa,        align="right")
            _cell(ws, r, 5,  placar,        fill=fscore,    align="center", bold=True)
            _cell(ws, r, 6,  f"{la:.3f}",   fill=fb,        align="right")
            _cell(ws, r, 7,  t2,            fill=fb,        align="left",  bold=(t2 == winner))
            _cell(ws, r, 8,  sd[1],         fill=FILL_ORIG, align="center", color="595959")
            _cell(ws, r, 9,  mode,          fill=FILL_PEN if is_pen else FILL_WIN,
                  align="center", color="CC6600" if is_pen else "006400")
            _cell(ws, r, 10, venue,         fill=FILL_ALT,  align="center")
            r += 1

        r += 1  # espaçador entre rounds

    # Legenda
    r += 1
    ws.cell(row=r, column=1, value="LEGENDA").font = Font(bold=True, size=9)
    r += 1
    for lbl, fill in [
        ("Vencedor / Classificado",               FILL_WIN),
        ("Eliminado",                             FILL_LOSE),
        ("Pênaltis (placar reg. + prorr.)",       FILL_PEN),
        ("Campeão (Final)",                       FILL_GOLD),
        ("* = decisão por pênaltis",              FILL_NONE),
    ]:
        c = ws.cell(row=r, column=1, value=lbl)
        c.font = Font(size=9); c.fill = fill; c.border = BORDER_THIN
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        r += 1


# ── Aba 1: Fase de Grupos ─────────────────────────────────────────────────────
def write_grupos(wb, det_standings, det_thirds):
    ws = wb.create_sheet("Fase de Grupos")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    for col in "BCDEFG": ws.column_dimensions[col].width = 8
    ws.column_dimensions["H"].width = 2  # espaçador

    q3_teams = {t["team"] for t in det_thirds}
    COLS = ["Time","Pts","GF","GA","GD","Pos"]

    r = 1
    for gi, (gname, gdf) in enumerate(det_standings.items()):
        fill_g = PatternFill("solid", fgColor=GROUP_COLORS[gi % len(GROUP_COLORS)])
        # Cabeçalho do grupo
        c = ws.cell(row=r, column=1, value=f"GRUPO {gname}")
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER_THIN
        for col in range(2, 7):
            cc = ws.cell(row=r, column=col)
            cc.fill = FILL_HEADER; cc.border = BORDER_THIN
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

        # Headers das colunas
        for ci, h in enumerate(COLS, 1):
            _header(ws, r, ci, h)
        r += 1

        # Times
        for _, row in gdf.iterrows():
            team = row["team"]
            pos  = int(row["pos"])
            if pos <= 2:          fill = FILL_TOP2
            elif team in q3_teams: fill = FILL_3RD
            else:                  fill = FILL_4TH
            _cell(ws, r, 1, team,         fill=fill, bold=(pos <= 2))
            _cell(ws, r, 2, int(row["pts"]), fill=fill, align="center")
            _cell(ws, r, 3, int(row["gf"]),  fill=fill, align="center")
            _cell(ws, r, 4, int(row["ga"]),  fill=fill, align="center")
            gd = int(row["gd"])
            _cell(ws, r, 5, f"+{gd}" if gd > 0 else str(gd), fill=fill, align="center")
            _cell(ws, r, 6, f"{pos}º",       fill=fill, align="center")
            r += 1

        r += 1  # linha vazia entre grupos

    # Legenda
    r += 1
    for lbl, fill in [("Classificado (Top 2)", FILL_TOP2),
                       ("Melhor 3º classificado", FILL_3RD),
                       ("Eliminado", FILL_4TH)]:
        c = ws.cell(row=r, column=1, value=lbl)
        c.font = Font(size=9); c.fill = fill; c.border = BORDER_THIN
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    # 8 melhores 3ºs
    r += 1
    c = ws.cell(row=r, column=1, value="8 MELHORES TERCEIROS CLASSIFICADOS")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="left"); c.border = BORDER_THIN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    for h, col in zip(["#","Grupo","Time","Pts","GD","GF"], range(1,7)):
        _header(ws, r, col, h)
    r += 1
    for rank, t in enumerate(det_thirds, 1):
        fill = FILL_3RD if rank <= 8 else FILL_4TH
        _cell(ws, r, 1, rank,         fill=fill, align="center")
        _cell(ws, r, 2, t["group"],   fill=fill, align="center")
        _cell(ws, r, 3, t["team"],    fill=fill)
        _cell(ws, r, 4, t["pts"],     fill=fill, align="center")
        gd = t["gd"]
        _cell(ws, r, 5, f"+{gd}" if gd > 0 else str(gd), fill=fill, align="center")
        _cell(ws, r, 6, t["gf"],     fill=fill, align="center")
        r += 1


# ── Aba 2: Monte Carlo ────────────────────────────────────────────────────────
def write_mc(wb, mc_df: pd.DataFrame):
    ws = wb.create_sheet("Monte Carlo")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 7
    for col in "CDEFGHIJ": ws.column_dimensions[col].width = 10

    COLS = ["Time","Grupo","Classif. Grupos%","R32%","R16%","QF%","SF%","3ºLugar%","Campeão%"]
    r = 1
    # Título
    c = ws.cell(row=r, column=1, value=f"Copa 2026 — Monte Carlo ({N_SIM:,} simulações) | modelo v10_clean")
    c.font = Font(bold=True, size=11, color="FFFFFF"); c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="left"); c.border = BORDER_THIN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    r += 1

    for ci, h in enumerate(COLS, 1):
        _header(ws, r, ci, h)
    r += 1

    grp_order = list(GROUPS.keys())
    prev_grp = None
    for i, row in mc_df.iterrows():
        fill = FILL_ALT if i % 2 == 0 else FILL_NONE
        _cell(ws, r, 1, row["time"],  fill=fill)
        _cell(ws, r, 2, row["grupo"], fill=fill, align="center")
        for ci, col in enumerate(["grupo%","r32%","r16%","qf%","sf%","3lugar%","campeon%"], 3):
            val = row[col]
            _cell(ws, r, ci, f"{val:.1f}%", fill=fill, align="right")
        r += 1

    # Top-10 campeões (sumário no rodapé)
    r += 2
    c = ws.cell(row=r, column=1, value="TOP-10 CAMPEÕES")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="left"); c.border = BORDER_THIN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    r += 1
    for ci, h in enumerate(["#","Time","Grupo","Campeão%"], 1):
        _header(ws, r, ci, h)
    r += 1
    for rank, (_, row) in enumerate(mc_df.head(10).iterrows(), 1):
        fill = PatternFill("solid", fgColor="FFD700") if rank == 1 else FILL_ALT if rank % 2 == 0 else FILL_NONE
        _cell(ws, r, 1, rank,               fill=fill, align="center")
        _cell(ws, r, 2, row["time"],        fill=fill, bold=(rank == 1))
        _cell(ws, r, 3, row["grupo"],       fill=fill, align="center")
        _cell(ws, r, 4, f"{row['campeon%']:.1f}%", fill=fill, align="right")
        r += 1


# ── Aba 3: Diagnóstico opp_saves ─────────────────────────────────────────────
def write_diag(wb, diag_df: pd.DataFrame):
    ws = wb.create_sheet("Diag opp_saves")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 17
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 22

    COLS = ["Time","Confederação","opp_saves_decay_valor","opp_saves_imputed",
            "n_jogos_saves_reais","lambda_médio_previsto"]
    r = 1
    # Título
    c = ws.cell(row=r, column=1,
                value="Diagnóstico opp_saves_decay por time — ordenado crescente (mais penalizado primeiro)")
    c.font = Font(bold=True, size=10, color="FFFFFF"); c.fill = FILL_HEADER
    c.alignment = Alignment(horizontal="left"); c.border = BORDER_THIN
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    r += 1

    # Nota explicativa
    note_txt = ("opp_saves_decay_valor = taxa de defesas do goleiro do time. "
                "Baixo valor = GK fraco = oponentes marcam mais. "
                "lambda_médio = média de gols previstos pelo time vs todos os oponentes da Copa.")
    c = ws.cell(row=r, column=1, value=note_txt)
    c.font = Font(size=9, italic=True, color="595959")
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    ws.row_dimensions[r].height = 30
    r += 1

    for ci, h in enumerate(COLS, 1):
        _header(ws, r, ci, h)
    r += 1

    for i, row in diag_df.iterrows():
        fill = FILL_ALT if i % 2 == 0 else FILL_NONE
        # Destaque para imputed
        if row["opp_saves_imputed"]:
            fill = PatternFill("solid", fgColor="FFE4B5")  # laranja suave = imputed

        _cell(ws, r, 1, row["time"],         fill=fill)
        _cell(ws, r, 2, row["confederacao"], fill=fill, align="center")
        _cell(ws, r, 3, row["opp_saves_decay_valor"], fill=fill, align="right")
        _cell(ws, r, 4, str(row["opp_saves_imputed"]), fill=fill, align="center",
              color="CC0000" if row["opp_saves_imputed"] else "006400")
        _cell(ws, r, 5, row["n_jogos_saves_reais"],    fill=fill, align="center")
        _cell(ws, r, 6, row["lambda_medio_previsto"],  fill=fill, align="right")
        r += 1

    # Legenda
    r += 1
    for lbl, fill in [("Valor imputed (sem dados reais de defesas)", PatternFill("solid", fgColor="FFE4B5")),
                       ("Valor real (dados históricos de defesas)",   FILL_NONE)]:
        c = ws.cell(row=r, column=1, value=lbl)
        c.font = Font(size=9); c.fill = fill; c.border = BORDER_THIN
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    SEP = "=" * 65
    print(f"\n{SEP}")
    print("  Copa 2026 — Simulação v11")
    print(f"  Modelo: {MODEL_PKL.name}  |  {N_SIM:,} MC")
    print(f"{SEP}\n")

    # Carregar dados
    print("  Carregando modelo e dataset...")
    model = load_model()
    wide  = pd.read_csv(MODEL_DATASET, parse_dates=["date"])
    print(f"  Dataset: {len(wide)} jogos × {len(wide.columns)} cols")

    # Extrair features
    team_feats = extract_team_feats(wide)
    hf_cohost  = load_cohost_factors()

    # Lambda cache
    print("  Pré-computando lambdas...")
    lam = precompute_lambdas(model, team_feats, ALL_COPA_TEAMS)
    print(f"  {len(lam)} pares computados")

    # Simulação determinística
    print("\n  Simulação determinística (grupos + bracket)...")
    np.random.seed(SEED); random.seed(SEED)
    det_std = simulate_all_groups(lam, hf_cohost, stochastic=False)
    det_thirds_all = []
    for g, df in det_std.items():
        row = df[df["pos"] == 3].iloc[0]
        det_thirds_all.append({"group":g,"team":row["team"],
                               "pts":int(row["pts"]),"gd":int(row["gd"]),"gf":int(row["gf"]),"_tb":0.0})
    det_thirds_all.sort(key=lambda x: (x["pts"],x["gd"],x["gf"]), reverse=True)
    det_thirds = det_thirds_all[:8]

    # Imprimir grupos determinísticos
    for g, df in det_std.items():
        q3 = {t["team"] for t in det_thirds}
        print(f"\n  Grupo {g}:", end="")
        for _, row in df.iterrows():
            flag = " ✓" if row["pos"] <= 2 else (" (3º)" if row["team"] in q3 else "")
            print(f"\n    {int(row['pos'])}. {row['team']:22s} Pts:{int(row['pts'])}  "
                  f"GD:{int(row['gd']):+d}  GF:{int(row['gf'])}{flag}", end="")

    print(f"\n\n  Melhores 3ºs:")
    for i, t in enumerate(det_thirds, 1):
        print(f"    {i}. Grp {t['group']}: {t['team']:22s} {t['pts']}pts  GD{t['gd']:+d}  GF{t['gf']}")

    # Bracket determinístico completo (com scores)
    print("\n\n  Simulando mata-mata determinístico...")
    det_mw, det_bracket, det_slot_desc, _ = simulate_knockouts_det(
        det_std, lam, hf_cohost, det_thirds
    )
    for round_key, mids in ROUND_MIDS:
        print(f"\n  {ROUND_NAMES[round_key]}:")
        for mid in mids:
            d = det_bracket[mid]
            pen = " *pen" if d["mode"] == "Pênaltis" else ""
            print(f"    M{mid:3d}: {d['t1']:22s} {d['g1']}-{d['g2']}{pen:5s} {d['t2']:22s}"
                  f"  -> {d['winner']}")

    # Monte Carlo
    print(f"\n  Monte Carlo ({N_SIM:,} simulações)...")
    mc_df = run_mc(lam, hf_cohost)

    print(f"\n  Top-10 campeões:")
    for i, row in mc_df.head(10).iterrows():
        print(f"    {i+1:2d}. {row['time']:22s} {row['campeon%']:5.1f}%")

    # Diagnóstico opp_saves
    print("\n  Construindo diagnóstico opp_saves...")
    diag_df = build_opp_saves_diag(wide, team_feats, lam)

    # Gerar Excel
    print("\n  Gerando Excel...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_grupos(wb, det_std, det_thirds)
    write_bracket(wb, det_bracket, det_slot_desc)
    write_mc(wb, mc_df)
    write_diag(wb, diag_df)

    wb.save(OUT_XLSX)
    print(f"\n  Salvo: {OUT_XLSX}")
    print(f"\n{SEP}\n")


if __name__ == "__main__":
    main()
