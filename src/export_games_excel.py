"""
export_games_excel.py — Exporta todos os jogos do dataset com gols para Excel.

Aba 1: Todos os jogos (data, times, gols, competição)
Aba 2: Resumo por confederação
Aba 3: Jogos Copa 2026 — bracket determinístico (v4)
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                                 numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Instale openpyxl: pip install openpyxl")
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MODEL_CSV  = Path("data/processed/model_dataset.csv")
OUTPUT     = Path("outputs/copa2026_jogos_gols.xlsx")
OUTPUT.parent.mkdir(parents=True, exist_ok=True)

# ─── Cores por confederação ───────────────────────────────────────────────────
CONF_COLORS = {
    "UEFA":       "BDD7EE",   # azul claro
    "CONMEBOL":   "C6EFCE",   # verde claro
    "AFC":        "FFEB9C",   # amarelo
    "CAF":        "FCE4D6",   # laranja claro
    "CONCACAF":   "E2EFDA",   # verde-água
    "OFC":        "D9D9D9",   # cinza
    "FRIENDLY":   "FFF2CC",   # creme
}

# match_type label
TYPE_LABELS = {
    "WCQ":        "WCQ",
    "friendly":   "Amistoso",
    "tournament": "Torneio",
}

HEADER_FILL    = PatternFill("solid", fgColor="1F3864")
HEADER_FONT    = Font(color="FFFFFF", bold=True, size=10)
SUBHDR_FILL    = PatternFill("solid", fgColor="2E75B6")
SUBHDR_FONT    = Font(color="FFFFFF", bold=True, size=10)
RESULT_FONT_W  = Font(bold=True, color="375623")   # verde escuro — vitória mandante
RESULT_FONT_A  = Font(bold=True, color="833C00")   # laranja — vitória visitante
RESULT_FONT_D  = Font(bold=True, color="595959")   # cinza — empate
THIN_SIDE      = Side(style="thin", color="BFBFBF")
THIN_BORDER    = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")


def _apply_header(ws, row_idx: int, cols: list[str],
                  fill=None, font=None) -> None:
    fill = fill or HEADER_FILL
    font = font or HEADER_FONT
    for j, label in enumerate(cols, 1):
        cell = ws.cell(row=row_idx, column=j, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _conf_from_row(row) -> str:
    for col in row.index:
        if col.startswith("conf_") and row[col] == 1:
            return col.replace("conf_", "")
    return "FRIENDLY"


def _result_label(hg, ag) -> str:
    if hg > ag:   return "V"
    if hg < ag:   return "D"
    return "E"


def build_sheet_jogos(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Jogos", 0)

    # Colunas
    headers = [
        "#", "Data", "Mandante", "Gols", "", "Gols", "Visitante",
        "Placar", "Res.", "Competição", "Conf.", "Tipo", "Fase",
    ]

    _apply_header(ws, 1, headers)

    # Larguras
    widths = [5, 12, 22, 6, 3, 6, 22, 9, 5, 32, 10, 10, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 18

    # Extrair confederação
    df2 = df.copy()
    df2["conf"] = df2.apply(_conf_from_row, axis=1)
    df2["date"] = pd.to_datetime(df2["date"])
    df2 = df2.sort_values(["date", "conf"]).reset_index(drop=True)

    for i, (_, row) in enumerate(df2.iterrows(), 2):
        hg = int(row["home_gols"]) if pd.notna(row["home_gols"]) else "-"
        ag = int(row["away_gols"]) if pd.notna(row["away_gols"]) else "-"
        placar = f"{hg}–{ag}"
        res    = _result_label(hg, ag) if isinstance(hg, int) else "?"
        conf   = row["conf"]
        mt     = TYPE_LABELS.get(row.get("match_type", ""), str(row.get("match_type", "")))

        # Fase (WCQ / Torneio / Copa / ...)
        comp = str(row.get("competition", ""))
        fase = "WCQ" if "Qual" in comp or "WCQ" in comp or mt == "WCQ" else mt

        vals = [
            i - 1,
            row["date"].strftime("%Y-%m-%d"),
            row["home_team"],
            hg,
            "×",
            ag,
            row["away_team"],
            placar,
            res,
            comp,
            conf,
            mt,
            fase,
        ]

        fill_hex = CONF_COLORS.get(conf, "FFFFFF")
        row_fill = PatternFill("solid", fgColor=fill_hex)

        for j, val in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill   = row_fill
            c.border = THIN_BORDER
            c.alignment = CENTER if j not in (3, 7, 10) else LEFT

        # Placar em negrito + cor
        pc = ws.cell(row=i, column=8)
        if res == "V":
            pc.font = RESULT_FONT_W
        elif res == "D":
            pc.font = RESULT_FONT_A
        else:
            pc.font = RESULT_FONT_D

        ws.row_dimensions[i].height = 14

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df2)+1}"
    print(f"  Aba 'Jogos': {len(df2)} jogos")


def build_sheet_resumo(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Resumo por Conf.", 1)

    df2 = df.copy()
    df2["conf"] = df2.apply(_conf_from_row, axis=1)
    df2["home_gols"] = pd.to_numeric(df2["home_gols"], errors="coerce")
    df2["away_gols"] = pd.to_numeric(df2["away_gols"], errors="coerce")

    # Estatísticas por confederação
    stats = []
    for conf, grp in df2.groupby("conf"):
        n = len(grp)
        total_gols = grp["home_gols"].sum() + grp["away_gols"].sum()
        avg_total  = total_gols / n
        avg_home   = grp["home_gols"].mean()
        avg_away   = grp["away_gols"].mean()
        pct_v      = ((grp["home_gols"] > grp["away_gols"]).sum() / n * 100)
        pct_e      = ((grp["home_gols"] == grp["away_gols"]).sum() / n * 100)
        pct_d      = ((grp["home_gols"] < grp["away_gols"]).sum() / n * 100)
        stats.append({
            "Confederação": conf,
            "Jogos": n,
            "Total Gols": int(total_gols),
            "Média Gols/Jogo": round(avg_total, 2),
            "Média Gols Casa": round(avg_home, 2),
            "Média Gols Visit.": round(avg_away, 2),
            "Δ Casa-Visita": round(avg_home - avg_away, 2),
            "% V Casa": round(pct_v, 1),
            "% Empate": round(pct_e, 1),
            "% V Visit.": round(pct_d, 1),
        })

    sdf = pd.DataFrame(stats).sort_values("Jogos", ascending=False)

    # Total geral
    n_tot = len(df2)
    tg    = df2["home_gols"].sum() + df2["away_gols"].sum()
    total_row = {
        "Confederação": "TOTAL",
        "Jogos": n_tot,
        "Total Gols": int(tg),
        "Média Gols/Jogo": round(tg / n_tot, 2),
        "Média Gols Casa": round(df2["home_gols"].mean(), 2),
        "Média Gols Visit.": round(df2["away_gols"].mean(), 2),
        "Δ Casa-Visita": round(df2["home_gols"].mean() - df2["away_gols"].mean(), 2),
        "% V Casa": round((df2["home_gols"] > df2["away_gols"]).mean() * 100, 1),
        "% Empate":  round((df2["home_gols"] == df2["away_gols"]).mean() * 100, 1),
        "% V Visit.": round((df2["home_gols"] < df2["away_gols"]).mean() * 100, 1),
    }
    sdf = pd.concat([sdf, pd.DataFrame([total_row])], ignore_index=True)

    cols = list(sdf.columns)
    _apply_header(ws, 1, cols)

    col_widths = [15, 8, 12, 17, 17, 18, 16, 10, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for i, (_, row) in enumerate(sdf.iterrows(), 2):
        conf = row["Confederação"]
        fill_hex = CONF_COLORS.get(conf, "F2F2F2") if conf != "TOTAL" else "1F3864"
        row_fill = PatternFill("solid", fgColor=fill_hex)
        is_total = (conf == "TOTAL")

        for j, val in enumerate(row.values, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill   = row_fill
            c.border = THIN_BORDER
            c.alignment = CENTER
            if is_total:
                c.font = Font(color="FFFFFF", bold=True)

        ws.row_dimensions[i].height = 16

    ws.freeze_panes = "A2"
    print(f"  Aba 'Resumo por Conf.': {len(sdf)} linhas")


def build_sheet_placares(wb, df: pd.DataFrame) -> None:
    """Distribuição de placares mais frequentes."""
    ws = wb.create_sheet("Dist. Placares", 2)

    df2 = df.copy()
    df2["home_gols"] = pd.to_numeric(df2["home_gols"], errors="coerce")
    df2["away_gols"] = pd.to_numeric(df2["away_gols"], errors="coerce")
    df2 = df2.dropna(subset=["home_gols", "away_gols"])
    df2["placar"] = df2.apply(lambda r: f"{int(r['home_gols'])}–{int(r['away_gols'])}", axis=1)

    top = df2["placar"].value_counts().reset_index()
    top.columns = ["Placar", "Frequência"]
    top["% do Total"] = (top["Frequência"] / len(df2) * 100).round(1)
    top = top.head(20)

    headers = ["#", "Placar", "Frequência", "% do Total", "Tipo de Resultado"]
    _apply_header(ws, 1, headers)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18

    for i, (_, row) in enumerate(top.iterrows(), 2):
        sc = str(row["Placar"])
        parts = sc.split("–")
        h, a = int(parts[0]), int(parts[1])
        res_type = "Vitória Mandante" if h > a else ("Vitória Visitante" if h < a else "Empate")

        fill_hex = ("C6EFCE" if h > a else ("FCE4D6" if h < a else "FFF2CC"))
        row_fill = PatternFill("solid", fgColor=fill_hex)

        for j, val in enumerate([i-1, row["Placar"], row["Frequência"], row["% do Total"], res_type], 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill = row_fill
            c.border = THIN_BORDER
            c.alignment = CENTER

    ws.freeze_panes = "A2"
    print(f"  Aba 'Dist. Placares': top-20 placares")


def build_sheet_por_time(wb, df: pd.DataFrame) -> None:
    """Estatísticas de gols por seleção (Copa 2026)."""
    COPA_TEAMS = {
        "Argentina","Brazil","Colombia","Ecuador","Uruguay","Paraguay",
        "Germany","France","Spain","England","Netherlands","Portugal",
        "Belgium","Austria","Switzerland","Croatia","Norway","Sweden",
        "Czech Republic","Turkey","Scotland","Bosnia and Herzegovina",
        "Japan","South Korea","Iran","Australia","Saudi Arabia",
        "Uzbekistan","Jordan","Iraq","Qatar",
        "Morocco","Senegal","Egypt","Ghana","Cape Verde",
        "DR Congo","Ivory Coast","South Africa","Algeria","Tunisia",
        "United States","Mexico","Canada","Panama","Curacao","Haiti",
        "New Zealand",
    }

    ws = wb.create_sheet("Times Copa 2026", 3)

    df2 = df.copy()
    df2["home_gols"] = pd.to_numeric(df2["home_gols"], errors="coerce")
    df2["away_gols"] = pd.to_numeric(df2["away_gols"], errors="coerce")

    rows_data = []
    all_teams = sorted(set(df2["home_team"].tolist()) | set(df2["away_team"].tolist()))
    copa_teams = [t for t in all_teams if t in COPA_TEAMS]

    for team in copa_teams:
        as_home = df2[df2["home_team"] == team]
        as_away = df2[df2["away_team"] == team]

        n = len(as_home) + len(as_away)
        gf = (as_home["home_gols"].sum() + as_away["away_gols"].sum())
        ga = (as_home["away_gols"].sum() + as_away["home_gols"].sum())

        # WCQ only
        h_wcq = as_home[as_home["match_type_wcq"] == 1]
        a_wcq = as_away[as_away["match_type_wcq"] == 1]
        n_wcq = len(h_wcq) + len(a_wcq)
        gf_wcq = h_wcq["home_gols"].sum() + a_wcq["away_gols"].sum()
        ga_wcq = h_wcq["away_gols"].sum() + a_wcq["home_gols"].sum()

        vitórias = (
            (as_home["home_gols"] > as_home["away_gols"]).sum()
            + (as_away["away_gols"] > as_away["home_gols"]).sum()
        )
        empates = (
            (as_home["home_gols"] == as_home["away_gols"]).sum()
            + (as_away["away_gols"] == as_away["home_gols"]).sum()
        )
        derrotas = n - vitórias - empates

        rows_data.append({
            "Seleção": team,
            "Jogos (total)": n,
            "GF": int(gf) if pd.notna(gf) else 0,
            "GA": int(ga) if pd.notna(ga) else 0,
            "Saldo": int(gf - ga) if pd.notna(gf) else 0,
            "Média GF": round(gf / n, 2) if n > 0 else 0,
            "V": int(vitórias),
            "E": int(empates),
            "D": int(derrotas),
            "% Vitórias": round(vitórias / n * 100, 1) if n > 0 else 0,
            "Jogos WCQ": n_wcq,
            "GF WCQ": int(gf_wcq) if pd.notna(gf_wcq) else 0,
            "GA WCQ": int(ga_wcq) if pd.notna(ga_wcq) else 0,
            "Média GF WCQ": round(gf_wcq / n_wcq, 2) if n_wcq > 0 else 0,
        })

    sdf = pd.DataFrame(rows_data).sort_values("% Vitórias", ascending=False)

    headers = list(sdf.columns)
    _apply_header(ws, 1, headers)

    col_widths = [22, 13, 6, 6, 8, 10, 6, 6, 6, 12, 12, 9, 9, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Gradiente de cor por % vitórias
    max_pct = sdf["% Vitórias"].max() or 1

    for i, (_, row) in enumerate(sdf.iterrows(), 2):
        pct = row["% Vitórias"] / 100
        g = int(210 - 80 * pct)
        fill_hex = f"C6{g:02X}CE"
        row_fill = PatternFill("solid", fgColor="C6EFCE" if pct > 0.6 else
                               ("FFEB9C" if pct > 0.4 else "FCE4D6"))

        for j, val in enumerate(row.values, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill = row_fill
            c.border = THIN_BORDER
            c.alignment = LEFT if j == 1 else CENTER

        ws.row_dimensions[i].height = 14

    ws.freeze_panes = "B2"
    print(f"  Aba 'Times Copa 2026': {len(sdf)} seleções")


def main() -> None:
    print("=" * 60)
    print("  export_games_excel.py — Jogos + Gols")
    print("=" * 60)

    df = pd.read_csv(MODEL_CSV, parse_dates=["date"])
    print(f"\n  {len(df)} jogos carregados de {MODEL_CSV}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_sheet_jogos(wb, df)
    build_sheet_resumo(wb, df)
    build_sheet_placares(wb, df)
    build_sheet_por_time(wb, df)

    wb.save(OUTPUT)
    print(f"\n  → Salvo: {OUTPUT}")
    print(f"  Abas: {wb.sheetnames}")


if __name__ == "__main__":
    main()
