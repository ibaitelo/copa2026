"""
export_copa_features.py — Exporta features de todas as 48 seleções da Copa 2026
organizadas por grupo em Excel.

Saída: outputs/copa2026_team_features.xlsx
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT         = Path(__file__).parent.parent
FULL_RAW     = ROOT / "data/raw/full_dataset_raw.csv"
RATINGS_CSV  = ROOT / "data/raw/player_ratings_per_game.csv"
OUTPUT_XLSX  = ROOT / "outputs/copa2026_team_features.xlsx"
OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

WINDOW = 5

# ---------------------------------------------------------------------------
# Grupos oficiais Copa 2026 (sorteio Miami 05/12/2024)
# ---------------------------------------------------------------------------
GROUPS: dict[str, list[str]] = {
    "A": ["Mexico",        "South Korea",   "South Africa",        "Czech Republic"],
    "B": ["Canada",        "Switzerland",   "Qatar",               "Bosnia and Herzegovina"],
    "C": ["Brazil",        "Morocco",       "Haiti",               "Scotland"],
    "D": ["United States", "Paraguay",      "Australia",           "Turkey"],
    "E": ["Germany",       "Ivory Coast",   "Ecuador",             "Curacao"],
    "F": ["Netherlands",   "Sweden",        "Tunisia",             "Japan"],
    "G": ["Belgium",       "Iran",          "New Zealand",         "Egypt"],
    "H": ["Spain",         "Saudi Arabia",  "Uruguay",             "Cape Verde"],
    "I": ["France",        "Senegal",       "Iraq",                "Norway"],
    "J": ["Argentina",     "Algeria",       "Austria",             "Jordan"],
    "K": ["Portugal",      "DR Congo",      "Uzbekistan",          "Colombia"],
    "L": ["England",       "Croatia",       "Ghana",               "Panama"],
}

CONF_MAP: dict[str, str] = {
    # CONMEBOL
    "Argentina": "CONMEBOL", "Brazil": "CONMEBOL", "Colombia": "CONMEBOL",
    "Ecuador": "CONMEBOL",   "Uruguay": "CONMEBOL", "Paraguay": "CONMEBOL",
    # UEFA
    "Germany": "UEFA",    "France": "UEFA",   "Spain": "UEFA",
    "England": "UEFA",    "Netherlands": "UEFA", "Portugal": "UEFA",
    "Belgium": "UEFA",    "Austria": "UEFA",  "Switzerland": "UEFA",
    "Croatia": "UEFA",    "Norway": "UEFA",   "Sweden": "UEFA",
    "Czech Republic": "UEFA", "Turkey": "UEFA", "Scotland": "UEFA",
    "Bosnia and Herzegovina": "UEFA",
    # AFC
    "Japan": "AFC",       "South Korea": "AFC", "Iran": "AFC",
    "Australia": "AFC",   "Saudi Arabia": "AFC", "Uzbekistan": "AFC",
    "Jordan": "AFC",      "Iraq": "AFC",          "Qatar": "AFC",
    # CAF
    "Morocco": "CAF",     "Senegal": "CAF",   "Egypt": "CAF",
    "Ghana": "CAF",       "Cape Verde": "CAF", "DR Congo": "CAF",
    "Ivory Coast": "CAF", "South Africa": "CAF", "Algeria": "CAF",
    "Tunisia": "CAF",
    # CONCACAF
    "United States": "CONCACAF", "Mexico": "CONCACAF", "Canada": "CONCACAF",
    "Panama": "CONCACAF",        "Curacao": "CONCACAF", "Haiti": "CONCACAF",
    # OFC
    "New Zealand": "OFC",
}

FEATURE_COLS = [
    "gls_avg5",
    "gls_ponderado_avg5",
    "win_rate_avg5",
    "shots_avg5",
    "shots_on_goal_avg5",
    "blocked_shots_avg5",
    "ball_possession_avg5",
    "passes_accurate_avg5",
    "fouls_avg5",
    "corners_avg5",
    "saves_avg5",
    "ast_avg5",
    "rating_medio_avg5",
    "opp_rating_medio_avg5",
    "delta_rating_avg5",
    "gls_avg5_vs_forte",
    "gls_avg5_vs_fraco",
    "shots_avg5_vs_forte",
]

COL_LABELS = {
    "gls_avg5":              "Gols/Jogo\n(últ.5)",
    "gls_ponderado_avg5":    "Gols\nPonderados\n(últ.5)",
    "win_rate_avg5":         "Taxa\nVitória\n(últ.5)",
    "shots_avg5":            "Chutes/\nJogo (últ.5)",
    "shots_on_goal_avg5":    "Chutes no\nGol (últ.5)",
    "blocked_shots_avg5":    "Chutes\nBloqueados\n(últ.5)",
    "ball_possession_avg5":  "Posse\nBola %\n(últ.5)",
    "passes_accurate_avg5":  "Passes\nCertos\n(últ.5)",
    "fouls_avg5":            "Faltas/\nJogo (últ.5)",
    "corners_avg5":          "Escanteios\n(últ.5)",
    "saves_avg5":            "Defesas\n(últ.5)",
    "ast_avg5":              "Assist.\n(últ.5)",
    "rating_medio_avg5":     "Rating\nMédio\n(últ.5)",
    "opp_rating_medio_avg5": "Rating\nOponente\n(últ.5)",
    "delta_rating_avg5":     "Delta\nRating\n(últ.5)",
    "gls_avg5_vs_forte":     "Gols vs\nForte\n(últ.5)",
    "gls_avg5_vs_fraco":     "Gols vs\nFraco\n(últ.5)",
    "shots_avg5_vs_forte":   "Chutes vs\nForte\n(últ.5)",
}


# ---------------------------------------------------------------------------
# Rolling helpers (anti-leakage garantido: posição i só vê 0..i-1)
# ---------------------------------------------------------------------------

def _nonnan_rolling(values: np.ndarray, window: int, min_periods: int = 1) -> np.ndarray:
    n = len(values)
    result = np.full(n, np.nan)
    buf: list[float] = []
    for i in range(n):
        if len(buf) >= min_periods:
            result[i] = float(np.mean(buf[-window:]))
        if not np.isnan(values[i]):
            buf.append(float(values[i]))
    return result


def _winrate_rolling(gm: np.ndarray, gs: np.ndarray, window: int) -> np.ndarray:
    n = len(gm)
    result = np.full(n, np.nan)
    for i in range(1, n):
        start = max(0, i - window)
        _gm = gm[start:i]
        _gs = gs[start:i]
        ok = ~(np.isnan(_gm) | np.isnan(_gs))
        if ok.any():
            pts = np.where(_gm[ok] > _gs[ok], 1.0,
                           np.where(_gm[ok] == _gs[ok], 0.5, 0.0))
            result[i] = pts.mean()
    return result


def _opp_weighted_rolling(values: np.ndarray, weights: np.ndarray,
                           window: int, min_periods: int = 2,
                           max_lookback: int = 20) -> np.ndarray:
    """
    gls_ponderado = sum(gols_i * opp_r_i) / sum(opp_r_i) sobre os últimos
    `window` jogos COM rating válido do adversário, olhando até `max_lookback`
    jogos para trás. Friendlies sem rating não comprimem a janela — o rolling
    busca os N jogos COM rating mais recentes. min_periods=2 exige ao menos 2
    jogos com rating para evitar distorção por partida única.
    """
    n = len(values)
    result = np.full(n, np.nan)
    for i in range(1, n):
        start = max(0, i - max_lookback)
        v = values[start:i]
        w = weights[start:i]
        mask = ~(np.isnan(v) | np.isnan(w))
        valid_idx = np.where(mask)[0]
        if len(valid_idx) >= min_periods:
            last_k = valid_idx[-window:]
            v_k = v[last_k]
            w_k = w[last_k]
            if w_k.sum() > 0:
                result[i] = np.dot(v_k, w_k) / w_k.sum()
    return result


def _conditional_nonnan_rolling(values: np.ndarray, condition: np.ndarray,
                                 window: int, min_periods: int = 2) -> np.ndarray:
    n = len(values)
    result = np.full(n, np.nan)
    buf: list[float] = []
    for i in range(n):
        if len(buf) >= min_periods:
            result[i] = float(np.mean(buf[-window:]))
        if (not np.isnan(values[i]) and not np.isnan(condition[i])
                and condition[i] > 0.5):
            buf.append(float(values[i]))
    return result


# ---------------------------------------------------------------------------
# Pipeline: carrega + merge ratings + rolling features
# ---------------------------------------------------------------------------

def compute_team_features() -> pd.DataFrame:
    print("  Carregando full_dataset_raw.csv ...")
    df = pd.read_csv(FULL_RAW)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")

    num_cols = ["gols_marcados", "gols_sofridos", "shots", "shots_on_goal",
                "blocked_shots", "ball_possession", "fouls", "corners",
                "passes_accurate", "saves", "ast"]
    for col in num_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    df["has_shot_data"] = df["has_shot_data"].map(
        {"True": True, "False": False, True: True, False: False}
    ).fillna(False)

    # Merge player ratings (próprio + adversário)
    df["rating_medio_game"]     = np.nan
    df["opp_rating_medio_game"] = np.nan

    if RATINGS_CSV.exists():
        ratings = pd.read_csv(RATINGS_CSV)
        ratings["match_id"] = ratings["match_id"].astype(str)
        df["match_id"]      = df["match_id"].astype(str)

        own = ratings[["match_id", "team", "rating_medio"]].rename(
            columns={"rating_medio": "rating_medio_game"})
        df = df.merge(own, on=["match_id", "team"], how="left", suffixes=("", "_r"))
        if "rating_medio_game_r" in df.columns:
            df["rating_medio_game"] = df["rating_medio_game_r"].combine_first(
                df["rating_medio_game"])
            df.drop(columns=["rating_medio_game_r"], inplace=True)

        opp = ratings[["match_id", "team", "rating_medio"]].rename(
            columns={"team": "opponent", "rating_medio": "opp_rating_medio_game"})
        df = df.merge(opp, on=["match_id", "opponent"], how="left", suffixes=("", "_r"))
        if "opp_rating_medio_game_r" in df.columns:
            df["opp_rating_medio_game"] = df["opp_rating_medio_game_r"].combine_first(
                df["opp_rating_medio_game"])
            df.drop(columns=["opp_rating_medio_game_r"], inplace=True)

        n_own = int(df["rating_medio_game"].notna().sum())
        n_opp = int(df["opp_rating_medio_game"].notna().sum())
        print(f"  Ratings: {n_own} linhas com rating próprio, {n_opp} com rating oponente")
    else:
        print("  player_ratings_per_game.csv não encontrado — rating features serão NaN")

    # Rolling features por time (ordenado por data)
    all_cols = ["gls_avg5", "ast_avg5", "shots_avg5", "shots_on_goal_avg5",
                "blocked_shots_avg5", "ball_possession_avg5", "fouls_avg5",
                "corners_avg5", "passes_accurate_avg5", "saves_avg5",
                "win_rate_avg5", "rating_medio_avg5", "opp_rating_medio_avg5",
                "delta_rating_avg5", "gls_avg5_vs_forte", "gls_avg5_vs_fraco",
                "shots_avg5_vs_forte", "gls_ponderado_avg5"]
    for col in all_cols:
        df[col] = np.nan

    print("  Computando rolling features ...")
    df = df.sort_values(["team", "date"]).reset_index(drop=True)

    simple_map = {
        "gols_marcados":   "gls_avg5",
        "ast":             "ast_avg5",
        "shots":           "shots_avg5",
        "shots_on_goal":   "shots_on_goal_avg5",
        "blocked_shots":   "blocked_shots_avg5",
        "ball_possession": "ball_possession_avg5",
        "fouls":           "fouls_avg5",
        "corners":         "corners_avg5",
        "passes_accurate": "passes_accurate_avg5",
        "saves":           "saves_avg5",
    }
    min_periods_map = {"shots": 3, "shots_on_goal": 3, "blocked_shots": 3}

    for team, grp in df.groupby("team", sort=False):
        grp = grp.sort_values("date")
        idx = grp.index

        for src, dst in simple_map.items():
            df.loc[idx, dst] = _nonnan_rolling(
                grp[src].values.astype(float) if src in grp.columns
                else np.full(len(grp), np.nan),
                WINDOW,
                min_periods=min_periods_map.get(src, 1),
            )

        df.loc[idx, "win_rate_avg5"] = _winrate_rolling(
            grp["gols_marcados"].values.astype(float),
            grp["gols_sofridos"].values.astype(float),
            WINDOW,
        )

        own_r = grp["rating_medio_game"].values.astype(float)
        opp_r = grp["opp_rating_medio_game"].values.astype(float)

        r5 = _nonnan_rolling(own_r, WINDOW, 1)
        o5 = _nonnan_rolling(opp_r, WINDOW, 1)
        df.loc[idx, "rating_medio_avg5"]     = r5
        df.loc[idx, "opp_rating_medio_avg5"] = o5
        df.loc[idx, "delta_rating_avg5"]     = r5 - o5

        is_forte = np.where(np.isnan(opp_r), np.nan, (opp_r > 6.8).astype(float))
        is_fraco = np.where(np.isnan(opp_r), np.nan, (opp_r <= 6.8).astype(float))
        gls  = grp["gols_marcados"].values.astype(float)
        shts = grp["shots"].values.astype(float) if "shots" in grp.columns \
            else np.full(len(grp), np.nan)

        df.loc[idx, "gls_avg5_vs_forte"]   = _conditional_nonnan_rolling(
            gls, is_forte, WINDOW, min_periods=2)
        df.loc[idx, "gls_avg5_vs_fraco"]   = _conditional_nonnan_rolling(
            gls, is_fraco, WINDOW, min_periods=2)
        df.loc[idx, "shots_avg5_vs_forte"] = _conditional_nonnan_rolling(
            shts, is_forte, WINDOW, min_periods=2)

        df.loc[idx, "gls_ponderado_avg5"] = _opp_weighted_rolling(gls, opp_r, WINDOW)

    return df


def get_latest_per_team(df: pd.DataFrame) -> pd.DataFrame:
    """Pega a linha mais recente de cada seleção (última feature antes da Copa)."""
    all_copa = [t for teams in GROUPS.values() for t in teams]
    sub = df[df["team"].isin(all_copa)].sort_values("date")
    latest = sub.groupby("team").last().reset_index()

    # Imputa NaN remanescentes pela mediana da confederação
    conf_col = [c for c in df.columns if c == "confederation"]
    if conf_col:
        latest["confederation"] = latest["team"].map(CONF_MAP)
    else:
        latest["confederation"] = latest["team"].map(CONF_MAP)

    for feat in FEATURE_COLS:
        if feat not in latest.columns:
            latest[feat] = np.nan
        for conf in latest["confederation"].unique():
            mask_nan  = (latest["confederation"] == conf) & latest[feat].isna()
            mask_val  = (latest["confederation"] == conf) & latest[feat].notna()
            if mask_nan.any() and mask_val.any():
                median_val = latest.loc[mask_val, feat].median()
                latest.loc[mask_nan, feat] = median_val

    return latest


# ---------------------------------------------------------------------------
# Exportação Excel
# ---------------------------------------------------------------------------

# Paleta de cores por confederação
CONF_COLORS = {
    "UEFA":     "DBEAFE",   # azul claro
    "CONMEBOL": "D1FAE5",   # verde claro
    "AFC":      "FEF3C7",   # amarelo claro
    "CAF":      "FCE7F3",   # rosa claro
    "CONCACAF": "E0E7FF",   # índigo claro
    "OFC":      "F3F4F6",   # cinza claro
}

HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")
GROUP_FILL   = PatternFill("solid", fgColor="2563EB")
SUBHDR_FILL  = PatternFill("solid", fgColor="BFDBFE")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def export_excel(latest: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Copa 2026 — Features"

    meta_cols  = ["Grupo", "Seleção", "Confederação"]
    feat_labels = [COL_LABELS[c] for c in FEATURE_COLS]
    all_headers = meta_cols + feat_labels

    n_meta = len(meta_cols)
    n_feat = len(FEATURE_COLS)
    n_cols = n_meta + n_feat

    # ── Linha 1: cabeçalho principal ──────────────────────────────────────
    ws.append(all_headers)
    header_row = ws.row_dimensions[1]
    header_row.height = 52
    for col_idx, hdr in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = hdr
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = THIN_BORDER

    # ── Dados por grupo ───────────────────────────────────────────────────
    current_row = 2
    for group_letter, teams in GROUPS.items():

        # Sub-cabeçalho do grupo
        ws.cell(row=current_row, column=1).value = f"GRUPO {group_letter}"
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.fill      = GROUP_FILL
            cell.font      = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = THIN_BORDER
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Times do grupo
        for pos, team in enumerate(teams, 1):
            row_data = latest[latest["team"] == team]
            conf     = CONF_MAP.get(team, "")
            fill     = _make_fill(CONF_COLORS.get(conf, "FFFFFF"))

            ws.cell(row=current_row, column=1).value = group_letter
            ws.cell(row=current_row, column=2).value = team
            ws.cell(row=current_row, column=3).value = conf

            for feat_idx, feat in enumerate(FEATURE_COLS, n_meta + 1):
                val = np.nan
                if not row_data.empty and feat in row_data.columns:
                    val = row_data.iloc[0][feat]
                cell = ws.cell(row=current_row, column=feat_idx)
                if pd.isna(val):
                    cell.value = None
                else:
                    cell.value = round(float(val), 3)
                    cell.number_format = "0.00"

            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center" if col_idx != 2
                                           else "left", vertical="center")
                cell.border    = THIN_BORDER
                cell.font      = Font(size=9)

            # Seleção em negrito
            ws.cell(row=current_row, column=2).font = Font(bold=True, size=9)
            ws.row_dimensions[current_row].height   = 15
            current_row += 1

        current_row += 1  # espaço entre grupos

    # ── Legendas de confederação ──────────────────────────────────────────
    current_row += 1
    ws.cell(row=current_row, column=1).value = "Legenda de cores:"
    ws.cell(row=current_row, column=1).font  = Font(bold=True, size=9)
    current_row += 1
    for conf, color in CONF_COLORS.items():
        c = ws.cell(row=current_row, column=1)
        c.value = conf
        c.fill  = _make_fill(color)
        c.font  = Font(size=9)
        c.border = THIN_BORDER
        current_row += 1

    # ── Larguras das colunas ──────────────────────────────────────────────
    col_widths = [7, 26, 12] + [11] * n_feat
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Congelar linha do cabeçalho + 2 primeiras colunas ─────────────────
    ws.freeze_panes = "C2"

    # ── Aba de metadados ──────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Descrição das Features")
    desc_rows = [
        ["Feature", "Descrição"],
        ["gls_avg5",             "Média simples de gols marcados nos últimos 5 jogos"],
        ["gls_ponderado_avg5",   "★ NOVA — Média de gols ponderada pelo rating do adversário: sum(gols × opp_rating) / sum(opp_rating). Goleadas contra times fracos valem menos."],
        ["win_rate_avg5",        "Taxa de vitória: vitória=1, empate=0.5, derrota=0 (média dos últimos 5)"],
        ["shots_avg5",           "Chutes totais por jogo (últimos 5)"],
        ["shots_on_goal_avg5",   "Chutes no gol por jogo (últimos 5)"],
        ["blocked_shots_avg5",   "Chutes bloqueados pelo adversário (últimos 5)"],
        ["ball_possession_avg5", "Posse de bola média % (últimos 5)"],
        ["passes_accurate_avg5", "Passes certos por jogo (últimos 5)"],
        ["fouls_avg5",           "Faltas cometidas por jogo (últimos 5)"],
        ["corners_avg5",         "Escanteios por jogo (últimos 5)"],
        ["saves_avg5",           "Defesas do goleiro por jogo (últimos 5)"],
        ["ast_avg5",             "Assistências por jogo (últimos 5, apenas amistosos FBref)"],
        ["rating_medio_avg5",    "Rating médio dos jogadores (SofaScore, últimos 5 jogos)"],
        ["opp_rating_medio_avg5","Rating médio dos jogadores do adversário (últimos 5 jogos)"],
        ["delta_rating_avg5",    "Diferença de rating: próprio − adversário (últimos 5)"],
        ["gls_avg5_vs_forte",    "Gols por jogo apenas contra adversários com rating > 6.8 (últimos 5)"],
        ["gls_avg5_vs_fraco",    "Gols por jogo apenas contra adversários com rating ≤ 6.8 (últimos 5)"],
        ["shots_avg5_vs_forte",  "Chutes por jogo apenas contra adversários com rating > 6.8 (últimos 5)"],
        ["", ""],
        ["Fonte de dados", "API-Football v3 (WCQ + torneios), FBref (amistosos), SofaScore via API (ratings)"],
        ["Imputação", "NaN restantes preenchidos pela mediana da confederação"],
        ["Cutoff",    "Features calculadas com todos os jogos até 05/06/2026 (incluindo amistosos de junho/2026)"],
    ]
    for r_data in desc_rows:
        ws_meta.append(r_data)

    ws_meta.column_dimensions["A"].width = 26
    ws_meta.column_dimensions["B"].width = 80
    ws_meta["A1"].font = Font(bold=True)
    ws_meta["B1"].font = Font(bold=True)

    wb.save(OUTPUT_XLSX)
    print(f"\n  ✓ Salvo: {OUTPUT_XLSX}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print("=" * 60)
    print("  export_copa_features.py — Features das 48 seleções Copa 2026")
    print("=" * 60)

    df = compute_team_features()
    latest = get_latest_per_team(df)

    # Diagnóstico rápido
    all_copa = [t for ts in GROUPS.values() for t in ts]
    print(f"\n  Seleções cobertas: {latest['team'].nunique()}/48")
    missing = set(all_copa) - set(latest["team"].unique())
    if missing:
        print(f"  Faltando: {sorted(missing)}")

    pct = latest[FEATURE_COLS].notna().mean().mul(100).round(1)
    print("\n  Cobertura por feature (% seleções com valor não-NaN):")
    for feat, p in pct.items():
        print(f"    {feat:<28}: {p:.1f}%")

    gls_simple   = latest["gls_avg5"].mean()
    gls_ponder   = latest["gls_ponderado_avg5"].mean()
    print(f"\n  gls_avg5 (simples) média geral        : {gls_simple:.3f}")
    print(f"  gls_ponderado_avg5 média geral        : {gls_ponder:.3f}")
    print(f"  Diferença média (simples − ponderado): {gls_simple - gls_ponder:.3f}")

    # Times onde o ponderado difere mais do simples (beneficiados por goleadas fáceis)
    diff = latest["gls_avg5"] - latest["gls_ponderado_avg5"]
    latest["_diff"] = diff
    print("\n  Top 10 seleções onde gols simples > ponderado (jogam contra times fracos):")
    top10 = latest.nlargest(10, "_diff")[["team", "gls_avg5", "gls_ponderado_avg5", "_diff"]]
    for _, row in top10.iterrows():
        print(f"    {row['team']:<28}: simples={row['gls_avg5']:.2f}  pond={row['gls_ponderado_avg5']:.2f}"
              f"  diff={row['_diff']:.2f}")
    latest.drop(columns=["_diff"], inplace=True)

    print("\n  Gerando Excel ...")
    export_excel(latest)
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
