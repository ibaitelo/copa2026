#!/usr/bin/env python3
"""
copa2026_v12.py — Pipeline Copa 2026 v12

MUDANÇAS vs v11:
1. Coleta jogos da fase de grupos Copa 2026 via API-Football (league=1, season=2026)
   match_type="WC", weight=1.0, home_advantage=0 (torneio em campo neutro)
2. home_advantage=0 para TODOS os jogos no histórico (não apenas amistosos)
3. CUTOFF=2026-06-28 (inclui fase de grupos no treino — decay temporal dá peso extra)
4. Retreina xgboost_v12.pkl (15 features, sem home_advantage no modelo)
5. Sem host_factor nas predições (home advantage ignorado completamente)
6. Usa resultados REAIS da fase de grupos para montar o chaveamento dos 16avos
7. Output: outputs/copa2026_r32_v12.xlsx
   Aba "Grupos"     — resultados reais + classificação final por grupo
   Aba "16avos"     — confrontos R32 com probabilidades Monte Carlo
   Aba "MC_Campeon" — top-20 campeões (MC 10k, a partir do chaveamento real)
"""
from __future__ import annotations

import json
import os
import pickle
import random
import sys
import time
import warnings
from itertools import combinations
from pathlib import Path
from shutil import copyfile

import numpy as np
import pandas as pd
import requests
from dotenv import load_dotenv
from scipy.stats import poisson as scipy_poisson
from xgboost import XGBRegressor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "src")
warnings.filterwarnings("ignore")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

load_dotenv(encoding="utf-8-sig")

# ── Paths ──────────────────────────────────────────────────────────────────────
RAW_CSV    = Path("data/raw/full_dataset_raw.csv")
RAW_BACKUP = Path("data/raw/full_dataset_raw_pre_v12.csv")
MODEL_CSV  = Path("data/processed/model_dataset.csv")
ELO_CSV    = Path("data/raw/elo_history.csv")
OUT_PKL    = Path("outputs/xgboost_v12.pkl")
OUT_XLSX   = Path("outputs/copa2026_r32_v12.xlsx")
CACHE_DIR  = Path("data/raw/api_cache")

CUTOFF = pd.Timestamp("2026-06-28")
SEED   = 42
N_SIM  = 10_000

# ── Feature set (15 features — idêntico ao v11) ────────────────────────────────
CONF_DUMMIES = ["conf_AFC","conf_CAF","conf_CONCACAF","conf_CONMEBOL","conf_OFC","conf_UEFA"]
RATING_FEATURES = [
    "elo_diff_decay","delta_rating_decay","gls_ponderado_decay",
    "gls_decay_vs_forte","gls_decay_vs_fraco","margem_gols_decay",
]
FEATURES = (["opp_saves_decay","shots_on_goal_decay","gols_sofridos_decay"]
            + RATING_FEATURES + CONF_DUMMIES)

XGB_PARAMS = dict(
    objective="count:poisson", n_estimators=300, max_depth=4,
    learning_rate=0.05, subsample=0.8, colsample_bytree=0.8,
    min_child_weight=3, random_state=42, n_jobs=-1, verbosity=0,
)

# ── Copa 2026 — chaveamento ────────────────────────────────────────────────────
GROUPS: dict[str, list[str]] = {
    "A": ["Mexico",        "South Korea",  "South Africa",        "Czech Republic"],
    "B": ["Canada",        "Switzerland",  "Qatar",               "Bosnia and Herzegovina"],
    "C": ["Brazil",        "Morocco",      "Haiti",               "Scotland"],
    "D": ["United States", "Paraguay",     "Australia",           "Turkey"],
    "E": ["Germany",       "Ivory Coast",  "Ecuador",             "Curacao"],
    "F": ["Netherlands",   "Sweden",       "Tunisia",             "Japan"],
    "G": ["Belgium",       "Iran",         "New Zealand",         "Egypt"],
    "H": ["Spain",         "Saudi Arabia", "Uruguay",             "Cape Verde"],
    "I": ["France",        "Senegal",      "Iraq",                "Norway"],
    "J": ["Argentina",     "Algeria",      "Austria",             "Jordan"],
    "K": ["Portugal",      "DR Congo",     "Uzbekistan",          "Colombia"],
    "L": ["England",       "Croatia",      "Ghana",               "Panama"],
}

ALL_COPA_TEAMS = [t for ts in GROUPS.values() for t in ts]

COPA_TEAM_CONF: dict[str, str] = {
    "Argentina":"CONMEBOL","Brazil":"CONMEBOL","Colombia":"CONMEBOL",
    "Ecuador":"CONMEBOL","Uruguay":"CONMEBOL","Paraguay":"CONMEBOL",
    "Germany":"UEFA","France":"UEFA","Spain":"UEFA","England":"UEFA",
    "Netherlands":"UEFA","Portugal":"UEFA","Belgium":"UEFA","Austria":"UEFA",
    "Switzerland":"UEFA","Croatia":"UEFA","Norway":"UEFA","Sweden":"UEFA",
    "Czech Republic":"UEFA","Turkey":"UEFA","Scotland":"UEFA",
    "Bosnia and Herzegovina":"UEFA",
    "Japan":"AFC","South Korea":"AFC","Iran":"AFC","Australia":"AFC",
    "Saudi Arabia":"AFC","Uzbekistan":"AFC","Jordan":"AFC","Iraq":"AFC",
    "Qatar":"AFC",
    "Morocco":"CAF","Senegal":"CAF","Egypt":"CAF","Ghana":"CAF",
    "Cape Verde":"CAF","DR Congo":"CAF","Ivory Coast":"CAF",
    "South Africa":"CAF","Algeria":"CAF","Tunisia":"CAF",
    "United States":"CONCACAF","Mexico":"CONCACAF","Canada":"CONCACAF",
    "Panama":"CONCACAF","Curacao":"CONCACAF","Haiti":"CONCACAF",
    "New Zealand":"OFC",
}

# R32 bracket — (match_id, slot_t1, slot_t2, eligible_groups_for_3rd|None)
R32 = [
    (73,  "2A",  "2B",  None),
    (74,  "1E",  "3",   ["A","B","C","D","F"]),
    (75,  "1F",  "2C",  None),
    (76,  "1C",  "2F",  None),
    (77,  "1I",  "3",   ["C","D","F","G","H"]),
    (78,  "2E",  "2I",  None),
    (79,  "1A",  "3",   ["C","E","F","H","I"]),
    (80,  "1L",  "3",   ["E","H","I","J","K"]),
    (81,  "1D",  "3",   ["B","E","F","I","J"]),
    (82,  "1G",  "3",   ["A","E","H","I","J"]),
    (83,  "2K",  "2L",  None),
    (84,  "1H",  "2J",  None),
    (85,  "1B",  "3",   ["E","F","G","I","J"]),
    (86,  "1J",  "2H",  None),
    (87,  "1K",  "3",   ["D","E","I","J","L"]),
    (88,  "2D",  "2G",  None),
]

THIRD_ELIGIBLE: dict[int, list[str]] = {
    mid: gs for mid, _, _, gs in R32 if gs is not None
}

MATCH_VENUES: dict[int, str] = {
    73:"USA", 74:"USA", 75:"USA", 76:"USA", 77:"USA", 78:"USA",
    79:"USA", 80:"USA", 81:"USA", 82:"USA", 83:"USA", 84:"USA",
    85:"Mexico", 86:"USA", 87:"Mexico", 88:"Mexico",
    89:"USA", 90:"USA", 91:"USA", 92:"USA",
    93:"USA", 94:"USA", 95:"Mexico", 96:"Mexico",
    97:"USA", 98:"USA", 99:"USA", 100:"Mexico",
    101:"USA", 102:"USA", 103:"USA", 104:"USA",
}

# ── API-Football ───────────────────────────────────────────────────────────────
API_KEY  = os.getenv("API_FOOTBALL_KEY", "")
BASE_URL = "https://v3.football.api-sports.io"
HEADERS  = {"x-apisports-key": API_KEY}

NAME_MAP: dict[str, str] = {
    "USA":                    "United States",
    "México":                 "Mexico",
    "Curaçao":                "Curacao",
    "Côte d'Ivoire":          "Ivory Coast",
    "Korea Republic":         "South Korea",
    "IR Iran":                "Iran",
    "FYR Macedonia":          "North Macedonia",
    "Bosnia & Herzegovina":   "Bosnia and Herzegovina",
    "Bosnia-Herzegovina":     "Bosnia and Herzegovina",
    "Türkiye":                "Turkey",
    "Cape Verde Islands":     "Cape Verde",
    "Congo DR":               "DR Congo",
    "Republic of Ireland":    "Ireland",
    "Czechia":                "Czech Republic",
    "Czech Rep.":             "Czech Republic",
}

STAT_MAP: dict[str, str] = {
    "Total Shots":      "shots",
    "Shots on Goal":    "shots_on_goal",
    "Blocked Shots":    "blocked_shots",
    "Ball Possession":  "ball_possession",
    "Fouls":            "fouls",
    "Yellow Cards":     "yellow_cards",
    "Red Cards":        "red_cards",
    "Corner Kicks":     "corners",
    "Offsides":         "offsides",
    "Passes accurate":  "passes_accurate",
    "Goalkeeper Saves": "saves",
}

COPA2026_TEAM_SET = set(ALL_COPA_TEAMS)

# ── Medianas de normalização de saves (mesmas do v11) ─────────────────────────
CONF_MEDIANS_SAVES = {
    "UEFA": 2.9353, "CONMEBOL": 2.7985, "AFC": 2.4974,
    "CONCACAF": 2.4799, "CAF": 1.8144, "OFC": 2.5,
}
GLOBAL_MEDIAN_SAVES = 2.5


# =============================================================================
# SEÇÃO 1 — API-Football: buscar fase de grupos Copa 2026
# =============================================================================

def _cache_path(endpoint: str, cache_id: str) -> Path:
    p = CACHE_DIR / endpoint.strip("/") / f"{cache_id}.json"
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def api_get(endpoint: str, params: dict, cache_id: str) -> dict:
    cp = _cache_path(endpoint, cache_id)
    if cp.exists():
        with open(cp, encoding="utf-8") as f:
            return json.load(f)
    url = f"{BASE_URL}/{endpoint.strip('/')}"
    resp = requests.get(url, headers=HEADERS, params=params, timeout=30, verify=False)
    resp.raise_for_status()
    data = resp.json()
    with open(cp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    time.sleep(0.5)
    return data


def _parse_val(val):
    if val is None or val == "null" or val == "":
        return None
    if isinstance(val, str):
        s = val.rstrip("%").strip()
        if not s:
            return None
        try:
            return float(s)
        except ValueError:
            return val
    return val


def fetch_copa2026_group_stage(existing_ids: set) -> list[dict]:
    """Busca jogos da fase de grupos Copa 2026 (league=1, season=2026, round contendo 'Group')."""
    print("  Buscando Copa 2026 (league=1, season=2026)...", flush=True)
    data = api_get(
        "fixtures",
        {"league": 1, "season": 2026, "status": "FT"},
        cache_id="wc2026_season2026_FT",
    )
    fixtures = data.get("response", [])
    print(f"  Total fixtures FT retornados: {len(fixtures)}")

    new_rows: list[dict] = []
    n_skip = n_irrel = n_not_gs = 0

    for fixture in fixtures:
        fid      = str(fixture["fixture"]["id"])
        rnd      = fixture.get("league", {}).get("round", "")
        home_raw = fixture["teams"]["home"]["name"]
        away_raw = fixture["teams"]["away"]["name"]
        home_n   = NAME_MAP.get(home_raw, home_raw)
        away_n   = NAME_MAP.get(away_raw, away_raw)

        # Apenas fase de grupos
        if "Group" not in rnd and "group" not in rnd.lower():
            n_not_gs += 1
            continue
        if home_n not in COPA2026_TEAM_SET and away_n not in COPA2026_TEAM_SET:
            n_irrel += 1
            continue
        if fid in existing_ids:
            n_skip += 1
            continue

        goals = fixture["goals"]
        date  = fixture["fixture"]["date"][:10]

        stats_raw  = api_get("fixtures/statistics", {"fixture": int(fid)}, cache_id=f"wc2026_{fid}")
        team_stats: dict[str, dict] = {}
        for entry in stats_raw.get("response", []):
            tname = entry["team"]["name"]
            team_stats[tname] = {s["type"]: s["value"] for s in entry.get("statistics", [])}

        for side, opp_side, t_raw, o_raw in [
            ("home","away", home_raw, away_raw),
            ("away","home", away_raw, home_raw),
        ]:
            t_n = NAME_MAP.get(t_raw, t_raw)
            o_n = NAME_MAP.get(o_raw, o_raw)
            row: dict = {
                "match_id":        fid,
                "date":            date,
                "competition":     "FIFA World Cup 2026",
                "confederation":   COPA_TEAM_CONF.get(t_n, "UEFA"),
                "match_type":      "WC",
                "season":          2026,
                "home_away":       side,
                "team":            t_n,
                "opponent":        o_n,
                "gols_marcados":   goals[side],
                "gols_sofridos":   goals[opp_side],
                "shots":           None,
                "shots_on_goal":   None,
                "blocked_shots":   None,
                "ball_possession": None,
                "fouls":           None,
                "yellow_cards":    None,
                "red_cards":       None,
                "corners":         None,
                "offsides":        None,
                "passes_accurate": None,
                "saves":           None,
                "ast":             None,
                "has_shot_data":   False,
                "weight":          1.0,
                "home_advantage":  0,   # campo neutro
            }
            stats = team_stats.get(t_raw) or team_stats.get(t_n)
            if stats:
                row["has_shot_data"] = True
                for api_type, col in STAT_MAP.items():
                    row[col] = _parse_val(stats.get(api_type))
            new_rows.append(row)

    n_new = len(new_rows) // 2
    print(f"  Copa 2026: {len(fixtures)} fixtures → {n_new} jogos novos "
          f"| {n_skip} já existiam | {n_not_gs} não são GS | {n_irrel} irrelevantes")
    return new_rows


# =============================================================================
# SEÇÃO 2 — Atualizar raw CSV
# =============================================================================

def update_raw_csv(new_rows: list[dict]) -> pd.DataFrame:
    raw = pd.read_csv(RAW_CSV)
    before = len(raw)
    if new_rows:
        copa_df = pd.DataFrame(new_rows)
        raw = pd.concat([raw, copa_df], ignore_index=True)
        raw = raw.drop_duplicates(subset=["match_id","team"], keep="last")
    # home_advantage = 0 para TODOS os jogos (nova política v12)
    n_changed = int((raw["home_advantage"] != 0).sum())
    raw["home_advantage"] = 0
    after = len(raw)
    print(f"  Raw: {before:,} → {after:,} linhas")
    print(f"  home_advantage=0 aplicado a {n_changed:,} linhas adicionais")
    raw.to_csv(RAW_CSV, index=False, encoding="utf-8")
    return raw


# =============================================================================
# SEÇÃO 3 — Rebuild model_dataset.csv
# =============================================================================

def rebuild_model_dataset(cutoff: pd.Timestamp) -> pd.DataFrame:
    from build_features import (
        load_and_merge_raw, add_decay_features, impute_by_confederation,
        build_match_dataset, XI_DEFAULT, save_data,
    )
    df_raw  = load_and_merge_raw()
    df_dec  = add_decay_features(df_raw, XI_DEFAULT)
    df_imp  = impute_by_confederation(df_dec)
    df_wide = build_match_dataset(df_imp)
    save_data(df_wide)
    return df_wide


# =============================================================================
# SEÇÃO 4 — Normalização de saves (idêntica ao v11)
# =============================================================================

def apply_saves_normalization(wide: pd.DataFrame) -> pd.DataFrame:
    """Normaliza opp_saves_decay por confederação (médias do v11)."""
    wide = wide.copy()
    for side, opp in [("home","away"),("away","home")]:
        saves_col = f"{side}_saves_decay"
        if saves_col not in wide.columns:
            continue
        # identifica confederação do time 'opp' (que é quem salva)
        opp_team_col = f"{opp}_team"
        def _med(team):
            c = COPA_TEAM_CONF.get(team, "UEFA")
            return CONF_MEDIANS_SAVES.get(c, GLOBAL_MEDIAN_SAVES)
        med_arr = wide[opp_team_col].map(_med).fillna(GLOBAL_MEDIAN_SAVES)
        wide[saves_col] = wide[saves_col].astype(float) / med_arr
    return wide


# =============================================================================
# SEÇÃO 5 — Treino + extração de features por time
# =============================================================================

def wide_to_long_v12(wide: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict] = []
    for _, r in wide.iterrows():
        for side, opp in [("home","away"),("away","home")]:
            row: dict = {
                "date":                r["date"],
                "team":                r[f"{side}_team"],
                "gols_marcados":       r[f"{side}_gols"],
                "opp_saves_decay":     float(r.get(f"{opp}_saves_decay", 1.0)),
                "shots_on_goal_decay": float(r[f"{side}_shots_on_goal_decay"]),
                "gols_sofridos_decay": float(r.get(f"{side}_gols_sofridos_decay", np.nan)),
                "saves_decay":         float(r.get(f"{side}_saves_decay", 1.0)),
            }
            for col in CONF_DUMMIES:
                row[col] = int(r.get(col, 0))
            for col in RATING_FEATURES:
                raw_v = r.get(f"{side}_{col}", np.nan)
                row[col] = float(raw_v) if pd.notna(raw_v) else np.nan
            rows.append(row)
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def train_v12(long: pd.DataFrame, cutoff: pd.Timestamp) -> XGBRegressor:
    train = long[long["date"] <= cutoff].copy()
    model = XGBRegressor(**XGB_PARAMS)
    model.fit(train[FEATURES], train["gols_marcados"])
    n_train = len(train)
    print(f"  Treino: {n_train:,} observações até {cutoff.date()}")
    return model


def extract_team_feats(long: pd.DataFrame) -> dict[str, dict]:
    team_feats: dict[str, dict] = {}
    for _, row in long.sort_values("date").iterrows():
        entry: dict = {
            "saves_decay":         float(row.get("saves_decay", 1.0)),
            "shots_on_goal_decay": float(row.get("shots_on_goal_decay", 3.5)),
            "gols_sofridos_decay": row.get("gols_sofridos_decay", np.nan),
        }
        for col in RATING_FEATURES:
            v = row.get(col, np.nan)
            entry[col] = float(v) if pd.notna(v) else np.nan
        team_feats[row["team"]] = entry

    if ELO_CSV.exists():
        elo = pd.read_csv(ELO_CSV, parse_dates=["date"])
        for team, val in elo.sort_values("date").groupby("team")["elo_after"].last().items():
            team_feats.setdefault(team, {})["_elo_current"] = float(val)
    return team_feats


def precompute_lambdas(model: XGBRegressor,
                       team_feats: dict[str, dict],
                       teams: list[str]) -> dict[tuple[str, str], float]:
    """Lambdas sem home_advantage (completamente neutro)."""
    rows, pairs = [], []
    for t1 in teams:
        for t2 in teams:
            if t1 == t2:
                continue
            f1 = team_feats.get(t1, {})
            f2 = team_feats.get(t2, {})
            row: dict = {
                "opp_saves_decay":     f2.get("saves_decay", 1.0),
                "shots_on_goal_decay": f1.get("shots_on_goal_decay", 3.5),
                "gols_sofridos_decay": f1.get("gols_sofridos_decay", np.nan),
                **{c: 0 for c in CONF_DUMMIES},
            }
            for col in RATING_FEATURES:
                if col == "elo_diff_decay":
                    e1 = f1.get("_elo_current", np.nan)
                    e2 = f2.get("_elo_current", np.nan)
                    row[col] = float(e1) - float(e2) if (pd.notna(e1) and pd.notna(e2)) else f1.get(col, np.nan)
                else:
                    row[col] = f1.get(col, np.nan)
            conf = COPA_TEAM_CONF.get(t1, "")
            ck = f"conf_{conf}"
            if ck in CONF_DUMMIES:
                row[ck] = 1
            rows.append(row)
            pairs.append((t1, t2))

    preds = model.predict(pd.DataFrame(rows)[FEATURES])
    return {pair: float(p) for pair, p in zip(pairs, preds)}


# =============================================================================
# SEÇÃO 6 — Standings reais da Copa 2026
# =============================================================================

def compute_real_standings(raw: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Calcula classificação real a partir dos jogos WC na raw dataset."""
    copa = raw[(raw["match_type"] == "WC") & (raw["home_away"] == "home")].copy()
    copa["date"] = pd.to_datetime(copa["date"])
    copa["home_gols"] = pd.to_numeric(copa["gols_marcados"], errors="coerce")
    copa["away_gols"] = pd.to_numeric(copa["gols_sofridos"], errors="coerce")

    team_to_group = {t: g for g, ts in GROUPS.items() for t in ts}
    standings: dict[str, dict[str, dict]] = {
        g: {t: {"pts":0,"gf":0,"ga":0,"gd":0,"w":0,"d":0,"l":0} for t in ts}
        for g, ts in GROUPS.items()
    }
    results_by_group: dict[str, list[dict]] = {g: [] for g in GROUPS}

    for _, row in copa.iterrows():
        home = NAME_MAP.get(row["team"], row["team"])
        away = NAME_MAP.get(row["opponent"], row["opponent"])
        gh   = int(row["home_gols"]) if pd.notna(row["home_gols"]) else None
        ga   = int(row["away_gols"]) if pd.notna(row["away_gols"]) else None
        if gh is None or ga is None:
            continue
        grp_h = team_to_group.get(home)
        grp_a = team_to_group.get(away)
        if grp_h is None or grp_a is None or grp_h != grp_a:
            continue

        g = grp_h
        results_by_group[g].append({
            "home": home, "away": away, "gh": gh, "ga": ga,
            "date": row["date"],
        })
        if gh > ga:
            standings[g][home]["pts"] += 3; standings[g][home]["w"] += 1
            standings[g][away]["l"] += 1
        elif ga > gh:
            standings[g][away]["pts"] += 3; standings[g][away]["w"] += 1
            standings[g][home]["l"] += 1
        else:
            standings[g][home]["pts"] += 1; standings[g][home]["d"] += 1
            standings[g][away]["pts"] += 1; standings[g][away]["d"] += 1
        standings[g][home]["gf"] += gh; standings[g][home]["ga"] += ga
        standings[g][home]["gd"] += gh - ga
        standings[g][away]["gf"] += ga; standings[g][away]["ga"] += gh
        standings[g][away]["gd"] += ga - gh

    dfs: dict[str, pd.DataFrame] = {}
    for g, st in standings.items():
        df = pd.DataFrame([{"team":t, **v} for t, v in st.items()])
        df["pgj"] = df["w"] + df["d"] + df["l"]
        df = df.sort_values(["pts","gd","gf"], ascending=False).reset_index(drop=True)
        df["pos"] = df.index + 1
        df["results"] = [results_by_group[g]] * len(df)
        dfs[g] = df
    return dfs


def best_third_place_real(standings: dict[str, pd.DataFrame]) -> list[dict]:
    thirds = []
    for g, df in standings.items():
        row = df[df["pos"] == 3].iloc[0]
        thirds.append({
            "group": g, "team": row["team"],
            "pts": row["pts"], "gd": row["gd"], "gf": row["gf"],
        })
    thirds.sort(key=lambda x: (x["pts"],x["gd"],x["gf"]), reverse=True)
    return thirds[:8]


# =============================================================================
# SEÇÃO 7 — assign_thirds (bipartite matching)
# =============================================================================

def _augment(mid: int, available: dict, assignment: dict, visited: set) -> bool:
    for group in available.get(mid, []):
        if group not in visited:
            visited.add(group)
            rev = {v: k for k, v in assignment.items()}
            if group not in rev or _augment(rev[group], available, assignment, visited):
                assignment[mid] = group
                return True
    return False


def assign_thirds(qualifying_thirds: list[dict]) -> dict[int, str]:
    q_groups  = {t["group"] for t in qualifying_thirds}
    available = {mid: [g for g in gs if g in q_groups]
                 for mid, gs in THIRD_ELIGIBLE.items()}
    match_order = sorted(available, key=lambda m: len(available[m]))
    assignment: dict[int, str] = {}
    for mid in match_order:
        _augment(mid, available, assignment, {mid})
    return assignment


def resolve_r32_slot(slot: str, mid: int,
                     standings: dict[str, pd.DataFrame],
                     thirds_by_group: dict[str, dict],
                     third_asgn: dict[int, str]) -> str:
    if slot[0] == "1":
        return standings[slot[1]][standings[slot[1]]["pos"] == 1].iloc[0]["team"]
    if slot[0] == "2":
        return standings[slot[1]][standings[slot[1]]["pos"] == 2].iloc[0]["team"]
    g = third_asgn.get(mid, "A")
    return thirds_by_group.get(g, {}).get("team", "TBD")


def fmt_slot(slot: str, mid: int, third_asgn: dict[int, str]) -> str:
    if slot[0] == "1": return f"1º Grp {slot[1]}"
    if slot[0] == "2": return f"2º Grp {slot[1]}"
    g = third_asgn.get(mid, "?")
    return f"3º Grp {g}"


# =============================================================================
# SEÇÃO 8 — Monte Carlo: probabilidades dos 16avos + campeão
# =============================================================================

def _sim_poisson_match(lh: float, la: float) -> tuple[float, float, float]:
    """P(T1 vence 90min), P(empate 90min), P(T2 vence 90min) — analítico."""
    max_g = 8
    probs_h = [scipy_poisson.pmf(k, lh) for k in range(max_g+1)]
    probs_a = [scipy_poisson.pmf(k, la) for k in range(max_g+1)]
    p1 = sum(probs_h[i] * probs_a[j] for i in range(max_g+1) for j in range(i))
    pe = sum(probs_h[i] * probs_a[i] for i in range(max_g+1))
    p2 = sum(probs_h[i] * probs_a[j] for j in range(max_g+1) for i in range(j))
    return p1, pe, p2


def compute_r32_matchups(standings: dict[str, pd.DataFrame],
                         lam: dict[tuple, float]) -> list[dict]:
    """Retorna lista de dicts com info de cada jogo do R32."""
    q_thirds = best_third_place_real(standings)
    thirds_by_group = {t["group"]: t for t in q_thirds}
    third_asgn = assign_thirds(q_thirds)

    matchups = []
    for mid, a_slot, b_slot, _ in R32:
        t1 = resolve_r32_slot(a_slot, mid, standings, thirds_by_group, third_asgn)
        t2 = resolve_r32_slot(b_slot, mid, standings, thirds_by_group, third_asgn)
        lh = lam.get((t1, t2), 1.2)
        la = lam.get((t2, t1), 1.0)
        p1, pe, p2 = _sim_poisson_match(lh, la)
        # P(avanço): T1 vence 90min OU empata e ganha ET/PK (50/50)
        pa1 = p1 + pe * 0.5
        pa2 = p2 + pe * 0.5
        matchups.append({
            "mid":      mid,
            "slot_t1":  fmt_slot(a_slot, mid, third_asgn),
            "slot_t2":  fmt_slot(b_slot, mid, third_asgn),
            "t1":       t1,
            "t2":       t2,
            "lh":       round(lh, 3),
            "la":       round(la, 3),
            "p1":       round(p1, 4),
            "pe":       round(pe, 4),
            "p2":       round(p2, 4),
            "pa1":      round(pa1, 4),
            "pa2":      round(pa2, 4),
            "venue":    MATCH_VENUES.get(mid, "USA"),
        })
    return matchups


def run_mc(lam: dict[tuple, float],
           standings: dict[str, pd.DataFrame],
           n_sim: int = N_SIM) -> tuple[dict[str, int], dict[str, int]]:
    """Monte Carlo: champions + R32 team adv rates."""
    random.seed(SEED); np.random.seed(SEED)
    q_thirds_real = best_third_place_real(standings)
    thirds_by_group = {t["group"]: t for t in q_thirds_real}
    third_asgn = assign_thirds(q_thirds_real)

    # Resolve R32 matchups (fixed, since group stage is done)
    r32_fixed: list[tuple[str,str,int]] = []
    for mid, a_slot, b_slot, _ in R32:
        t1 = resolve_r32_slot(a_slot, mid, standings, thirds_by_group, third_asgn)
        t2 = resolve_r32_slot(b_slot, mid, standings, thirds_by_group, third_asgn)
        r32_fixed.append((t1, t2, mid))

    r16_bracket = [(89,73,74),(90,75,76),(91,77,78),(92,79,80),
                   (93,81,82),(94,83,84),(95,85,86),(96,87,88)]
    qf_bracket  = [(97,89,90),(98,91,92),(99,93,94),(100,95,96)]
    sf_bracket  = [(101,97,98),(102,99,100)]

    def sim_ko(t1, t2, mid):
        lh = lam.get((t1,t2), 1.2); la = lam.get((t2,t1), 1.0)
        g1, g2 = int(np.random.poisson(lh)), int(np.random.poisson(la))
        if g1 != g2: return t1 if g1 > g2 else t2
        g1 += int(np.random.poisson(lh/3)); g2 += int(np.random.poisson(la/3))
        if g1 != g2: return t1 if g1 > g2 else t2
        return t1 if random.random() < 0.5 else t2

    champions: dict[str, int] = {}
    r32_adv:   dict[str, int] = {}

    print(f"  Monte Carlo ({n_sim:,} simulações)...", flush=True)
    for i in range(n_sim):
        if i % 2000 == 0:
            print(f"    {i/n_sim*100:5.0f}%", end="  ", flush=True)
        mw: dict[int, str] = {}

        # R32 — chaveamento fixo
        for t1, t2, mid in r32_fixed:
            w = sim_ko(t1, t2, mid)
            mw[mid] = w
            r32_adv[w] = r32_adv.get(w, 0) + 1

        # R16 → QF → SF → Final
        for mid, m1, m2 in r16_bracket:
            mw[mid] = sim_ko(mw[m1], mw[m2], mid)
        for mid, m1, m2 in qf_bracket:
            mw[mid] = sim_ko(mw[m1], mw[m2], mid)
        sf_losers = {}
        for mid, m1, m2 in sf_bracket:
            t1, t2 = mw[m1], mw[m2]
            w = sim_ko(t1, t2, mid)
            mw[mid] = w
            sf_losers[mid] = t2 if w == t1 else t1
        mw[103] = sim_ko(sf_losers[101], sf_losers[102], 103)
        mw[104] = sim_ko(mw[101], mw[102], 104)
        champions[mw[104]] = champions.get(mw[104], 0) + 1

    print("\n  Concluído.")
    return champions, r32_adv


# =============================================================================
# SEÇÃO 9 — Geração do Excel
# =============================================================================

# Estilos
GOLD   = PatternFill("solid", fgColor="FFD700")
SILVER = PatternFill("solid", fgColor="C0C0C0")
BRONZE = PatternFill("solid", fgColor="CD7F32")
QUAL   = PatternFill("solid", fgColor="90EE90")
ELIM   = PatternFill("solid", fgColor="FFB6B6")
HEADER = PatternFill("solid", fgColor="1A3A5C")
SUBHDR = PatternFill("solid", fgColor="2B5A8A")
GREY   = PatternFill("solid", fgColor="E8E8E8")

def _hfont(bold=True, color="FFFFFF", size=11):
    return Font(bold=bold, color=color, size=size)

def _border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def _align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)

def set_col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


def write_grupos_sheet(wb: openpyxl.Workbook, standings: dict[str, pd.DataFrame]) -> None:
    ws = wb.create_sheet("Grupos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A1"

    row = 1
    ws.cell(row, 1, "Copa 2026 — Classificação Final por Grupo").font = Font(bold=True, size=14, color="1A3A5C")
    ws.merge_cells(f"A1:K1")
    ws.cell(row, 1).alignment = _align("left")
    row += 2

    headers = ["Pos","Time","PGJ","V","E","D","GP","GC","SG","Pts","Situação"]
    for g in sorted(GROUPS.keys()):
        df = standings[g]
        n_games = df["pgj"].max()

        # Cabeçalho do grupo
        ws.cell(row, 1, f"GRUPO {g}").font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row, 1).fill = HEADER
        ws.cell(row, 1).alignment = _align()
        ws.merge_cells(f"A{row}:K{row}")
        row += 1

        # Cabeçalhos das colunas
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row, ci, h)
            c.fill = SUBHDR; c.font = _hfont(size=10); c.alignment = _align(); c.border = _border()
        row += 1

        # Times
        for _, r in df.iterrows():
            pos = int(r["pos"])
            team = r["team"]
            situacao = ""
            fill = None
            if pos == 1:
                situacao = "✓ Classificado (1º)"; fill = QUAL
            elif pos == 2:
                situacao = "✓ Classificado (2º)"; fill = QUAL
            elif pos == 3:
                situacao = "? Depende (3º)"; fill = PatternFill("solid", fgColor="FFFACD")
            else:
                situacao = "✗ Eliminado"; fill = ELIM

            vals = [pos, team, r["pgj"], r["w"], r["d"], r["l"],
                    r["gf"], r["ga"], r["gd"], r["pts"], situacao]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row, ci, v)
                if fill: c.fill = fill
                c.alignment = _align("center" if ci != 2 else "left")
                c.border = _border()
                if ci == 2: c.font = Font(bold=(pos <= 2))
            row += 1
        row += 1  # espaço entre grupos

    set_col_widths(ws, {
        "A":"6","B":"22","C":"5","D":"5","E":"5","F":"5",
        "G":"5","H":"5","I":"5","J":"5","K":"20",
    })


def write_r32_sheet(wb: openpyxl.Workbook,
                    matchups: list[dict],
                    r32_adv: dict[str, int],
                    n_sim: int) -> None:
    ws = wb.create_sheet("16avos")
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1, "Copa 2026 — 16 Avos de Final (Previsões)").font = Font(bold=True, size=14, color="1A3A5C")
    ws.merge_cells("A1:K1")
    ws.cell(1, 1).alignment = _align("left")

    hdrs = ["Jogo","Origem T1","Time 1","λ1","P(T1 vence 90')","P(Empate 90')",
            "P(T2 vence 90')","λ2","Time 2","Origem T2","P(T1 avança)"]
    row = 3
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row, ci, h)
        c.fill = HEADER; c.font = _hfont(size=10); c.alignment = _align(); c.border = _border()
    row += 1

    for m in matchups:
        mid = m["mid"]
        t1, t2 = m["t1"], m["t2"]
        pa1 = m["pa1"]

        # cor do favorito
        fill_t1 = QUAL if pa1 >= 0.5 else None
        fill_t2 = QUAL if pa1 < 0.5  else None
        fill_even = PatternFill("solid", fgColor="FFFACD") if abs(pa1 - 0.5) < 0.05 else None

        if fill_even:
            fill_t1 = fill_even; fill_t2 = fill_even

        vals = [
            f"M{mid}", m["slot_t1"], t1,
            m["lh"], f"{m['p1']:.1%}", f"{m['pe']:.1%}",
            f"{m['p2']:.1%}", m["la"], t2, m["slot_t2"],
            f"{pa1:.1%}",
        ]
        fills = [None, None, fill_t1, None, None, None, None, None, fill_t2, None, None]
        for ci, (v, f) in enumerate(zip(vals, fills), 1):
            c = ws.cell(row, ci, v)
            if f: c.fill = f
            c.alignment = _align("center" if ci not in (2,3,9,10) else "left")
            c.border = _border()
            if ci in (3,9):
                c.font = Font(bold=True)
        row += 1

    set_col_widths(ws, {
        "A":"8","B":"16","C":"22","D":"7","E":"16","F":"16",
        "G":"16","H":"7","I":"22","J":"16","K":"14",
    })


def write_mc_sheet(wb: openpyxl.Workbook,
                   champions: dict[str, int],
                   n_sim: int) -> None:
    ws = wb.create_sheet("MC_Campeon")
    ws.sheet_view.showGridLines = False

    ws.cell(1, 1, f"Copa 2026 — Monte Carlo ({n_sim:,} simulações) — Probabilidade de Título").font = Font(bold=True, size=14, color="1A3A5C")
    ws.merge_cells("A1:E1")
    ws.cell(1, 1).alignment = _align("left")

    hdrs = ["Rank","Time","Simulações","Prob. (%)","Barra"]
    row = 3
    for ci, h in enumerate(hdrs, 1):
        c = ws.cell(row, ci, h)
        c.fill = HEADER; c.font = _hfont(size=10); c.alignment = _align(); c.border = _border()
    row += 1

    sorted_champs = sorted(champions.items(), key=lambda x: x[1], reverse=True)
    for rank, (team, cnt) in enumerate(sorted_champs, 1):
        pct = cnt / n_sim * 100
        fills_map = {1: GOLD, 2: SILVER, 3: BRONZE}
        fill = fills_map.get(rank, GREY if rank % 2 == 0 else None)
        bar = "█" * int(pct * 2)
        vals = [rank, team, cnt, f"{pct:.2f}%", bar]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row, ci, v)
            if fill: c.fill = fill
            c.alignment = _align("center" if ci != 2 else "left")
            c.border = _border()
            if ci == 2: c.font = Font(bold=(rank <= 3))
        row += 1

    set_col_widths(ws, {"A":"6","B":"22","C":"14","D":"12","E":"30"})


def generate_excel(matchups: list[dict],
                   standings: dict[str, pd.DataFrame],
                   champions: dict[str, int],
                   r32_adv: dict[str, int],
                   n_sim: int) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_grupos_sheet(wb, standings)
    write_r32_sheet(wb, matchups, r32_adv, n_sim)
    write_mc_sheet(wb, champions, n_sim)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"  Excel salvo: {OUT_XLSX}")


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    SEP = "=" * 72
    print(f"\n{SEP}")
    print("  COPA 2026 v12 — Pipeline completo")
    print(f"{SEP}\n")

    # ── 1. Backup ──────────────────────────────────────────────────────────────
    copyfile(RAW_CSV, RAW_BACKUP)
    print(f"  Backup: {RAW_BACKUP}")

    # ── 2. Fetch Copa 2026 group stage ─────────────────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 1 — Buscar fase de grupos Copa 2026")
    print(f"{SEP}")
    raw_current = pd.read_csv(RAW_CSV)
    existing_ids = set(raw_current["match_id"].astype(str).unique())
    print(f"  IDs existentes no raw: {len(existing_ids):,}")

    new_rows = fetch_copa2026_group_stage(existing_ids)
    n_new = len(new_rows) // 2

    # ── 3. Atualizar raw CSV ───────────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 2 — Atualizar raw dataset")
    print(f"{SEP}")
    raw = update_raw_csv(new_rows)
    n_copa = int((raw["match_type"] == "WC").sum())
    print(f"  Jogos Copa 2026 (WC) no raw: {n_copa // 2} jogos ({n_copa} linhas)")

    # ── 4. Rebuild model_dataset ───────────────────────────────────────────────
    print(f"\n{SEP}")
    print(f"  SEÇÃO 3 — Rebuild model_dataset.csv (CUTOFF={CUTOFF.date()})")
    print(f"{SEP}")
    wide = rebuild_model_dataset(CUTOFF)
    print(f"  model_dataset: {len(wide):,} jogos wide")

    # ── 5. Normalização de saves ───────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 4 — Normalização de saves por confederação")
    print(f"{SEP}")
    wide = apply_saves_normalization(wide)
    print("  Normalização aplicada.")

    # ── 6. Long format + treino ────────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 5 — Treino xgboost_v12")
    print(f"{SEP}")
    long = wide_to_long_v12(wide)
    model = train_v12(long, CUTOFF)
    OUT_PKL.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PKL, "wb") as f:
        pickle.dump(model, f)
    print(f"  Modelo salvo: {OUT_PKL}")

    # ── 7. Extrair features por time e lambdas ─────────────────────────────────
    team_feats = extract_team_feats(long)
    lam = precompute_lambdas(model, team_feats, ALL_COPA_TEAMS)
    print(f"  Lambdas computados: {len(lam)} pares")

    # ── 8. Standings reais ─────────────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 6 — Classificação real dos grupos")
    print(f"{SEP}")
    standings = compute_real_standings(raw)

    for g in sorted(GROUPS.keys()):
        df = standings[g]
        copa_games = int(df["pgj"].max())
        print(f"  Grupo {g}: {copa_games} rodadas jogadas")
        for _, r in df.iterrows():
            print(f"    {r['pos']}. {r['team']:<25} "
                  f"{r['pts']}pts  {r['gf']}-{r['ga']} (SG{r['gd']:+})")

    # ── 9. R32 matchups + probabilidades analíticas ────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 7 — 16 Avos de Final")
    print(f"{SEP}")
    matchups = compute_r32_matchups(standings, lam)
    for m in matchups:
        fav = m["t1"] if m["pa1"] >= 0.5 else m["t2"]
        prob_fav = max(m["pa1"], m["pa2"])
        print(f"  M{m['mid']:>2}  {m['t1']:<22} vs {m['t2']:<22}  "
              f"→ {fav} ({prob_fav:.0%})")

    # ── 10. Monte Carlo ────────────────────────────────────────────────────────
    print(f"\n{SEP}")
    print(f"  SEÇÃO 8 — Monte Carlo ({N_SIM:,} simulações)")
    print(f"{SEP}")
    champions, r32_adv = run_mc(lam, standings, N_SIM)
    total = sum(champions.values())

    print(f"\n  Top 20 Campeões:")
    for rank, (team, cnt) in enumerate(sorted(champions.items(), key=lambda x: x[1], reverse=True)[:20], 1):
        pct = cnt / total * 100
        print(f"  {rank:>2}. {team:<25} {pct:5.1f}%")

    # ── 11. Gerar Excel ────────────────────────────────────────────────────────
    print(f"\n{SEP}")
    print("  SEÇÃO 9 — Gerar Excel")
    print(f"{SEP}")
    generate_excel(matchups, standings, champions, r32_adv, N_SIM)

    print(f"\n{SEP}")
    print(f"  Copa 2026 v12 concluído!")
    print(f"  Modelo: {OUT_PKL}")
    print(f"  Excel:  {OUT_XLSX}")
    print(f"{SEP}\n")


if __name__ == "__main__":
    main()
