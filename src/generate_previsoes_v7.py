"""
generate_previsoes_v7.py — Copa 2026 Excel completo com SHAP por partida.

Estrutura do Excel:
  "Copa 2026"          : todos os 104+ jogos (grupos + bracket det.)
                         → colunas A–K visíveis (resultados)
                         → colunas L–Z recolhíveis (features + SHAP top-5)
  "Fase de Grupos"     : 48 jogos com λ, V%/E%/D%, top-3 placares
  "Mata-Mata"          : bracket determinístico
  "Classificação"      : esperança analítica por grupo
  "Probabilidades MC"  : probabilidades Monte Carlo 10k
  "Rodada 1"           : previsões 8 jogos Jun 11-13 com SHAP

Para investigar um jogo: expandir o grupo de colunas na aba "Copa 2026"
  → ver features de cada time + contribuição SHAP top-5
Para ver apenas resultados: recolher o grupo (botão [-] no topo da planilha)
"""

import sys
import warnings
from pathlib import Path
from itertools import combinations

import numpy as np
import pandas as pd
from scipy.stats import poisson as sp_poisson
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent))
from simulate_copa2026 import (
    train_model, precompute_lambdas, load_host_factors,
    simulate_all_groups, simulate_knockouts,
    get_host_factor, get_match_lambdas,
    GROUP_VENUE_COUNTRY, MATCH_VENUES,
    GROUPS, R32, ROUND1_GAMES, FEATURES, RATING_FEATURES,
)

OUT_XLSX = Path("outputs/copa2026_previsoes_v7.xlsx")

# ---------------------------------------------------------------------------
# Traduções
# ---------------------------------------------------------------------------

TEAM_PT: dict[str, str] = {
    "Mexico": "México", "South Korea": "Coreia do Sul", "South Africa": "África do Sul",
    "Czech Republic": "Rep. Tcheca", "Canada": "Canadá", "Switzerland": "Suíça",
    "Qatar": "Catar", "Bosnia and Herzegovina": "Bósnia-Herz.", "Brazil": "Brasil",
    "Morocco": "Marrocos", "Haiti": "Haiti", "Scotland": "Escócia",
    "United States": "EUA", "Paraguay": "Paraguai", "Australia": "Austrália",
    "Turkey": "Turquia", "Germany": "Alemanha", "Ivory Coast": "Costa do Marfim",
    "Ecuador": "Equador", "Curacao": "Curaçao", "Netherlands": "Holanda",
    "Sweden": "Suécia", "Tunisia": "Tunísia", "Japan": "Japão",
    "Belgium": "Bélgica", "Iran": "Irã", "New Zealand": "Nova Zelândia",
    "Egypt": "Egito", "Spain": "Espanha", "Saudi Arabia": "Arábia Saudita",
    "Uruguay": "Uruguai", "Cape Verde": "Cabo Verde", "France": "França",
    "Senegal": "Senegal", "Iraq": "Iraque", "Norway": "Noruega",
    "Argentina": "Argentina", "Algeria": "Argélia", "Austria": "Áustria",
    "Jordan": "Jordânia", "Portugal": "Portugal", "DR Congo": "RD Congo",
    "Uzbekistan": "Uzbequistão", "Colombia": "Colômbia", "England": "Inglaterra",
    "Croatia": "Croácia", "Ghana": "Gana", "Panama": "Panamá",
}

GROUP_HEADER_PT: dict[str, str] = {
    "A": "GRUPO A  —  México · Coreia do Sul · África do Sul · Rep. Tcheca",
    "B": "GRUPO B  —  Canadá · Suíça · Catar · Bósnia-Herz.",
    "C": "GRUPO C  —  Brasil · Marrocos · Haiti · Escócia",
    "D": "GRUPO D  —  EUA · Paraguai · Austrália · Turquia",
    "E": "GRUPO E  —  Alemanha · Costa do Marfim · Equador · Curaçao",
    "F": "GRUPO F  —  Holanda · Suécia · Tunísia · Japão",
    "G": "GRUPO G  —  Bélgica · Irã · Nova Zelândia · Egito",
    "H": "GRUPO H  —  Espanha · Arábia Saudita · Uruguai · Cabo Verde",
    "I": "GRUPO I  —  França · Senegal · Iraque · Noruega",
    "J": "GRUPO J  —  Argentina · Argélia · Áustria · Jordânia",
    "K": "GRUPO K  —  Portugal · RD Congo · Uzbequistão · Colômbia",
    "L": "GRUPO L  —  Inglaterra · Croácia · Gana · Panamá",
}


def pt(team: str) -> str:
    return TEAM_PT.get(team, team)


# ---------------------------------------------------------------------------
# Poisson helpers
# ---------------------------------------------------------------------------

def poisson_probs(lh: float, la: float, max_g: int = 10) -> tuple[float, float, float]:
    p_w = p_d = p_l = 0.0
    for h in range(max_g + 1):
        ph = sp_poisson.pmf(h, lh)
        for a in range(max_g + 1):
            p = ph * sp_poisson.pmf(a, la)
            if h > a:    p_w += p
            elif h == a: p_d += p
            else:        p_l += p
    return p_w, p_d, p_l


def top3_scores(lh: float, la: float, max_g: int = 8) -> list[str]:
    scores = [(h, a, sp_poisson.pmf(h, lh) * sp_poisson.pmf(a, la))
              for h in range(max_g + 1) for a in range(max_g + 1)]
    scores.sort(key=lambda x: x[2], reverse=True)
    return [f"{h}–{a} ({p*100:.1f}%)" for h, a, p in scores[:3]]


# ---------------------------------------------------------------------------
# ELO
# ---------------------------------------------------------------------------

def load_elo(team_feats: dict) -> dict[str, int]:
    elo: dict[str, int] = {}
    elo_path = Path("data/raw/elo_history.csv")
    if elo_path.exists():
        eh = pd.read_csv(elo_path, parse_dates=["date"])
        for team, val in (eh.sort_values("date")
                            .groupby("team")["elo_after"].last().items()):
            elo[team] = int(round(float(val)))
    for team, feats in team_feats.items():
        if team not in elo and "_elo_current" in feats:
            elo[team] = int(round(feats["_elo_current"]))
    return elo


# ---------------------------------------------------------------------------
# SHAP por jogo
# ---------------------------------------------------------------------------

def compute_shap_per_game(model, all_matchups: list[dict]) -> list[dict]:
    """Calcula SHAP top-5 para cada (team, opponent, venue) na perspectiva do time."""
    try:
        import shap as shap_lib
    except ImportError:
        return [{} for _ in all_matchups]

    if not all_matchups:
        return []

    df = pd.DataFrame([m["features"] for m in all_matchups])
    explainer  = shap_lib.TreeExplainer(model)
    shap_vals  = explainer(df[FEATURES], check_additivity=False)

    results = []
    for i, _ in enumerate(all_matchups):
        vals = dict(zip(FEATURES, shap_vals.values[i]))
        top5 = sorted(vals.items(), key=lambda x: abs(x[1]), reverse=True)[:5]
        results.append({"top5": top5, "all": vals})
    return results


def build_feature_row_for_shap(team: str, opponent: str, venue_country: str,
                                team_feats: dict, hf_home: dict, hf_cohost: dict) -> dict:
    f1 = team_feats.get(team, {})
    f2 = team_feats.get(opponent, {})
    hf = get_host_factor(team, venue_country, hf_home, hf_cohost)
    row = {
        "opp_saves_decay":     f2.get("saves_decay", 2.0),
        "shots_on_goal_decay": f1.get("shots_on_goal_decay", 3.5),
        "gols_sofridos_decay": f1.get("gols_sofridos_decay", np.nan),
        "host_factor":         hf,
        "conf_AFC": 0, "conf_CAF": 0, "conf_CONCACAF": 0,
        "conf_CONMEBOL": 0, "conf_OFC": 0, "conf_UEFA": 0,
    }
    for col in RATING_FEATURES:
        if col == "elo_diff_decay":
            elo1 = f1.get("_elo_current", np.nan)
            elo2 = f2.get("_elo_current", np.nan)
            row[col] = (float(elo1) - float(elo2)
                        if pd.notna(elo1) and pd.notna(elo2)
                        else f1.get(col, np.nan))
        else:
            row[col] = f1.get(col, np.nan)
    return row


# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------

FILL_TITLE   = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER  = PatternFill("solid", fgColor="2E75B6")
FILL_GROUP   = PatternFill("solid", fgColor="D6E4F0")
FILL_PHASE   = PatternFill("solid", fgColor="E2EFDA")
FILL_HOST    = PatternFill("solid", fgColor="FFF3CD")
FILL_COHOST  = PatternFill("solid", fgColor="FFF8E7")

FONT_TITLE   = Font(bold=True, color="FFFFFF", size=12)
FONT_HEADER  = Font(bold=True, color="FFFFFF", size=10)
FONT_SUBHDR  = Font(bold=True, color="1F4E79", size=9)
FONT_GROUP   = Font(bold=True, color="1F4E79", size=10)
FONT_PHASE   = Font(bold=True, color="375623", size=10)
FONT_NORMAL  = Font(size=10)
FONT_SMALL   = Font(size=9, color="595959")
FONT_SHAP_P  = Font(size=9, color="006400")
FONT_SHAP_N  = Font(size=9, color="8B0000")


def _c(ws, r, c, val=None, font=None, fill=None, align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    if font:  cell.font = font
    if fill:  cell.fill = fill
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return cell


def _style_row(ws, r_idx, fill, font, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=r_idx, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Aba "Copa 2026" — visão completa com SHAP recolhível
# ---------------------------------------------------------------------------

def build_copa2026_sheet(ws, model, team_feats, hf_home, hf_cohost, elo,
                         det_standings, det_ko, lam_full, lam_cohost, lam_neutral):
    """Aba principal: todos os jogos com colunas de features/SHAP agrupadas."""

    # Colunas:
    # A=Fase B=Jogo# C=Data D=Time1 E=ELO1 F=λ1 G=× H=λ2 I=ELO2 J=Time2
    # K=V% L=E% M=D% N=Favorito O=Placar1 P=Placar2
    # --- GRUPO RECOLHÍVEL (features + SHAP) ---
    # Q=HF_T1 R=HF_T2 S=ELO_diff T=shots_T1 U=gols_sofridos_T1 V=gls_pond_T1
    # W=gls_forte_T1 X=margem_T1 Y=SHAP#1 Z=SHAP#2 AA=SHAP#3 AB=SHAP#4 AC=SHAP#5

    N_MAIN = 16   # colunas A-P sempre visíveis
    N_FEAT = 7    # colunas Q-W (features time1)
    N_SHAP = 5    # colunas X-AB (SHAP top-5)
    N_TOTAL = N_MAIN + N_FEAT + N_SHAP

    # Linha de cabeçalho dupla
    headers_main = [
        "Fase", "Jogo", "Data", "Time 1", "ELO1", "λ₁", "×",
        "λ₂", "ELO2", "Time 2", "V%", "E%", "D%", "Favorito", "Placar 1", "Placar 2",
    ]
    headers_feat = ["HF T1", "HF T2", "ΔELO", "Chutes/Gol T1", "Gols Sofr. T1",
                    "Gls Pond. T1", "Gls×Forte T1"]
    headers_shap = ["SHAP #1", "SHAP #2", "SHAP #3", "SHAP #4", "SHAP #5"]

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_TOTAL)
    _c(ws, 1, 1,
       "COPA DO MUNDO 2026 — Análise completa por jogo  (v10 · venue-aware host_factor · 15 features)",
       font=FONT_TITLE, fill=FILL_TITLE)
    ws.row_dimensions[1].height = 22

    # Sub-cabeçalho: "RESULTADOS" span + "FEATURES (expandir →)" span
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_MAIN)
    _c(ws, 2, 1, "RESULTADOS E PROBABILIDADES", font=FONT_HEADER, fill=FILL_HEADER)
    ws.merge_cells(start_row=2, start_column=N_MAIN+1, end_row=2, end_column=N_TOTAL)
    _c(ws, 2, N_MAIN+1,
       "◀ EXPANDIR PARA VER FEATURES E SHAP  (botão [ + ] na borda superior)",
       font=Font(bold=True, color="7F4800", size=9),
       fill=PatternFill("solid", fgColor="FFF3CD"))
    ws.row_dimensions[2].height = 16

    # Cabeçalhos de coluna
    for ci, h in enumerate(headers_main + headers_feat + headers_shap, 1):
        _c(ws, 3, ci, h, font=FONT_HEADER if ci <= N_MAIN else FONT_SUBHDR,
           fill=FILL_HEADER if ci <= N_MAIN else PatternFill("solid", fgColor="FFF8E7"))
    ws.row_dimensions[3].height = 28

    # Larguras
    col_widths = [
        14, 5, 11, 22, 6, 6, 3, 6, 6, 22,  # A-J
        7, 7, 7, 18, 16, 16,                 # K-P
        7, 7, 7, 10, 12, 10, 11,             # Q-W (features)
        22, 22, 22, 22, 22,                  # X-AB (SHAP)
    ]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Agrupar colunas de features+SHAP (colunas Q onwards)
    col_q = get_column_letter(N_MAIN + 1)
    col_last = get_column_letter(N_TOTAL)
    ws.column_dimensions.group(col_q, col_last, hidden=True, outline_level=1)
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Coletar todos os matchups para SHAP em batch
    all_matchups_home = []  # perspectiva do time1 (home)
    game_records = []       # metadata dos jogos

    # Grupos
    jogo_num = 1
    for g in sorted(GROUPS.keys()):
        teams    = GROUPS[g]
        venue    = GROUP_VENUE_COUNTRY[g]
        for t1, t2 in combinations(teams, 2):
            lh, la = get_match_lambdas(t1, t2, venue, lam_full, lam_cohost, lam_neutral)
            pw, pd_, pl = poisson_probs(lh, la)
            sc = top3_scores(lh, la)
            hf1 = get_host_factor(t1, venue, hf_home, hf_cohost)
            hf2 = get_host_factor(t2, venue, hf_home, hf_cohost)
            game_records.append({
                "fase": f"Grupo {g}", "jogo": jogo_num, "data": "",
                "t1": t1, "t2": t2, "lh": lh, "la": la,
                "pw": pw, "pd_": pd_, "pl": pl,
                "fav": t1 if pw >= pl else t2,
                "sc": sc, "elo1": elo.get(t1, ""), "elo2": elo.get(t2, ""),
                "hf1": hf1, "hf2": hf2,
            })
            all_matchups_home.append({
                "features": build_feature_row_for_shap(t1, t2, venue, team_feats, hf_home, hf_cohost),
                "t1": t1, "t2": t2,
            })
            jogo_num += 1

    # Mata-mata determinístico
    mt = det_ko["teams"]
    mw = det_ko["winner"]
    phase_map = {}
    for mid, _, _, _ in R32:
        phase_map[mid] = "R32 (16-avos)"
    for mid in [89,90,91,92,93,94,95,96]:
        phase_map[mid] = "R16 (Oitavas)"
    for mid in [97,98,99,100]:
        phase_map[mid] = "QF (Quartas)"
    for mid in [101,102]:
        phase_map[mid] = "SF (Semifinais)"
    phase_map[103] = "3º Lugar"
    phase_map[104] = "FINAL"

    ko_ids = [r[0] for r in R32] + [89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104]
    for mid in ko_ids:
        if mid not in mt:
            continue
        t1, t2  = mt[mid]
        venue   = MATCH_VENUES.get(mid, "USA")
        lh, la  = get_match_lambdas(t1, t2, venue, lam_full, lam_cohost, lam_neutral)
        pw, pd_, pl = poisson_probs(lh, la)
        sc      = top3_scores(lh, la)
        hf1     = get_host_factor(t1, venue, hf_home, hf_cohost)
        hf2     = get_host_factor(t2, venue, hf_home, hf_cohost)
        game_records.append({
            "fase": phase_map.get(mid, f"M{mid}"), "jogo": mid, "data": "",
            "t1": t1, "t2": t2, "lh": lh, "la": la,
            "pw": pw, "pd_": pd_, "pl": pl,
            "fav": mw.get(mid, t1 if lh >= la else t2),
            "sc": sc, "elo1": elo.get(t1, ""), "elo2": elo.get(t2, ""),
            "hf1": hf1, "hf2": hf2,
        })
        all_matchups_home.append({
            "features": build_feature_row_for_shap(t1, t2, venue, team_feats, hf_home, hf_cohost),
            "t1": t1, "t2": t2,
        })

    # Calcular SHAP em batch
    print("  Calculando SHAP por jogo...")
    shap_results = compute_shap_per_game(model, all_matchups_home)

    # Escrever linhas
    row = 4
    prev_fase = None
    for i, rec in enumerate(game_records):
        shap_r = shap_results[i] if i < len(shap_results) else {}

        if rec["fase"] != prev_fase:
            # Cabeçalho de fase
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_TOTAL)
            fase_label = rec["fase"]
            if rec["fase"].startswith("Grupo"):
                fill, font = FILL_GROUP, FONT_GROUP
                g_key = rec["fase"].split()[-1]
                if g_key in GROUP_HEADER_PT:
                    fase_label = GROUP_HEADER_PT[g_key]
            elif rec["fase"] == "FINAL":
                fill = PatternFill("solid", fgColor="FFD700")
                font = Font(bold=True, color="7F4800", size=11)
            else:
                fill, font = FILL_PHASE, FONT_PHASE
            _c(ws, row, 1, fase_label, font=font, fill=fill)
            ws.row_dimensions[row].height = 18
            row += 1
            prev_fase = rec["fase"]

        # Fundo para times-sede
        row_fill = None
        if rec["hf1"] > 0.1:
            row_fill = FILL_HOST
        elif rec["hf1"] > 0 or rec["hf2"] > 0:
            row_fill = FILL_COHOST

        # Colunas principais (A-P)
        vals_main = [
            rec["fase"], rec["jogo"], rec["data"],
            pt(rec["t1"]), rec["elo1"], round(rec["lh"], 3), "×",
            round(rec["la"], 3), rec["elo2"], pt(rec["t2"]),
            f"{rec['pw']*100:.1f}%", f"{rec['pd_']*100:.1f}%", f"{rec['pl']*100:.1f}%",
            pt(rec["fav"]), rec["sc"][0], rec["sc"][1] if len(rec["sc"]) > 1 else "",
        ]
        for ci, v in enumerate(vals_main, 1):
            cell = _c(ws, row, ci, v, font=FONT_NORMAL)
            if row_fill:
                cell.fill = row_fill

        # Colunas de features (Q-W)
        f1_feats = team_feats.get(rec["t1"], {})
        feat_vals = [
            round(rec["hf1"], 4),
            round(rec["hf2"], 4),
            round(all_matchups_home[i]["features"].get("elo_diff_decay", 0) if i < len(all_matchups_home) else 0, 1),
            round(f1_feats.get("shots_on_goal_decay", float("nan")), 3) if pd.notna(f1_feats.get("shots_on_goal_decay")) else "",
            round(f1_feats.get("gols_sofridos_decay", float("nan")), 3) if pd.notna(f1_feats.get("gols_sofridos_decay")) else "",
            round(f1_feats.get("gls_ponderado_decay", float("nan")), 3) if pd.notna(f1_feats.get("gls_ponderado_decay")) else "",
            round(f1_feats.get("gls_decay_vs_forte", float("nan")), 3) if pd.notna(f1_feats.get("gls_decay_vs_forte")) else "",
        ]
        for ci, v in enumerate(feat_vals, N_MAIN + 1):
            _c(ws, row, ci, v, font=FONT_SMALL,
               fill=PatternFill("solid", fgColor="FFFFF0") if row_fill is None else row_fill)

        # Colunas SHAP (X-AB)
        if shap_r and "top5" in shap_r:
            for si, (feat_name, shap_v) in enumerate(shap_r["top5"]):
                ci = N_MAIN + N_FEAT + 1 + si
                label = f"{feat_name}: {shap_v:+.3f}"
                font  = FONT_SHAP_P if shap_v >= 0 else FONT_SHAP_N
                _c(ws, row, ci, label, font=font,
                   fill=PatternFill("solid", fgColor="F0FFF0" if shap_v >= 0 else "FFF0F0"))

        ws.row_dimensions[row].height = 16
        row += 1

    # Freeze panes
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = True


# ---------------------------------------------------------------------------
# Abas auxiliares (reaproveitadas do v6)
# ---------------------------------------------------------------------------

def build_grupos_rows(lam_full, lam_cohost, lam_neutral,
                      hf_home, hf_cohost, elo) -> list[list]:
    rows = [["COPA DO MUNDO 2026 — PREVISÃO DE PLACARES  (v10 · venue-aware host_factor)"] + [None]*14]
    rows.append(["Grp", "Time 1", "ELO", "λ1", "×", "λ2", "ELO", "Time 2",
                 "V%", "E%", "D%", "Favorito", "1º Placar", "2º Placar", "3º Placar"])
    for g in sorted(GROUPS.keys()):
        teams = GROUPS[g]
        venue = GROUP_VENUE_COUNTRY[g]
        rows.append([GROUP_HEADER_PT[g]] + [None]*14)
        for t1, t2 in combinations(teams, 2):
            lh, la   = get_match_lambdas(t1, t2, venue, lam_full, lam_cohost, lam_neutral)
            pw, pd_, pl = poisson_probs(lh, la)
            sc       = top3_scores(lh, la)
            hf1 = get_host_factor(t1, venue, hf_home, hf_cohost)
            hf2 = get_host_factor(t2, venue, hf_home, hf_cohost)
            fav = pt(t1) if pw >= pl else pt(t2)
            rows.append([g, pt(t1), elo.get(t1,""), round(lh,3), "×", round(la,3),
                         elo.get(t2,""), pt(t2),
                         f"{pw*100:.1f}%", f"{pd_*100:.1f}%", f"{pl*100:.1f}%",
                         fav, sc[0], sc[1], sc[2]])
    return rows


def build_matamata_rows(lam_full, lam_cohost, lam_neutral,
                        hf_home, hf_cohost, elo, det_standings, det_ko) -> list[list]:
    rows = [["COPA DO MUNDO 2026 — MATA-MATA DETERMINÍSTICO  (v10)"] + [None]*14]
    rows.append(["Fase","Time 1","ELO","λ1","×","λ2","ELO","Time 2",
                 "V%","E%","D%","Vencedor","1º Placar","2º Placar","3º Placar"])
    mt = det_ko["teams"]

    def _add_phase(label, match_ids):
        rows.append([label] + [None]*14)
        for mid in match_ids:
            if mid not in mt:
                continue
            t1, t2  = mt[mid]
            venue   = MATCH_VENUES.get(mid, "USA")
            lh, la  = get_match_lambdas(t1, t2, venue, lam_full, lam_cohost, lam_neutral)
            pw, pd_, pl = poisson_probs(lh, la)
            sc      = top3_scores(lh, la)
            winner  = pt(det_ko["winner"].get(mid, ""))
            rows.append([label, pt(t1), elo.get(t1,""), round(lh,3), "×", round(la,3),
                         elo.get(t2,""), pt(t2),
                         f"{pw*100:.1f}%", f"{pd_*100:.1f}%", f"{pl*100:.1f}%",
                         winner, sc[0], sc[1], sc[2]])

    _add_phase("DEZESSEIS-AVOS", [r[0] for r in R32])
    _add_phase("OITAVAS",        [89,90,91,92,93,94,95,96])
    _add_phase("QUARTAS",        [97,98,99,100])
    _add_phase("SEMIFINAIS",     [101,102])
    _add_phase("3º LUGAR",       [103])
    _add_phase("FINAL",          [104])
    return rows


def build_classificacao_rows(lam_full, lam_cohost, lam_neutral) -> list[list]:
    exp = {t: {"pts": 0.0, "gf": 0.0, "ga": 0.0}
           for ts in GROUPS.values() for t in ts}
    for g, teams in GROUPS.items():
        venue = GROUP_VENUE_COUNTRY[g]
        for t1, t2 in combinations(teams, 2):
            lh, la  = get_match_lambdas(t1, t2, venue, lam_full, lam_cohost, lam_neutral)
            pw, pd_, pl = poisson_probs(lh, la)
            exp[t1]["pts"] += 3*pw + pd_
            exp[t2]["pts"] += 3*pl + pd_
            exp[t1]["gf"] += lh;  exp[t1]["ga"] += la
            exp[t2]["gf"] += la;  exp[t2]["ga"] += lh

    def sort_key(t):
        e = exp[t]; return (e["pts"], e["gf"]-e["ga"], e["gf"])

    thirds = []
    for g in sorted(GROUPS.keys()):
        ranked = sorted(GROUPS[g], key=sort_key, reverse=True)
        t3 = ranked[2]
        thirds.append({"team": t3, "pts": exp[t3]["pts"]})
    thirds.sort(key=lambda x: x["pts"], reverse=True)
    qual_thirds = {d["team"] for d in thirds[:8]}

    rows = [["COPA DO MUNDO 2026 — CLASSIFICAÇÃO ESPERADA (pontos × probabilidades)"] + [None]*7]
    rows.append(["Pos","Time","Pts Esp.","GF Esp.","GA Esp.","Saldo","Grupo","Obs."])
    for g in sorted(GROUPS.keys()):
        ranked = sorted(GROUPS[g], key=sort_key, reverse=True)
        for pos, team in enumerate(ranked, 1):
            e = exp[team]
            obs = ("Classifica" if pos <= 2
                   else ("3º qualif." if pos == 3 and team in qual_thirds else None))
            rows.append([pos, pt(team), round(e["pts"],2), round(e["gf"],2),
                         round(e["ga"],2), round(e["gf"]-e["ga"],2), g, obs])
        rows.append([None]*8)
    if rows and all(v is None for v in rows[-1]):
        rows.pop()
    return rows


def build_probabilidades_rows(mc_csv: Path) -> list[list]:
    if not mc_csv.exists():
        return []
    mc = pd.read_csv(mc_csv)
    rows = [["COPA DO MUNDO 2026 — PROBABILIDADES MONTE CARLO  (10.000 simulações — v10)"] + [None]*8]
    rows.append(["#","Time","Grupo","Grupos%","Oitavas%","Quartas%",
                 "Semi%","Final%","3ºLugar%","Campeão%"])
    mc = mc.sort_values("champion", ascending=False).reset_index(drop=True)
    for rank, (_, row) in enumerate(mc.iterrows(), 1):
        rows.append([rank, pt(row["team"]), row.get("group",""),
                     f"{row['grupo']:.1f}%", f"{row['r32']:.1f}%", f"{row['r16']:.1f}%",
                     f"{row['qf']:.1f}%", f"{row['sf']:.1f}%",
                     f"{row.get('3rd_place',0):.1f}%", f"{row['champion']:.1f}%"])
    return rows


def build_rodada1_rows(model, team_feats, hf_home, hf_cohost, elo) -> list[list]:
    """Aba Rodada 1 com previsões + features + SHAP."""
    rows = [["RODADA 1 — Copa 2026 (11–13/jun/2026)  ·  v10 · venue-aware"] + [None]*18]
    rows.append(["Data","Grupo","Venue","Time 1","ELO1","HF T1","λ₁","×",
                 "λ₂","HF T2","ELO2","Time 2","V%","E%","D%","Favorito",
                 "Placar 1","Placar 2","Placar 3"])
    rows.append(["---Detalhe SHAP (perspectiva Time 1)---"] + [None]*18)

    for t1, t2, date_str, group, venue in ROUND1_GAMES:
        lh_feat = build_feature_row_for_shap(t1, t2, venue, team_feats, hf_home, hf_cohost)
        la_feat = build_feature_row_for_shap(t2, t1, venue, team_feats, hf_home, hf_cohost)
        df_pred = pd.DataFrame([lh_feat, la_feat])
        lams = model.predict(df_pred[FEATURES])
        lh, la  = float(lams[0]), float(lams[1])
        hf1 = get_host_factor(t1, venue, hf_home, hf_cohost)
        hf2 = get_host_factor(t2, venue, hf_home, hf_cohost)
        pw, pd_, pl = poisson_probs(lh, la)
        sc = top3_scores(lh, la)
        fav = pt(t1) if pw >= pl else pt(t2)
        rows.append([date_str, group, venue,
                     pt(t1), elo.get(t1,""), round(hf1,4), round(lh,3), "×",
                     round(la,3), round(hf2,4), elo.get(t2,""), pt(t2),
                     f"{pw*100:.1f}%", f"{pd_*100:.1f}%", f"{pl*100:.1f}%",
                     fav, sc[0], sc[1] if len(sc)>1 else "", sc[2] if len(sc)>2 else ""])

    return rows


# ---------------------------------------------------------------------------
# Excel formatting helpers (reaproveitados)
# ---------------------------------------------------------------------------

FILL_PHASE_MAP = {
    "DEZESSEIS-AVOS": PatternFill("solid", fgColor="E2EFDA"),
    "OITAVAS":        PatternFill("solid", fgColor="FFF2CC"),
    "QUARTAS":        PatternFill("solid", fgColor="FCE4D6"),
    "SEMIFINAIS":     PatternFill("solid", fgColor="F4B8C1"),
    "3º LUGAR":       PatternFill("solid", fgColor="E2EFDA"),
    "FINAL":          PatternFill("solid", fgColor="FFD700"),
}
FONT_PHASE_MAP = {
    "DEZESSEIS-AVOS": Font(bold=True, color="375623", size=10),
    "OITAVAS":        Font(bold=True, color="7F6000", size=10),
    "QUARTAS":        Font(bold=True, color="843C0C", size=10),
    "SEMIFINAIS":     Font(bold=True, color="99002B", size=10),
    "3º LUGAR":       Font(bold=True, color="375623", size=10),
    "FINAL":          Font(bold=True, color="7F4800", size=10),
}


def write_simple_sheet(ws, rows, n_cols, col_widths, title_row=1, header_row=2,
                       group_rows=None, phase_rows=None):
    ws.row_dimensions[title_row].height = 20
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row[:n_cols], 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Título
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=n_cols)
    _style_row(ws, title_row, FILL_TITLE, FONT_TITLE, n_cols)

    # Cabeçalho
    _style_row(ws, header_row, FILL_HEADER, FONT_HEADER, n_cols)

    # Linhas de grupo/fase
    for r_idx, row in enumerate(rows, 1):
        if r_idx <= header_row:
            continue
        first = row[0] if row else None
        if group_rows and isinstance(first, str) and any(first.startswith(g) for g in group_rows):
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=n_cols)
            _style_row(ws, r_idx, FILL_GROUP, FONT_GROUP, n_cols)
        elif phase_rows and isinstance(first, str) and first in phase_rows:
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=n_cols)
            fill = FILL_PHASE_MAP.get(first, FILL_PHASE)
            font = FONT_PHASE_MAP.get(first, FONT_PHASE)
            _style_row(ws, r_idx, fill, font, n_cols)
            for c in range(1, n_cols + 1):
                if isinstance(row[c-1] if c-1 < len(row) else None, str) and first in FILL_PHASE_MAP:
                    ws.cell(row=r_idx, column=c).fill = FILL_PHASE_MAP.get(first, FILL_PHASE)

    for col, width in enumerate(col_widths[:n_cols], 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = f"A{header_row+1}"


def write_probabilidades_sheet(ws, rows):
    if not rows:
        return
    n_cols = 10
    col_widths = [4, 22, 7, 10, 10, 10, 8, 8, 10, 10]
    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    _style_row(ws, 1, FILL_TITLE, FONT_TITLE, n_cols)
    _style_row(ws, 2, FILL_HEADER, FONT_HEADER, n_cols)
    FILLS_RANK = [
        PatternFill("solid", fgColor="1A7340"), PatternFill("solid", fgColor="217A43"),
        PatternFill("solid", fgColor="2E8A50"), PatternFill("solid", fgColor="3D9960"),
        PatternFill("solid", fgColor="52A870"), PatternFill("solid", fgColor="6ABF85"),
        PatternFill("solid", fgColor="85D09A"), PatternFill("solid", fgColor="A3E0B0"),
        PatternFill("solid", fgColor="C2EEC7"), PatternFill("solid", fgColor="E2F7E5"),
    ]
    rank_row = 0
    for r_idx, row in enumerate(rows, 1):
        if r_idx <= 2:
            continue
        if isinstance(row[0], int):
            rank_row += 1
            if rank_row <= 10:
                fill = FILLS_RANK[rank_row - 1]
                font = Font(bold=True, color="FFFFFF" if rank_row <= 6 else "000000", size=10)
                for c in range(1, n_cols + 1):
                    ws.cell(row=r_idx, column=c).fill  = fill
                    ws.cell(row=r_idx, column=c).font  = font
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    sep = "=" * 65
    print(f"\n{sep}")
    print("  generate_previsoes_v7.py — Copa 2026 Excel v7")
    print("  Modelo: XGBoost v10 · venue-aware host_factor · SHAP por jogo")
    print(sep)

    print("\n  Carregando modelo e features...")
    model, team_feats = train_model()

    all_teams = [t for ts in GROUPS.values() for t in ts]
    missing = [t for t in all_teams if t not in team_feats]
    if missing:
        for t in missing:
            team_feats[t] = {"shots_on_goal_decay": 3.0, "gols_sofridos_decay": float("nan")}

    print("  Carregando host factors e pré-computando lambdas...")
    hf_home, hf_cohost = load_host_factors()
    lam_full    = precompute_lambdas(model, team_feats, all_teams, hf_home)
    lam_cohost  = precompute_lambdas(model, team_feats, all_teams, hf_cohost)
    lam_neutral = precompute_lambdas(model, team_feats, all_teams, {})

    print("  Carregando ELO...")
    elo_path = Path("data/raw/elo_history.csv")
    elo: dict[str, int] = {}
    if elo_path.exists():
        eh = pd.read_csv(elo_path, parse_dates=["date"])
        for team, val in (eh.sort_values("date").groupby("team")["elo_after"].last().items()):
            elo[team] = int(round(float(val)))
    for team, feats in team_feats.items():
        if team not in elo and "_elo_current" in feats:
            elo[team] = int(round(feats["_elo_current"]))

    print("  Simulação determinística dos grupos e bracket...")
    np.random.seed(42)
    det_standings = simulate_all_groups(lam_full, lam_cohost, lam_neutral, stochastic=False)
    det_ko        = simulate_knockouts(det_standings, lam_full, lam_cohost, lam_neutral, stochastic=False)

    champion = pt(det_ko["winner"][104])
    finalist = pt(det_ko["teams"][104][1] if det_ko["teams"][104][1] != det_ko["winner"][104]
                  else det_ko["teams"][104][0])
    print(f"  Bracket det. → Final: {finalist} vs {champion}  ← Campeão: {champion}")

    print("  Construindo abas...")
    grupos_rows   = build_grupos_rows(lam_full, lam_cohost, lam_neutral, hf_home, hf_cohost, elo)
    matamata_rows = build_matamata_rows(lam_full, lam_cohost, lam_neutral, hf_home, hf_cohost, elo, det_standings, det_ko)
    classif_rows  = build_classificacao_rows(lam_full, lam_cohost, lam_neutral)
    prob_rows     = build_probabilidades_rows(Path("outputs/copa2026_previsoes_v7.csv"))
    r1_rows       = build_rodada1_rows(model, team_feats, hf_home, hf_cohost, elo)

    print(f"  Fase de Grupos  : {len(grupos_rows)} linhas")
    print(f"  Mata-Mata       : {len(matamata_rows)} linhas")
    print(f"  Classificação   : {len(classif_rows)} linhas")
    print(f"  Probabilidades  : {len(prob_rows)} linhas")
    print(f"  Rodada 1        : {len(r1_rows)} linhas")

    print("  Gerando Excel (com SHAP por jogo)...")
    wb = openpyxl.Workbook()

    # Aba principal: Copa 2026 (com SHAP recolhível)
    ws_main = wb.active
    ws_main.title = "Copa 2026"
    build_copa2026_sheet(ws_main, model, team_feats, hf_home, hf_cohost, elo,
                         det_standings, det_ko, lam_full, lam_cohost, lam_neutral)

    # Aba Fase de Grupos
    ws_grp = wb.create_sheet("Fase de Grupos")
    write_simple_sheet(ws_grp, grupos_rows, 15,
                       [5, 22, 6, 7, 3, 7, 6, 22, 7, 7, 7, 20, 16, 16, 16],
                       group_rows=["GRUPO"])

    # Aba Mata-Mata
    ws_ko = wb.create_sheet("Mata-Mata")
    write_simple_sheet(ws_ko, matamata_rows, 15,
                       [12, 22, 6, 7, 3, 7, 6, 22, 7, 7, 7, 20, 16, 16, 16],
                       phase_rows=list(FILL_PHASE_MAP.keys()))

    # Aba Classificação Esperada
    ws_cls = wb.create_sheet("Classificação Esperada")
    write_simple_sheet(ws_cls, classif_rows, 8,
                       [5, 22, 10, 10, 10, 8, 7, 12])

    # Aba Probabilidades MC
    if prob_rows:
        ws_prob = wb.create_sheet("Probabilidades MC")
        write_probabilidades_sheet(ws_prob, prob_rows)

    # Aba Rodada 1
    ws_r1 = wb.create_sheet("Rodada 1")
    write_simple_sheet(ws_r1, r1_rows, 19,
                       [11,7,8,22,6,7,6,3,6,7,6,22,7,7,7,18,14,14,14])

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"\n  Salvo: {OUT_XLSX}")
    print(f"  Dica: na aba 'Copa 2026', clique [ + ] na borda superior para ver features e SHAP")
    print(sep)


if __name__ == "__main__":
    main()
