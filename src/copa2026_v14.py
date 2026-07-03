#!/usr/bin/env python3
"""
copa2026_v14.py — Pipeline atualizado: R32 resultados reais + simulação restante

Mudanças vs v13:
- CUTOFF = 2026-07-03 (inclui 10 jogos R32 já finalizados no treino)
- Resultados reais dos 10 jogos R32 já jogados
- Belgium vs Senegal (correção vs v13 que tinha Belgium vs Algeria)
- Switzerland vs Algeria (real, já no DB)
- Simulação determinística: usa resultado real onde disponível, modelo onde não
- MC 10k a partir dos 6 jogos R32 pendentes (bracket já parcialmente resolvido)

Jogos R32 já disputados (resultados reais):
  South Africa  0-1  Canada        (28/Jun)
  Brazil        2-1  Japan         (29/Jun)  ← Brasil ganhou! vs previsão anterior
  Ivory Coast   1-2  Norway        (30/Jun)
  France        3-0  Sweden        (30/Jun)
  Mexico        2-0  Ecuador       (01/Jul)
  England       2-1  DR Congo      (01/Jul)
  United States 2-0  Bosnia        (02/Jul)
  Spain         3-0  Austria       (02/Jul)
  Portugal      2-1  Croatia       (02/Jul)
  Switzerland   2-0  Algeria       (03/Jul)  ← último jogo salvo

Jogos R32 pendentes (a simular pelo modelo):
  Germany       vs   Paraguay
  Netherlands   vs   Morocco
  Belgium       vs   Senegal
  Argentina     vs   Cape Verde
  Colombia      vs   Ghana
  Australia     vs   Egypt
"""
from __future__ import annotations

import pickle
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from scipy.stats import poisson as scipy_poisson
from xgboost import XGBRegressor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, "src")
warnings.filterwarnings("ignore")
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

RAW_CSV   = Path("data/raw/full_dataset_raw.csv")
MODEL_CSV = Path("data/processed/model_dataset.csv")
ELO_CSV   = Path("data/raw/elo_history.csv")
OUT_PKL   = Path("outputs/xgboost_v14.pkl")
OUT_XLSX  = Path("outputs/copa2026_full_v14.xlsx")

CUTOFF = pd.Timestamp("2026-07-03")
SEED   = 42
N_SIM  = 10_000

CONF_DUMMIES = ["conf_AFC","conf_CAF","conf_CONCACAF","conf_CONMEBOL","conf_OFC","conf_UEFA"]
RATING_FEATURES = [
    "elo_diff_decay","delta_rating_decay","gls_ponderado_decay",
    "gls_decay_vs_forte","gls_decay_vs_fraco","margem_gols_decay",
]
FEATURES = (["opp_saves_decay","shots_on_goal_decay","gols_sofridos_decay"]
            + RATING_FEATURES + CONF_DUMMIES)

XGB_PARAMS = dict(
    objective="count:poisson", n_estimators=300, max_depth=4,
    learning_rate=0.05, subsample=0.8, colsample_bytree=0.8,
    min_child_weight=3, random_state=42, n_jobs=-1, verbosity=0,
)

GROUPS: dict[str, list[str]] = {
    "A": ["Mexico",        "South Korea",  "South Africa",        "Czech Republic"],
    "B": ["Canada",        "Switzerland",  "Qatar",               "Bosnia and Herzegovina"],
    "C": ["Brazil",        "Morocco",      "Haiti",               "Scotland"],
    "D": ["United States", "Paraguay",     "Australia",           "Turkey"],
    "E": ["Germany",       "Ivory Coast",  "Ecuador",             "Curacao"],
    "F": ["Netherlands",   "Sweden",       "Tunisia",             "Japan"],
    "G": ["Belgium",       "Iran",         "New Zealand",         "Egypt"],
    "H": ["Spain",         "Saudi Arabia", "Uruguay",             "Cape Verde"],
    "I": ["France",        "Senegal",      "Iraq",                "Norway"],
    "J": ["Argentina",     "Algeria",      "Austria",             "Jordan"],
    "K": ["Portugal",      "DR Congo",     "Uzbekistan",          "Colombia"],
    "L": ["England",       "Croatia",      "Ghana",               "Panama"],
}
ALL_COPA_TEAMS = [t for ts in GROUPS.values() for t in ts]

COPA_TEAM_CONF: dict[str, str] = {
    "Argentina":"CONMEBOL","Brazil":"CONMEBOL","Colombia":"CONMEBOL",
    "Ecuador":"CONMEBOL","Uruguay":"CONMEBOL","Paraguay":"CONMEBOL",
    "Germany":"UEFA","France":"UEFA","Spain":"UEFA","England":"UEFA",
    "Netherlands":"UEFA","Portugal":"UEFA","Belgium":"UEFA","Austria":"UEFA",
    "Switzerland":"UEFA","Croatia":"UEFA","Norway":"UEFA","Sweden":"UEFA",
    "Czech Republic":"UEFA","Turkey":"UEFA","Scotland":"UEFA",
    "Bosnia and Herzegovina":"UEFA",
    "Japan":"AFC","South Korea":"AFC","Iran":"AFC","Australia":"AFC",
    "Saudi Arabia":"AFC","Uzbekistan":"AFC","Jordan":"AFC","Iraq":"AFC",
    "Qatar":"AFC",
    "Morocco":"CAF","Senegal":"CAF","Egypt":"CAF","Ghana":"CAF",
    "Cape Verde":"CAF","DR Congo":"CAF","Ivory Coast":"CAF",
    "South Africa":"CAF","Algeria":"CAF","Tunisia":"CAF",
    "United States":"CONCACAF","Mexico":"CONCACAF","Canada":"CONCACAF",
    "Panama":"CONCACAF","Curacao":"CONCACAF","Haiti":"CONCACAF",
    "New Zealand":"OFC",
}

CONF_MEDIANS_SAVES = {
    "UEFA":2.9353,"CONMEBOL":2.7985,"AFC":2.4974,
    "CONCACAF":2.4799,"CAF":1.8144,"OFC":2.4974,
}
GLOBAL_MEDIAN_SAVES = 2.6

NAME_MAP = {
    "Czechia":"Czech Republic","Czech Rep.":"Czech Republic",
    "Côte d'Ivoire":"Ivory Coast","Cote d'Ivoire":"Ivory Coast",
    "USA":"United States","IR Iran":"Iran",
    "Korea Republic":"South Korea","Bosnia & Herzegovina":"Bosnia and Herzegovina",
    "Türkiye":"Turkey","Cape Verde Islands":"Cape Verde","Congo DR":"DR Congo",
    "México":"Mexico","Curaçao":"Curacao",
}

# ── R32 COMPLETO: resultados reais + matchups restantes ───────────────────────
# (mid, t1, t2, score_real_g1|None, score_real_g2|None, went_penalties)
# Legenda de status:
#   score != None  → resultado real (API confirmado)
#   score = None   → a disputar (previsão do modelo)
R32_FULL = [
    # ✅ RESULTADOS REAIS (API confirmados)
    (73,  "South Africa",          "Canada",                   0, 1, False),  # 28/Jun
    (76,  "Brazil",                "Japan",                    2, 1, False),  # 29/Jun — Brasil ganhou!
    (77,  "France",                "Sweden",                   3, 0, False),  # 30/Jun
    (78,  "Ivory Coast",           "Norway",                   1, 2, False),  # 30/Jun
    (79,  "Mexico",                "Ecuador",                  2, 0, False),  # 01/Jul
    (80,  "England",               "DR Congo",                 2, 1, False),  # 01/Jul
    (81,  "United States",         "Bosnia and Herzegovina",   2, 0, False),  # 02/Jul
    (83,  "Portugal",              "Croatia",                  2, 1, False),  # 02/Jul
    (84,  "Spain",                 "Austria",                  3, 0, False),  # 02/Jul
    (85,  "Switzerland",           "Algeria",                  2, 0, False),  # 03/Jul — último confirmado
    # 🔮 PREVISÃO MODELO (jogados mas API sem dados / ainda pendentes)
    (74,  "Germany",               "Paraguay",                 None, None, None),
    (75,  "Netherlands",           "Morocco",                  None, None, None),
    (82,  "Belgium",               "Senegal",                  None, None, None),
    (86,  "Argentina",             "Cape Verde",               None, None, None),
    (87,  "Colombia",              "Ghana",                    None, None, None),
    (88,  "Australia",             "Egypt",                    None, None, None),
]

# Vencedores já conhecidos do R32 (resultados reais da API)
KNOWN_R32_WINNERS = {
    73: "Canada",
    76: "Brazil",
    77: "France",
    78: "Norway",
    79: "Mexico",
    80: "England",
    81: "United States",
    83: "Portugal",
    84: "Spain",
    85: "Switzerland",
}

# ── BRACKET COPA 2026 ─────────────────────────────────────────────────────────
# Formato: R32 (16avos) → Oitavas → Quartas → Semifinais → Final + 3° Lugar
#
# R32 → Oitavas (M89-M96):
#   M73-M74 → M89   (Canada vs Germany/Paraguay)
#   M75-M77 → M90   (Netherlands/Morocco vs France)    ← M75 vs M77 (não M76)
#   M76-M78 → M91   (Brazil vs Norway)                 ← M76 vs M78 (não M77)
#   M79-M80 → M92   (Mexico vs England)
#   M81-M82 → M93   (USA vs Belgium/Senegal)
#   M83-M84 → M94   (Portugal vs Spain)
#   M85-M86 → M95   (Switzerland vs Argentina/Cape Verde)
#   M87-M88 → M96   (Colombia/Ghana vs Australia/Egypt)
R16_BRACKET = [   # Oitavas de Final
    (89,  73, 74),
    (90,  75, 77),
    (91,  76, 78),
    (92,  79, 80),
    (93,  81, 82),
    (94,  83, 84),
    (95,  85, 86),
    (96,  87, 88),
]
QF_BRACKET = [    # Quartas de Final
    (97,  89, 90),
    (98,  91, 92),
    (99,  93, 94),
    (100, 95, 96),
]
SF_BRACKET = [    # Semifinais
    (101, 97,  98),
    (102, 99, 100),
]
# Final: M103 = winner(101) vs winner(102)
# 3° Lugar: M104 = loser(101) vs loser(102)

SEP = "=" * 72


# =============================================================================
# 1. Dados e treino
# =============================================================================

def wide_to_long(wide: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for _, r in wide.iterrows():
        for side, opp in [("home","away"),("away","home")]:
            row: dict = {
                "date":                r["date"],
                "team":                r[f"{side}_team"],
                "gols_marcados":       r[f"{side}_gols"],
                "match_type":          r.get("match_type", ""),
                "opp_saves_decay":     float(r.get(f"{opp}_saves_decay", 1.0)),
                "shots_on_goal_decay": float(r[f"{side}_shots_on_goal_decay"]),
                "gols_sofridos_decay": float(r.get(f"{side}_gols_sofridos_decay", np.nan)),
                "saves_decay":         float(r.get(f"{side}_saves_decay", 1.0)),
            }
            for col in CONF_DUMMIES:
                row[col] = int(r.get(col, 0))
            for col in RATING_FEATURES:
                v = r.get(f"{side}_{col}", np.nan)
                row[col] = float(v) if pd.notna(v) else np.nan
            rows.append(row)
    return pd.DataFrame(rows).sort_values("date").reset_index(drop=True)


def apply_saves_norm(wide: pd.DataFrame) -> pd.DataFrame:
    wide = wide.copy()
    for side, opp in [("home","away"),("away","home")]:
        col = f"{side}_saves_decay"
        if col not in wide.columns:
            continue
        opp_col = f"{opp}_team"
        med = wide[opp_col].map(
            lambda t: CONF_MEDIANS_SAVES.get(COPA_TEAM_CONF.get(t, "UEFA"), GLOBAL_MEDIAN_SAVES)
        ).fillna(GLOBAL_MEDIAN_SAVES)
        wide[col] = wide[col].astype(float) / med
    return wide


def winsorize(long: pd.DataFrame) -> pd.DataFrame:
    feats = ["shots_on_goal_decay","gols_sofridos_decay","opp_saves_decay",
             "gls_ponderado_decay","gls_decay_vs_forte","gls_decay_vs_fraco",
             "margem_gols_decay","delta_rating_decay"]
    long = long.copy()
    for feat in feats:
        if feat not in long.columns:
            continue
        col = long[feat].dropna()
        p1, p99 = col.quantile(0.01), col.quantile(0.99)
        long[feat] = long[feat].clip(lower=p1, upper=p99)
    return long


def train_model(long: pd.DataFrame, cutoff: pd.Timestamp) -> XGBRegressor:
    tr = long[long["date"] <= cutoff].copy()
    X  = tr[FEATURES].fillna(tr[FEATURES].median())
    y  = tr["gols_marcados"]
    m  = XGBRegressor(**XGB_PARAMS)
    m.fit(X, y)
    yp   = m.predict(X)
    mae  = np.mean(np.abs(yp - y))
    bias = yp.mean() - y.mean()
    print(f"  n={len(tr):,}  MAE={mae:.4f}  bias={bias:+.4f}")
    wc = tr["date"] >= pd.Timestamp("2026-06-28")
    if wc.sum() > 0:
        print(f"  WC (GS+R32) n={wc.sum()}  MAE={np.mean(np.abs(yp[wc]-y.values[wc])):.4f}"
              f"  bias={yp[wc].mean()-y.values[wc].mean():+.4f}")
    return m


# =============================================================================
# 2. Features e lambdas
# =============================================================================

def extract_feats(long: pd.DataFrame) -> dict[str, dict]:
    tf: dict[str, dict] = {}
    for _, row in long.sort_values("date").iterrows():
        entry: dict = {
            "saves_decay":         float(row.get("saves_decay", 1.0)),
            "shots_on_goal_decay": float(row.get("shots_on_goal_decay", 3.5)),
            "gols_sofridos_decay": row.get("gols_sofridos_decay", np.nan),
        }
        for col in RATING_FEATURES:
            v = row.get(col, np.nan)
            entry[col] = float(v) if pd.notna(v) else np.nan
        tf[row["team"]] = entry
    if ELO_CSV.exists():
        elo = pd.read_csv(ELO_CSV, parse_dates=["date"])
        for team, val in elo.sort_values("date").groupby("team")["elo_after"].last().items():
            tf.setdefault(team, {})["_elo_current"] = float(val)
    return tf


def compute_lambdas(model, tf: dict, teams: list[str]) -> dict[tuple[str,str], float]:
    rows, pairs = [], []
    for t1 in teams:
        for t2 in teams:
            if t1 == t2:
                continue
            f1, f2 = tf.get(t1, {}), tf.get(t2, {})
            row: dict = {
                "opp_saves_decay":     f2.get("saves_decay", 1.0),
                "shots_on_goal_decay": f1.get("shots_on_goal_decay", 3.5),
                "gols_sofridos_decay": f1.get("gols_sofridos_decay", np.nan),
                **{c: 0 for c in CONF_DUMMIES},
            }
            for col in RATING_FEATURES:
                if col == "elo_diff_decay":
                    e1 = f1.get("_elo_current", np.nan)
                    e2 = f2.get("_elo_current", np.nan)
                    row[col] = float(e1)-float(e2) if (pd.notna(e1) and pd.notna(e2)) else f1.get(col, np.nan)
                else:
                    row[col] = f1.get(col, np.nan)
            c = COPA_TEAM_CONF.get(t1, "")
            ck = f"conf_{c}"
            if ck in CONF_DUMMIES:
                row[ck] = 1
            rows.append(row)
            pairs.append((t1, t2))
    preds = model.predict(pd.DataFrame(rows)[FEATURES])
    return {p: float(v) for p, v in zip(pairs, preds)}


# =============================================================================
# 3. Poisson utils
# =============================================================================

def best_score(l1: float, l2: float, max_g: int = 8) -> tuple[int, int, float]:
    bp, bg1, bg2 = -1.0, 0, 0
    for g1 in range(max_g+1):
        for g2 in range(max_g+1):
            p = scipy_poisson.pmf(g1, l1) * scipy_poisson.pmf(g2, l2)
            if p > bp:
                bp, bg1, bg2 = p, g1, g2
    return bg1, bg2, bp


def match_probs(l1: float, l2: float, max_g: int = 8) -> tuple[float, float, float]:
    p1 = pd = p2 = 0.0
    for g1 in range(max_g+1):
        for g2 in range(max_g+1):
            p = scipy_poisson.pmf(g1, l1) * scipy_poisson.pmf(g2, l2)
            if g1 > g2:    p1 += p
            elif g1 == g2: pd += p
            else:          p2 += p
    return p1, pd, p2


def knock_prob(l1: float, l2: float) -> float:
    p1, pd, p2 = match_probs(l1, l2)
    return p1 + 0.5 * pd


def sim_ko(l1: float, l2: float, rng: np.random.Generator) -> int:
    g1, g2 = rng.poisson(l1), rng.poisson(l2)
    if g1 > g2:   return 0
    elif g2 > g1: return 1
    else:         return int(rng.random() > 0.5)


# =============================================================================
# 4. Simulação determinística
# =============================================================================

def simulate_deterministic(lam: dict) -> dict:
    winners: dict[int, str] = {}
    games:   list[dict] = []

    def play(mid: int, t1: str, t2: str, stage: str,
             real_g1=None, real_g2=None, real_pen=False) -> str:
        l1 = lam.get((t1, t2), 1.3)
        l2 = lam.get((t2, t1), 1.3)
        p1, pd_, p2 = match_probs(l1, l2)
        pk = knock_prob(l1, l2)

        if real_g1 is not None:
            # Resultado real
            g1, g2 = real_g1, real_g2
            if g1 > g2:    winner, note, is_real = t1, "", True
            elif g2 > g1:  winner, note, is_real = t2, "", True
            elif real_pen:
                winner = t1 if pk >= 0.5 else t2
                note, is_real = " (pên)", True
            else:          winner, note, is_real = t1, " (pên)", True
        else:
            # Predição do modelo
            g1, g2, _ = best_score(l1, l2)
            is_real = False
            if g1 > g2:    winner, note = t1, ""
            elif g2 > g1:  winner, note = t2, ""
            else:          winner, note = (t1 if l1 >= l2 else t2), " (pên)"

        games.append({
            "stage": stage, "mid": mid,
            "t1": t1, "t2": t2,
            "g1": g1, "g2": g2,
            "score": f"{g1}–{g2}{note}",
            "winner": winner,
            "l1": round(l1, 3), "l2": round(l2, 3),
            "p1": round(p1, 4), "pd": round(pd_, 4), "p2": round(p2, 4),
            "pk": round(pk, 4),
            "is_real": is_real,
        })
        return winner

    # R32 — mistura de resultados reais e previsões
    for mid, t1, t2, rg1, rg2, rpen in R32_FULL:
        winner = play(mid, t1, t2, "R32", rg1, rg2, rpen)
        winners[mid] = winner

    # Oitavas de Final
    for mid, ra, rb in R16_BRACKET:
        t1, t2 = winners[ra], winners[rb]
        winners[mid] = play(mid, t1, t2, "Oitavas")

    # Quartas de Final
    for mid, ra, rb in QF_BRACKET:
        t1, t2 = winners[ra], winners[rb]
        winners[mid] = play(mid, t1, t2, "Quartas")

    # Semifinais
    for mid, ra, rb in SF_BRACKET:
        t1, t2 = winners[ra], winners[rb]
        winners[mid] = play(mid, t1, t2, "Semifinais")

    def loser(mid: int) -> str:
        g = next(gm for gm in games if gm["mid"] == mid)
        return g["t2"] if g["winner"] == g["t1"] else g["t1"]

    # 3° Lugar
    play(103, loser(101), loser(102), "3o Lugar")

    # Final
    play(104, winners[101], winners[102], "Final")

    return {"games": games, "winners": winners}


# =============================================================================
# 5. Monte Carlo
# =============================================================================

def run_mc(lam: dict) -> tuple[dict, dict]:
    rng = np.random.default_rng(SEED)
    champ: dict[str, int] = {}
    radv: dict[str, dict[str, int]] = {
        t: {"R32": 0, "Oitavas": 0, "Quartas": 0, "Semifinais": 0, "Final": 0, "Campeon": 0}
        for t in ALL_COPA_TEAMS
    }

    for _ in range(N_SIM):
        w: dict[int, str] = {}

        # R32 (16avos)
        for mid, t1, t2, rg1, rg2, rpen in R32_FULL:
            if rg1 is not None:
                if rg1 > rg2:    winner = t1
                elif rg2 > rg1:  winner = t2
                else:            winner = t1 if lam.get((t1,t2),1.3) >= lam.get((t2,t1),1.3) else t2
            else:
                l1, l2 = lam.get((t1, t2), 1.3), lam.get((t2, t1), 1.3)
                winner = t1 if sim_ko(l1, l2, rng) == 0 else t2
            w[mid] = winner
            radv[winner]["R32"] += 1

        # Oitavas de Final
        for mid, ra, rb in R16_BRACKET:
            t1, t2 = w[ra], w[rb]
            l1, l2 = lam.get((t1, t2), 1.3), lam.get((t2, t1), 1.3)
            w[mid] = t1 if sim_ko(l1, l2, rng) == 0 else t2
            radv[w[mid]]["Oitavas"] += 1

        # Quartas de Final
        for mid, ra, rb in QF_BRACKET:
            t1, t2 = w[ra], w[rb]
            l1, l2 = lam.get((t1, t2), 1.3), lam.get((t2, t1), 1.3)
            w[mid] = t1 if sim_ko(l1, l2, rng) == 0 else t2
            radv[w[mid]]["Quartas"] += 1

        # Semifinais
        for mid, ra, rb in SF_BRACKET:
            t1, t2 = w[ra], w[rb]
            l1, l2 = lam.get((t1, t2), 1.3), lam.get((t2, t1), 1.3)
            w[mid] = t1 if sim_ko(l1, l2, rng) == 0 else t2
            radv[w[mid]]["Semifinais"] += 1

        # Final
        ft1, ft2 = w[101], w[102]
        l1, l2 = lam.get((ft1, ft2), 1.3), lam.get((ft2, ft1), 1.3)
        radv[ft1]["Final"] += 1; radv[ft2]["Final"] += 1
        campeon = ft1 if sim_ko(l1, l2, rng) == 0 else ft2
        radv[campeon]["Campeon"] += 1
        champ[campeon] = champ.get(campeon, 0) + 1

    return champ, radv


# =============================================================================
# 6. Standings reais
# =============================================================================

def compute_standings(raw: pd.DataFrame) -> dict[str, pd.DataFrame]:
    copa = raw[(raw["match_type"] == "WC") & (raw["home_away"] == "home")].copy()
    copa["date"] = pd.to_datetime(copa["date"])

    team_to_group = {t: g for g, ts in GROUPS.items() for t in ts}
    s: dict[str, dict[str, dict]] = {
        g: {t: {"pts":0,"gf":0,"ga":0,"gd":0,"w":0,"d":0,"l":0,"pgj":0}
            for t in ts}
        for g, ts in GROUPS.items()
    }

    for _, row in copa.iterrows():
        home = NAME_MAP.get(row["team"], row["team"])
        away = NAME_MAP.get(row["opponent"], row["opponent"])
        gh = pd.to_numeric(row["gols_marcados"], errors="coerce")
        ga = pd.to_numeric(row["gols_sofridos"], errors="coerce")
        if pd.isna(gh) or pd.isna(ga):
            continue
        gh, ga = int(gh), int(ga)
        grp = team_to_group.get(home)
        if grp is None or team_to_group.get(away) != grp:
            continue
        for team, gf, gc in [(home, gh, ga), (away, ga, gh)]:
            if team not in s[grp]:
                continue
            d = s[grp][team]
            d["gf"] += gf; d["ga"] += gc; d["gd"] = d["gf"] - d["ga"]; d["pgj"] += 1
            if gf > gc:    d["pts"] += 3; d["w"] += 1
            elif gf == gc: d["pts"] += 1; d["d"] += 1
            else:          d["l"] += 1

    result: dict[str, pd.DataFrame] = {}
    for g, td in s.items():
        df = pd.DataFrame([{"team": t, **v} for t, v in td.items()])
        df = df.sort_values(["pts","gd","gf"], ascending=[False,False,False]).reset_index(drop=True)
        df["pos"] = df.index + 1
        result[g] = df
    return result


# =============================================================================
# 7. Excel
# =============================================================================

def _hdr(ws, r, c, val, bg="1A5276", fg="FFFFFF", bold=True, size=11):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _cell(ws, r, c, val, bg=None, bold=False, fmt=None, align="center"):
    cell = ws.cell(row=r, column=c, value=val)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt:
        cell.number_format = fmt
    s = Side(border_style="thin", color="CCCCCC")
    cell.border = Border(left=s, right=s, top=s, bottom=s)
    return cell


STAGE_BG = {
    "R32": "D6EAF8", "Oitavas": "D5F5E3", "Quartas": "FDEBD0",
    "Semifinais": "FDEDEC", "Final": "FFD700", "3o Lugar": "E8DAEF",
}


def build_excel(out: Path, det, mc_champ, radv, lam):
    wb = openpyxl.Workbook()

    # ── Aba 1: Bracket Eliminatório ───────────────────────────────────────────
    ws1 = wb.active; ws1.title = "Bracket"; ws1.sheet_view.showGridLines = False

    # Legenda
    _hdr(ws1, 1, 1, "Copa 2026 — Bracket Eliminatório  |  ✅ Resultado Real   🔮 Previsão Modelo",
         bg="0B2545")
    ws1.merge_cells("A1:J1")
    ws1.row_dimensions[1].height = 22

    row = 2
    # Separamos R32 em: reais já no DB e pendentes
    r32_real    = [gm for gm in det["games"] if gm["stage"] == "R32" and gm["is_real"]]
    r32_pending = [gm for gm in det["games"] if gm["stage"] == "R32" and not gm["is_real"]]

    stages_sequence = [
        ("16avos — Resultados já confirmados",  r32_real,    "1A5276", "D6EAF8"),
        ("16avos — Jogos pendentes (previsão)", r32_pending, "117A65", "D5F5E3"),
        ("Oitavas de Final",   [gm for gm in det["games"] if gm["stage"] == "Oitavas"],    "117A65", "D5F5E3"),
        ("Quartas de Final",   [gm for gm in det["games"] if gm["stage"] == "Quartas"],    "784212", "FDEBD0"),
        ("Semifinais",         [gm for gm in det["games"] if gm["stage"] == "Semifinais"], "922B21", "FDEDEC"),
        ("Final",              [gm for gm in det["games"] if gm["stage"] == "Final"],      "7B0000", "FFD700"),
        ("3° Lugar",           [gm for gm in det["games"] if gm["stage"] == "3o Lugar"],   "4A235A", "E8DAEF"),
    ]

    hdrs_row = ["Tipo", "#", "Time 1", "λ1", "Placar", "λ2", "Time 2", "P(T1 adv)", "P(Empate 90')", "Vencedor"]
    col_w = [14, 4, 26, 7, 10, 7, 26, 10, 13, 26]

    for section_label, games, hdr_bg, body_bg in stages_sequence:
        if not games:
            continue
        # Section header
        _hdr(ws1, row, 1, section_label, bg=hdr_bg)
        ws1.merge_cells(f"A{row}:J{row}")
        ws1.row_dimensions[row].height = 18
        row += 1
        # Column headers
        for c, h in enumerate(hdrs_row, 1):
            _hdr(ws1, row, c, h, bg="2C3E50", size=10)
        row += 1

        for gm in games:
            is_real = gm["is_real"]
            tipo    = "✅ Real" if is_real else "🔮 Previsão"
            bg      = "D5F5E3" if is_real else body_bg
            winner  = gm["winner"]

            vals  = [tipo, gm["mid"],
                     gm["t1"], gm["l1"], gm["score"], gm["l2"], gm["t2"],
                     gm["pk"], gm["pd"], winner]
            fmts  = [None, None, None, "0.00", None, "0.00", None, "0.0%", "0.0%", None]
            bolds = [False, False,
                     winner == gm["t1"], False, True, False, winner == gm["t2"],
                     False, False, True]

            for c, (v, fmt, bold) in enumerate(zip(vals, fmts, bolds), 1):
                bg_c = "D4EFDF" if c == 10 else ("F0FFF4" if (is_real and c != 1) else bg)
                _cell(ws1, row, c, v, bg=bg_c, bold=bold, fmt=fmt,
                      align="left" if c in (3, 7, 10) else "center")
            row += 1
        row += 1  # blank line between sections

    for i, w in enumerate(col_w, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Aba 2: MC Probabilidades (apenas times ainda em campo ou eliminados recentemente) ──
    ws2 = wb.create_sheet("MC_Probabilidades"); ws2.sheet_view.showGridLines = False
    _hdr(ws2, 1, 1, f"Monte Carlo {N_SIM:,} simulações — P(avançar) por fase (apenas times do R32)",
         bg="154360")
    ws2.merge_cells("A1:I1")
    hdrs2 = ["Conf", "Time", "Status", "P(Oitavas)", "P(Quartas)", "P(Semi)", "P(Final)", "P(Campeão)", "λ_atq"]
    for c, h in enumerate(hdrs2, 1):
        _hdr(ws2, 2, c, h, bg="1F618D")
    row = 3

    # Somente times que participaram do R32
    r32_teams = set()
    for mid, t1, t2, *_ in R32_FULL:
        r32_teams.add(t1); r32_teams.add(t2)

    for t in sorted(r32_teams, key=lambda t: radv.get(t, {}).get("Campeon", 0), reverse=True):
        conf  = COPA_TEAM_CONF.get(t, "?")
        ra    = radv.get(t, {})
        pcamp = ra.get("Campeon", 0) / N_SIM

        eliminated = any(
            mid in KNOWN_R32_WINNERS and KNOWN_R32_WINNERS[mid] != t and t in (t1, t2)
            for mid, t1, t2, *_ in R32_FULL
        )
        pending = any(t in (t1, t2) and g1 is None for mid, t1, t2, g1, *_ in R32_FULL)

        if eliminated:
            status = "❌ Eliminado"
            bg = "F5F5F5"
        elif pending:
            status = "⏳ Jogo pendente"
            bg = "FFF9C4"
        else:
            status = "🟢 Em campo"
            bg = ("FFDDD5" if pcamp>=0.10 else ("FFF3CD" if pcamp>=0.05 else
                  ("D5F5E3" if pcamp>=0.02 else ("EBF5FB" if row%2==0 else "FDFEFE"))))

        lam_a = np.mean([lam.get((t, o), 1.3) for o in ALL_COPA_TEAMS if o != t])
        vals = [conf, t, status,
                ra.get("Oitavas", 0) / N_SIM,
                ra.get("Quartas", 0) / N_SIM,
                ra.get("Semifinais", 0) / N_SIM,
                ra.get("Final", 0) / N_SIM,
                pcamp,
                round(lam_a, 3)]
        fmts = [None, None, None, "0.0%", "0.0%", "0.0%", "0.0%", "0.0%", "0.000"]
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            _cell(ws2, row, c, v, bg=bg, bold=(pcamp >= 0.05), fmt=fmt,
                  align="left" if c in (2, 3) else "center")
        row += 1

    col_w2 = [10, 26, 16, 11, 11, 10, 10, 12, 9]
    for i, w in enumerate(col_w2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Aba 3: MC Campeão ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("MC_Campeon"); ws3.sheet_view.showGridLines = False
    _hdr(ws3, 1, 1, f"MC {N_SIM:,} sim — Probabilidades de Campeão Copa 2026", bg="154360")
    ws3.merge_cells("A1:E1")
    for c, h in enumerate(["Rank", "Conf", "Time", "Simulações", "P(Campeão)"], 1):
        _hdr(ws3, 2, c, h, bg="1F618D")
    row = 3
    fills = ["FFD700", "C0C0C0", "CD7F32"]
    total = sum(mc_champ.values())
    for rank, (team, cnt) in enumerate(
        sorted(mc_champ.items(), key=lambda x: x[1], reverse=True), 1
    ):
        pct  = cnt / total
        conf = COPA_TEAM_CONF.get(team, "?")
        bg   = (fills[rank-1] if rank <= 3 else
                ("FFDDD5" if pct >= 0.10 else ("FFF3CD" if pct >= 0.05 else
                ("D5F5E3" if pct >= 0.02 else ("EBF5FB" if row % 2 == 0 else "FDFEFE")))))
        for c, (v, fmt) in enumerate(
            zip([rank, conf, team, cnt, pct], [None, None, None, "0", "0.0%"]), 1
        ):
            _cell(ws3, row, c, v, bg=bg, bold=(rank <= 3), fmt=fmt,
                  align="left" if c == 3 else "center")
        row += 1
    for i, w in enumerate([6, 10, 26, 12, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(out)
    print(f"  Excel salvo: {out}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    print(f"\n{SEP}")
    print("  COPA 2026 v14 — R32 resultados reais + simulação restante")
    print(f"{SEP}\n")

    # 1. Rebuild model_dataset com CUTOFF=2026-07-03
    print("1. Reconstruindo model_dataset.csv (CUTOFF=2026-07-03)...")
    from build_features import (load_and_merge_raw, add_decay_features,
                                 impute_by_confederation, build_match_dataset,
                                 XI_DEFAULT, save_data)
    df_raw  = load_and_merge_raw()
    df_dec  = add_decay_features(df_raw, XI_DEFAULT)
    df_imp  = impute_by_confederation(df_dec)
    df_wide = build_match_dataset(df_imp)
    save_data(df_wide)
    wide    = pd.read_csv(MODEL_CSV, parse_dates=["date"])
    wide    = apply_saves_norm(wide)
    print(f"   model_dataset: {len(wide):,} jogos wide")
    wc_wide = wide[wide["match_type"] == "WC"]
    print(f"   WC no wide: {len(wc_wide)} jogos ({len(wc_wide[(wc_wide['date'] >= pd.Timestamp('2026-06-28'))])} inclui KO)")

    # 2. Long + winsorize + treino
    print("\n2. Treinando xgboost_v14...")
    long    = wide_to_long(wide)
    long_w  = winsorize(long)
    model   = train_model(long_w, CUTOFF)
    with open(OUT_PKL, "wb") as f:
        pickle.dump(model, f)
    print(f"   Salvo: {OUT_PKL}")

    # 3. Lambdas
    print("\n3. Computando lambdas...")
    tf  = extract_feats(long_w)
    lam = compute_lambdas(model, tf, ALL_COPA_TEAMS)
    top5_atk = sorted(
        [(t, np.mean([lam.get((t,o),1.3) for o in ALL_COPA_TEAMS if o!=t]))
         for t in ALL_COPA_TEAMS], key=lambda x: x[1], reverse=True
    )[:5]
    print(f"   Top-5 λ_atk: {[(t, round(v,3)) for t,v in top5_atk]}")

    # 4. Bracket determinístico
    print("\n4. Simulação determinística R32→Final...")
    det = simulate_deterministic(lam)

    for stage_label, stage_key in [
        ("R32", "R32"), ("Oitavas", "Oitavas"), ("Quartas", "Quartas"),
        ("Semifinais", "Semifinais"), ("Final", "Final"), ("3o Lugar", "3o Lugar"),
    ]:
        gms = [g for g in det["games"] if g["stage"] == stage_key]
        if gms:
            print(f"\n   {stage_label}:")
            for gm in gms:
                tag = "✅" if gm["is_real"] else "🔮"
                print(f"   {tag} M{gm['mid']:>3}  {gm['t1']:<25} {gm['score']:>12}  "
                      f"{gm['t2']:<25}  → {gm['winner']}")

    # 5. MC
    print(f"\n5. Monte Carlo {N_SIM:,} simulações...")
    mc_champ, radv = run_mc(lam)
    total = sum(mc_champ.values())
    print("   Top-10 campeões:")
    for rank, (team, cnt) in enumerate(
        sorted(mc_champ.items(), key=lambda x: x[1], reverse=True)[:10], 1
    ):
        print(f"   {rank:>2}. {team:<25} {cnt/total:5.1%}  (n={cnt})")

    # 6. Excel
    print("\n6. Gerando Excel...")
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    build_excel(OUT_XLSX, det, mc_champ, radv, lam)

    print(f"\n{SEP}")
    print(f"  CONCLUÍDO")
    print(f"  Modelo:  {OUT_PKL}")
    print(f"  Excel:   {OUT_XLSX}")
    print(f"{SEP}\n")


if __name__ == "__main__":
    main()
